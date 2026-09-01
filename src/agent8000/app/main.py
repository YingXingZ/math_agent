from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from html import escape
from io import BytesIO, StringIO
from PIL import Image
from pathlib import Path
from typing import Literal
import json
import re
import asyncio
import os
import csv
import sqlite3
import httpx
import time
import uuid
from urllib.parse import quote, urlencode


# --- LaTeX helpers for HTML rendering ----------------------------------------
# 8014 content_text often contains raw LaTeX without $ delimiters (e.g.
# \begin{cases}..., \frac{...}{...}).  Wrap them so MathJax can render them.
_LATEX_INLINE_RE = re.compile(
    r'\\(?!(?:begin|end)\{)[a-zA-Z]+.*?(?=[\u4e00-\u9fa5，。；：？！、;!]|\\begin|\$\$|$)',
    re.DOTALL,
)
_LATEX_ENV_RE = re.compile(r'\\begin\{[^}]+\}.*?\\end\{[^}]+\}', re.DOTALL)
_PRINT_PAGE_HEADER_RE = re.compile(r'(?m)^\s*\d{3}\s*第[一二三四五六七八九十]+章[^\n]*$')


def _clean_assignment_math(text: str) -> str:
    """Remove OCR-created empty/unmatched display delimiters for worksheet display only."""
    text = (text or '').replace('\r', '')
    # Some OCR imports leave a bare $$ after an otherwise inline formula.
    # A delimiter on a line by itself cannot render useful student content here.
    text = re.sub(r'(?m)^\s*\$\$\s*$', '', text)
    # If OCR left one final delimiter after removing blank delimiter lines,
    # discard only that unmatched tail; valid $$...$$ pairs remain intact.
    if text.count('$$') % 2:
        text = text.rsplit('$$', 1)[0]
    return re.sub(r'\n{3,}', '\n\n', text).strip()


def _strip_selected_subpart_prefix(text: str, original_no: str) -> str:
    """The worksheet header already carries （n）; avoid printing it twice."""
    match = re.search(r'[（(]\s*(\d+)\s*[)）]\s*$', str(original_no or ''))
    if not match:
        return text or ''
    marker = match.group(1)
    return re.sub(r'^\s*[（(]\s*' + re.escape(marker) + r'\s*[)）]\s*', '', text or '', count=1)


def _wrap_latex_for_html(text: str) -> str:
    """Add $ / $$ delimiters around raw LaTeX fragments without double-wrapping."""
    # Answer-book imports occasionally retain a printed page header.  It is
    # useful provenance in the source record but must not appear in a student
    # worksheet.  This is display-only; the stored question is untouched.
    text = _clean_assignment_math(_PRINT_PAGE_HEADER_RE.sub('', text or ''))
    placeholders: list[str] = []

    def _stash(m: re.Match) -> str:
        placeholders.append(m.group(0))
        return f'\x00MATH{len(placeholders) - 1}\x00'

    # 1. Preserve every supported existing delimiter.  In particular, do this
    # before searching for raw commands: otherwise ``\\( \\frac{1}{2} \\)``
    # becomes the invalid nested form ``\\( $\\frac{1}{2}$ \\)``.
    text = re.sub(r'\\\(.*?\\\)', _stash, text, flags=re.DOTALL)
    text = re.sub(r'\\\[.*?\\\]', _stash, text, flags=re.DOTALL)
    text = re.sub(r'\$\$.*?\$\$', _stash, text, flags=re.DOTALL)
    text = re.sub(r'(?<!\$)\$(?!\$)[^\$]*?\$(?!\$)', _stash, text, flags=re.DOTALL)

    # 2. Convert \begin{...} ... \end{...} environments to display math.
    def _wrap_env(m: re.Match) -> str:
        wrapped = f'\n$$\n{m.group(0)}\n$$\n'
        placeholders.append(wrapped)
        return f'\x00MATH{len(placeholders) - 1}\x00'

    text = _LATEX_ENV_RE.sub(_wrap_env, text)

    # 3. Wrap remaining inline LaTeX commands (but not \begin / \end).
    def _wrap_inline(m: re.Match) -> str:
        s = m.group(0).strip()
        return f'${s}$' if s else m.group(0)

    text = _LATEX_INLINE_RE.sub(_wrap_inline, text)

    # 4. Restore stashed math.
    for i, ph in enumerate(placeholders):
        text = text.replace(f'\x00MATH{i}\x00', ph)
    return text

from fastapi import BackgroundTasks, FastAPI, File, Form, HTTPException, Request, Response, UploadFile
from fastapi.responses import HTMLResponse, FileResponse, RedirectResponse
from pydantic import BaseModel, Field

from .config import settings
from .evidence_client import client as evidence_client, url as evidence_url
from .db import connection, init_db, queue_due_grading as enqueue_due_grading, normalize_question_type
from .auth import audit, current_user, ensure_bootstrap_admin, login, require_roles, teacher_id_for_scope, token_hash
from .dify import run_workflow
from .knowledge_bridge import build_image_solve_candidate, evidence_status, list_evidence_sections, rescue_formula_from_crop, retrieve_section_problems
from .grading_pipeline import run_grading_job
from .queueing import QueueUnavailable, _rq_objects, dispatch_grading_job, enqueue_pending_grading_jobs, queue_health, rq_enabled
from .mineru_staging import match_staged, stage_markdown
from .answer_matcher import answer_json_to_staged, canonical_section, match_section_questions
from .question_validation import validate_question, first_issue_message
from .assignment_pdf import (POINTS_PER_QUESTION, build_assignment_pdf,
                             display_problem_no, export_latex_source,
                             latex_document, strip_source_problem_prefix)
from .orchestrator import publish_homework
from .mineru_review import (
    approve_item as approve_mineru_item,
    create_session as create_mineru_session,
    get_session as get_mineru_session,
    list_audit_log as list_mineru_audit_log,
    list_items as list_mineru_items,
    list_pending_sync as list_mineru_pending_sync,
    mark_section_synced,
    reject_item as reject_mineru_item,
)
from .ai_stem_review import (
    approve_candidate as approve_ai_stem_candidate,
    list_candidates as list_ai_stem_candidates,
    reject_candidate as reject_ai_stem_candidate,
    store_pending_candidate,
)
from .question_bank_review import _looks_corrupt, _scan_looks_corrupt


@asynccontextmanager
async def lifespan(_: FastAPI):
    settings.prepare_dirs()
    init_db()
    ensure_bootstrap_admin()
    yield


app = FastAPI(title="高数作业助手", version="0.1.0", lifespan=lifespan)


class QuestionIn(BaseModel):
    content: str = Field(min_length=3)
    chapter: str
    difficulty: Literal["基础", "提高", "综合"] = "基础"
    question_type: str = "计算题"
    answer: str = ""
    rubric: str = ""
    source_page: int | None = None


class QuestionUpdateIn(BaseModel):
    """Manual edit of an existing question.  All fields optional (PATCH-style);
    only the provided fields are written.  Intentionally NOT gated by
    validate_question — a teacher correcting OCR garble or raw LaTeX must be able
    to save text the auto-validator would otherwise flag."""

    content: str | None = None
    chapter: str | None = None
    difficulty: Literal["基础", "提高", "综合"] | None = None
    question_type: str | None = None
    answer: str | None = None
    rubric: str | None = None
    knowledge_points: str | None = None
    review_status: str | None = None
    sync_8014: bool = True  # when True, also push the edited fields back to the 8014 evidence DB


class AssignmentIn(BaseModel):
    title: str
    chapter: str
    class_id: int = Field(gt=0)
    due_at: datetime
    semester: str = ""
    question_count: int = Field(ge=1, le=30, default=6)
    basic_ratio: float = Field(ge=0, le=1, default=.5)
    advanced_ratio: float = Field(ge=0, le=1, default=.35)
    subpart_limit: int = Field(ge=3, le=6, default=3)


class QuestionReviewDecisionIn(BaseModel):
    sort_order: int = Field(ge=1)
    score: float = Field(ge=0)
    feedback: str = Field(default="", max_length=2000)


class ReviewDecisionIn(BaseModel):
    # score keeps the legacy whole-submission confirmation path compatible.
    # New teacher UI sends question_decisions and server derives the total itself.
    score: float | None = Field(default=None, ge=0)
    feedback: str = Field(default="", max_length=4000)
    question_decisions: list[QuestionReviewDecisionIn] = Field(default_factory=list, max_length=100)


class BatchReviewIn(BaseModel):
    submission_ids: list[int] = Field(min_length=1, max_length=100)


class DemoCleanupIn(BaseModel):
    confirm: bool = False


class SubmissionRegionIn(BaseModel):
    question_id: int = Field(gt=0)
    subpart_no: str = Field(default="", max_length=30)
    sort_order: int = Field(ge=1)
    page_no: int = Field(ge=1)
    x: float = Field(default=0.0, ge=0.0, le=1.0)
    y: float = Field(default=0.0, ge=0.0, le=1.0)
    width: float = Field(default=1.0, gt=0.0, le=1.0)
    height: float = Field(default=1.0, gt=0.0, le=1.0)


class SubmissionRegionBatchIn(BaseModel):
    mappings: list[SubmissionRegionIn] = Field(min_length=1, max_length=100)


class AssignmentUpdateIn(BaseModel):
    title: str = Field(min_length=1)
    due_at: datetime


class AssignmentDraftItemIn(BaseModel):
    question_id: int = Field(gt=0)
    subpart_no: str = Field(default='', max_length=30)
    content: str = Field(min_length=1, max_length=20000)
    score: float = Field(gt=0, le=100)


class ClassIn(BaseModel):
    name: str = Field(min_length=1, max_length=80)
    semester: str = Field(default="", max_length=40)


class StudentListIn(BaseModel):
    students: list[dict]


class LoginIn(BaseModel):
    username: str = Field(min_length=1, max_length=80)
    password: str = Field(min_length=1, max_length=200)


class TeacherCreateIn(BaseModel):
    username: str = Field(min_length=3, max_length=80)
    display_name: str = Field(min_length=1, max_length=80)
    temporary_password: str = Field(min_length=10, max_length=200)


class InviteCreateIn(BaseModel):
    max_uses: int = Field(default=1, ge=1, le=500)
    expires_days: int = Field(default=14, ge=1, le=90)


class StudentActivateIn(BaseModel):
    invite_code: str = Field(min_length=16, max_length=160)
    student_no: str = Field(min_length=1, max_length=32)
    name: str = Field(min_length=1, max_length=80)
    username: str = Field(min_length=3, max_length=80)
    password: str = Field(min_length=10, max_length=200)


DEMO_CLASS_NAME = "演示班级（不计入正式报表）"
DEMO_SEMESTER = "DEMO"
DEMO_STUDENT_NO = "DEMO001"
DEMO_STUDENT_NAME = "演示学生"


def _create_demo_package(conn: sqlite3.Connection, actor: dict) -> dict:
    """Create a reusable, isolated grade-flow demo without publishing a question."""
    teacher_id = actor["id"] if actor["role"] == "teacher" else None
    class_row = conn.execute("SELECT * FROM classes WHERE name=? AND semester=?", (DEMO_CLASS_NAME, DEMO_SEMESTER)).fetchone()
    if class_row is None:
        class_id = conn.execute("INSERT INTO classes(name,semester,teacher_user_id) VALUES(?,?,?)", (DEMO_CLASS_NAME, DEMO_SEMESTER, teacher_id)).lastrowid
    else:
        class_id = class_row["id"]
    conn.execute("INSERT OR IGNORE INTO students(class_id,student_no,name) VALUES(?,?,?)", (class_id, DEMO_STUDENT_NO, DEMO_STUDENT_NAME))
    question = conn.execute("SELECT id FROM questions WHERE content=? AND review_status='blocked' LIMIT 1", ("【演示测试题】计算 1+0 的值。",)).fetchone()
    if question is None:
        question_id = conn.execute(
            """INSERT INTO questions(content,chapter,difficulty,question_type,answer,rubric,review_status,owner_teacher_id)
               VALUES(?,?,?,?,?,?,?,?)""",
            ("【演示测试题】计算 1+0 的值。", "DEMO", "基础", "计算题", "1", "写出结果 1，得 10 分。", "blocked", teacher_id),
        ).lastrowid
    else:
        question_id = question["id"]
    assignment = conn.execute("SELECT id FROM assignments WHERE title='演示：自动批改流程' AND is_demo=1 LIMIT 1").fetchone()
    if assignment is None:
        assignment_id = conn.execute(
            """INSERT INTO assignments(title,chapter,class_name,class_id,due_at,total_score,status,semester,is_demo)
               VALUES(?,?,?,?,?,?,?,?,1)""",
            ("演示：自动批改流程", "DEMO", DEMO_CLASS_NAME, class_id, "2099-12-31T23:59:59+00:00", 10, "published", DEMO_SEMESTER),
        ).lastrowid
        conn.execute("INSERT INTO assignment_questions(assignment_id,question_id,sort_order,score,original_no) VALUES(?,?,?,?,?)", (assignment_id, question_id, 1, 10, "1"))
    else:
        assignment_id = assignment["id"]
    return {"class_id": class_id, "assignment_id": assignment_id, "student_no": DEMO_STUDENT_NO,
            "student_name": DEMO_STUDENT_NAME, "question": "计算 1+0 的值", "expected_answer": "1"}


@app.post("/api/admin/demo-package", status_code=201)
def create_demo_package(request: Request):
    """Create a disposable-looking demo package. It is explicitly excluded from reports."""
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        package = _create_demo_package(conn, actor)
    audit(actor, "demo.create", "demo_package", package["assignment_id"], actor["id"] if actor["role"] == "teacher" else None)
    return {**package, "sample_download_url": "/api/admin/demo-package/handwriting-sample",
            "submit_url": f"/submit?assignment_id={package['assignment_id']}",
            "note": "演示数据仅用于测试，不进入正式教学报表。"}


@app.get("/api/admin/demo-package/handwriting-sample")
def download_demo_handwriting_sample(request: Request):
    require_roles(request, {"admin", "teacher"})
    path = Path(__file__).with_name("demo_handwriting_sample.png")
    if not path.is_file():
        raise HTTPException(404, "演示手写样本文件不存在")
    return FileResponse(path, filename="高数智能体_手写作业演示样本.png", media_type="image/png")


class MineruStageIn(BaseModel):
    role: Literal["textbook", "answer_book"]
    name: str = Field(min_length=1, max_length=120)
    markdown: str = Field(min_length=10)


class MineruMatchIn(BaseModel):
    textbook: dict
    answer_book: dict


class MineruReviewStartIn(BaseModel):
    answer_book: dict


class AnswerMatchIn(BaseModel):
    answer_parse_result: dict
    threshold: float = Field(default=0.90, gt=0, le=1)
    create_review_on_block: bool = True


class MineruReviewDecisionIn(BaseModel):
    action: Literal["approve", "reject"]
    std_answer: str = ""
    full_solution: str = ""
    note: str = ""
    overwrite_verified: bool = False


@app.post("/api/auth/login")
def auth_login(payload: LoginIn, response: Response, request: Request):
    user, token, expires_at = login(payload.username, payload.password)
    response.set_cookie(settings.session_cookie_name, token, httponly=True,
                        secure=settings.cookie_secure, samesite="lax", expires=expires_at)
    audit(user, "login", "session", tenant_teacher_id=(user["id"] if user["role"] == "teacher" else None),
          ip=request.client.host if request.client else None)
    return {"user": user, "auth_required": settings.auth_required}


@app.post("/api/auth/logout")
def auth_logout(request: Request, response: Response):
    user = current_user(request)
    token = request.cookies.get(settings.session_cookie_name, "")
    if token:
        with connection() as conn:
            conn.execute("UPDATE user_sessions SET revoked_at=? WHERE token_hash=?", (datetime.now(timezone.utc).isoformat(), token_hash(token)))
    response.delete_cookie(settings.session_cookie_name)
    audit(user, "logout", "session")
    return {"ok": True}


@app.get("/api/auth/me")
def auth_me(request: Request):
    user = current_user(request)
    return {"user": user, "auth_required": settings.auth_required}


@app.get("/api/admin/status")
def admin_status(request: Request):
    require_roles(request, {"admin"})
    with connection() as conn:
        return {
            "auth_required": settings.auth_required,
            "users": conn.execute("SELECT COUNT(*) FROM users WHERE active=1").fetchone()[0],
            "teachers": conn.execute("SELECT COUNT(*) FROM users WHERE active=1 AND role='teacher'").fetchone()[0],
            "classes": conn.execute("SELECT COUNT(*) FROM classes WHERE teacher_user_id IS NOT NULL").fetchone()[0],
            "pending_reviews": conn.execute("SELECT COUNT(*) FROM submissions WHERE needs_review=1").fetchone()[0],
            "retention": {"submission_days": settings.submission_retention_days, "audit_days": settings.audit_retention_days},
        }


@app.get("/api/admin/teachers")
def admin_list_teachers(request: Request):
    require_roles(request, {"admin"})
    with connection() as conn:
        rows = conn.execute("""SELECT u.id,u.username,u.display_name,u.active,u.created_at,COUNT(c.id) AS class_count
                               FROM users u LEFT JOIN classes c ON c.teacher_user_id=u.id
                               WHERE u.role='teacher' GROUP BY u.id ORDER BY u.created_at DESC""").fetchall()
    return [dict(row) for row in rows]


@app.post("/api/admin/teachers", status_code=201)
def admin_create_teacher(payload: TeacherCreateIn, request: Request):
    admin = require_roles(request, {"admin"})
    from .auth import hash_password
    with connection() as conn:
        try:
            cur = conn.execute("INSERT INTO users(username,display_name,password_hash,role) VALUES(?,?,?,'teacher')",
                               (payload.username.strip(), payload.display_name.strip(), hash_password(payload.temporary_password)))
        except sqlite3.IntegrityError as exc:
            raise HTTPException(409, "该登录名已存在") from exc
    audit(admin, "teacher.create", "user", cur.lastrowid, metadata={"username": payload.username.strip()})
    return {"id": cur.lastrowid, "username": payload.username.strip(), "display_name": payload.display_name.strip(), "role": "teacher"}


@app.get("/api/admin/audit")
def admin_audit(request: Request, limit: int = 100):
    require_roles(request, {"admin"})
    limit = min(max(limit, 1), 500)
    with connection() as conn:
        rows = conn.execute("""SELECT l.*,u.username,u.display_name FROM audit_logs l
                               LEFT JOIN users u ON u.id=l.actor_user_id
                               ORDER BY l.id DESC LIMIT ?""", (limit,)).fetchall()
    return [dict(row) for row in rows]


@app.get("/api/admin/roster-template")
def admin_roster_template(request: Request):
    require_roles(request, {"admin"})
    return Response("学号,姓名\n20260001,示例同学\n", media_type="text/csv; charset=utf-8",
                    headers={"Content-Disposition": "attachment; filename=roster_template.csv"})


def _difficulty_label(value: object) -> str:
    """Normalise 8014's numeric/text difficulty to the local paper levels."""
    try:
        number = int(value)  # 8014 currently uses 1 / 2 / 3.
        return "基础" if number <= 1 else "提高" if number == 2 else "综合"
    except (TypeError, ValueError):
        text = str(value or "")
        return text if text in {"基础", "提高", "综合"} else "提高"


def _parse_sections(value: str) -> list[str]:
    """Accept teacher-friendly multi-section input: 1.1, 1.2 / 1.3."""
    sections: list[str] = []
    for section in re.split(r"[，,、;；\s]+", value or ""):
        section = section.strip()
        if section and section not in sections:
            sections.append(section)
    if not sections:
        raise HTTPException(422, "请至少选择一个章节，例如 1.1、1.2")
    return sections


def _problem_text_issue(item: dict) -> str | None:
    """Reject OCR fragments before they can enter an assignment.

    Older 8014 imports may contain a long string which is nevertheless not a
    readable question (for example dozens of one-character OCR lines).  Length
    alone is therefore deliberately not a publication criterion.
    """
    content = (item.get("content_text") or "").strip()
    if len(content) < 10:
        return "题干缺失"
    lines = [line.strip() for line in content.splitlines() if line.strip()]
    one_char_lines = sum(len(line) <= 1 for line in lines)
    cjk_count = len(re.findall(r"[\u4e00-\u9fff]", content))
    replacement_count = content.count("�")
    # Mathematical notation is allowed, but a question still needs natural
    # language context and must not be a vertical stream of OCR fragments.
    if replacement_count:
        return "题干含无法解码字符"
    if len(lines) >= 7 and one_char_lines / len(lines) >= 0.45:
        return "题干是公式碎片，需按原题图重识"
    if cjk_count < 3 and len(content) > 20:
        return "题干缺少可读文字，需按原题图重识"
    # These are not legitimate mathematical tokens.  They are the recurring
    # artefacts produced when the previous OCR confused radicals, infinity,
    # brackets and fractions with CJK/Latin glyphs.  Do not try to repair them
    # with regex: require the original image and a vision re-read instead.
    garbled_formula = re.compile(r"[←ζ]|(?:[一二三四五六七八九]oo)|(?:[JHVUR]{2,})|(?:[VJH][0-9])|(?:[一丨川]（)|例\s*\d")
    if garbled_formula.search(content):
        return "题干含数学 OCR 污染，需按原题图重识"
    return None


def _usable_problem(item: dict) -> bool:
    """Do not issue a paper containing a missing or garbled OCR result."""
    return _problem_text_issue(item) is None


async def _validate_or_rescue(item: dict) -> tuple[dict, bool]:
    """Apply the unified gate. Returns ``(usable_item_or_None, rescues_attempted)``.

    ``item`` may be mutated in place when an image rescue produces a usable text.

    Admission rule: a problem is usable for the local working cache when it is
    *readable and structurally sound* (``report["valid"]`` — no block/error
    severity issues).  OCR-sourced text rarely reaches ``decision == "pass"``
    because its source confidence (0.78) caps the score below ``PASS_THRESHOLD``
    (0.90); a ``review`` decision carrying only warnings is still perfectly
    usable here because the 8014 source answer is already teacher-verified and
    the generated paper is previewed by the teacher before distribution
    (设计一.1).  Items that are ``block``ed (unreadable) or carry structural
    errors are *not* admitted; they fall through to the Qwen vision rescue or
    remain unresolved for the teacher to fix at the source.
    """
    crop = (item.get("evidence") or {}).get("crop_image_path")
    report = validate_question(
        item.get("content_text") or "",
        source_type="ocr",
        source_confidence=0.78,
        crop_image_path=crop,
    )
    if report["valid"]:
        item = dict(item)
        item["evidence"] = {
            **item.get("evidence", {}),
            "validation": {
                "decision": report["decision"],
                "score": report["score"],
                "source_confidence": report["source_confidence"],
            },
        }
        return item, False
    if report["needs_formula_rescue"] and crop and not item.get("_rescued"):
        rescue = await rescue_formula_from_crop(crop)
        rescued_text = (rescue.get("candidate_text") or "").strip()
        if rescued_text:
            recheck = validate_question(
                rescued_text,
                source_type="pix2text",
                source_confidence=float(rescue.get("confidence", 0.45) or 0.45),
                crop_image_path=crop,
            )
            if recheck["valid"]:
                item = dict(item)
                item["content_text"] = rescued_text
                item["_rescued"] = True
                item["evidence"] = {
                    **item["evidence"],
                    "formula_rescue": rescue,
                    "validation": {"decision": recheck["decision"], "score": recheck["score"]},
                }
                return item, True
    return None, False


async def _sync_section_into_local_cache(section_no: str, limit: int = 80) -> dict:
    """Import a traceable working copy of one 8014 section without changing 8014."""
    packet = await retrieve_section_problems(section_no, limit)
    source_items = packet["items"]
    usable_items: list[dict] = []
    ai_candidate_count = 0
    pending_review_count = 0
    pending_review_items: list[dict] = []
    unresolved_count = 0
    unresolved_items: list[dict] = []
    # Per-item processing used to be a serial `for` loop; with VLM/Pix2Text in
    # the rescue path each problem can stall 5-10s, so a chapter of 80 items
    # easily blew past the 90s hard cap even after we made `publish_homework`
    # run its sections concurrently.  Run item handling through a small
    # semaphore so we fan out without flooding 8014.
    PER_ITEM_SEM = asyncio.Semaphore(6)
    PER_ITEM_TIMEOUT = 18  # seconds per problem (VLM 15 + headroom)

    async def _process_one(item: dict) -> str:
        """Return 'usable' | 'ai' | 'pending' | 'unresolved' to classify."""
        async with PER_ITEM_SEM:
            try:
                return await asyncio.wait_for(_handle_one_item(item), PER_ITEM_TIMEOUT)
            except asyncio.TimeoutError:
                return "timeout"
            except Exception:
                return "unresolved"

    async def _handle_one_item(item: dict) -> str:
        text_issue = _problem_text_issue(item)
        if not text_issue and (item.get("std_answer") or "").strip():
            usable, _ = await _validate_or_rescue(item)
            if usable is not None:
                usable_items.append(usable)
                return "usable"
        candidate = await build_image_solve_candidate(item)
        if candidate.get("status") == "eligible":
            item2 = dict(item)
            item2["content_text"] = candidate["problem_text"]
            item2["ptype"] = candidate["ptype"]
            item2["std_answer"] = candidate["std_answer"]
            item2["full_solution"] = candidate["full_solution"]
            item2["answer_status"] = "ai_candidate"
            item2["evidence"] = {**item2["evidence"], "ai_candidate": candidate}
            usable_items.append(item2)
            nonlocal ai_candidate_count; ai_candidate_count += 1
            return "ai"
        if candidate.get("status") == "pending_review":
            stored = store_pending_candidate(item, candidate)
            pending_review_count += 1
            pending_review_items.append({
                "candidate_id": stored.get("id"),
                "source_problem_id": str(item.get("source_problem_id") or ""),
                "problem_no": str(item.get("problem_no") or "?"),
                "difficulty": _difficulty_label(item.get("difficulty")),
                "ptype": str(item.get("ptype") or ""),
                "candidate_text": (candidate.get("problem_text") or "").strip(),
                "std_answer": str(candidate.get("std_answer") or ""),
                "confidence": candidate.get("confidence"),
                "reason": candidate.get("reason", ""),
                "has_source_image": bool(crop := (item.get("evidence") or {}).get("crop_image_path")),
                "next_action": "教师核对原题图后确认写回",
            })
            return "pending"
        text_issue_final = _problem_text_issue(item)
        unresolved_count += 1
        unresolved_items.append({
            "source_problem_id": str(item.get("source_problem_id") or ""),
            "problem_no": str(item.get("problem_no") or "?"),
            "ptype": str(item.get("ptype") or ""),
            "reason": text_issue_final or "缺少可用标准答案",
            "has_source_image": bool(crop := (item.get("evidence") or {}).get("crop_image_path")),
            "next_action": "系统将按原题图重识" if crop else "8014 缺少原题裁切图，需补绑定原题证据",
        })
        return "unresolved"

    if source_items:
        await asyncio.gather(*[_process_one(it) for it in source_items])
    if not usable_items:
        raise HTTPException(
            422,
            "该章节暂无可直接发布的可读题目；请先在 8014 修复题目截图或 OCR，避免向学生发出乱码题目。",
        )

    synced_ids: list[int] = []
    available_by_level = {"基础": 0, "提高": 0, "综合": 0}
    with connection() as conn:
        # Old local cache rows are never allowed to remain selectable after a
        # newer quality gate has found that their source text is unreadable.
        usable_source_ids = {str(item["source_problem_id"]) for item in usable_items}
        for item in source_items:
            source_id = str(item.get("source_problem_id") or "")
            if source_id and source_id not in usable_source_ids:
                conn.execute("UPDATE questions SET review_status='blocked' WHERE source_problem_id=?", (source_id,))
        for item in usable_items:
            source_id = str(item["source_problem_id"])
            evidence = json.dumps(item["evidence"], ensure_ascii=False)
            difficulty = _difficulty_label(item.get("difficulty"))
            available_by_level[difficulty] += 1
            values = (
                item["content_text"].strip(),
                section_no,
                difficulty,
                str(item.get("ptype") or "计算题"),
                str(item.get("std_answer") or ""),
                str(item.get("grading_steps") or item.get("full_solution") or ""),
                evidence,
                str(item.get("problem_no") or ""),
            )
            existing = conn.execute(
                "SELECT id FROM questions WHERE source_problem_id=?", (source_id,)
            ).fetchone()
            if existing:
                conn.execute(
                    """UPDATE questions SET content=?, chapter=?, difficulty=?, question_type=?,
                       answer=?, rubric=?, source_evidence_json=?, source_problem_no=?, review_status='published'
                       WHERE id=?""",
                    (*values, existing["id"]),
                )
                synced_ids.append(existing["id"])
            else:
                cursor = conn.execute(
                    """INSERT INTO questions
                       (content,chapter,difficulty,question_type,answer,rubric,source_evidence_json,source_problem_id,source_problem_no)
                       VALUES(?,?,?,?,?,?,?,?,?)""",
                    (*values[:7], source_id, values[7]),
                )
                synced_ids.append(cursor.lastrowid)
    return {
        "section_no": section_no,
        "source_total": packet["total"],
        "synced_count": len(synced_ids),
        "skipped_unreadable": unresolved_count,
        "ai_candidate_count": ai_candidate_count,
        "pending_review_count": pending_review_count,
        "pending_review_items": pending_review_items,
        "unresolved_items": unresolved_items,
        "available_by_level": available_by_level,
        "question_ids": synced_ids,
    }


@app.get("/api/questions")
def list_questions(request: Request, chapter: str | None = None):
    actor = require_roles(request, {"admin", "teacher"})
    sql, args = "SELECT * FROM questions WHERE review_status='published'", []
    if chapter:
        sql += " AND chapter=?"; args.append(chapter)
    scope = teacher_id_for_scope(actor)
    if scope is not None:
        sql += " AND (owner_teacher_id IS NULL OR owner_teacher_id=?)"; args.append(scope)
    with connection() as conn:
        return [dict(r) for r in conn.execute(sql + " ORDER BY id DESC", args)]


@app.post("/api/mineru/stage")
def stage_mineru_document(payload: MineruStageIn):
    """Preview MinerU chapter/question extraction without database writes."""
    return stage_markdown(payload.markdown, payload.role, payload.name)


@app.post("/api/mineru/match")
def match_mineru_documents(payload: MineruMatchIn):
    """Match staged textbook and answer-book JSON; confirmation is a separate step."""
    if payload.textbook.get("document", {}).get("role") != "textbook":
        raise HTTPException(422, "textbook 必须是教材暂存 JSON")
    if payload.answer_book.get("document", {}).get("role") != "answer_book":
        raise HTTPException(422, "answer_book 必须是答案书暂存 JSON")
    return match_staged(payload.textbook, payload.answer_book)


@app.post("/api/mineru/answer-match")
async def match_parsed_answers(payload: AnswerMatchIn):
    """Match server_answer_parse JSON to 8014 and enforce the 90% publish gate."""
    try:
        answer_book = answer_json_to_staged(payload.answer_parse_result)
        section = canonical_section(payload.answer_parse_result.get("meta", {}).get("target_section"))
        evidence = await retrieve_section_problems(section, limit=100)
        result = match_section_questions(section, evidence.get("items", []), answer_book, payload.threshold)
        result["answer_book"] = answer_book
        if not result["publish_gate"]["can_publish"] and payload.create_review_on_block:
            result["review_session"] = await create_mineru_session(answer_book)
        return result
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc
    except httpx.TimeoutException as exc:
        raise HTTPException(504, "8014 题目查询超时") from exc
    except httpx.HTTPStatusError as exc:
        raise HTTPException(502, "8014 题目查询失败：" + exc.response.text[:180]) from exc


@app.post("/api/mineru/review-sessions", status_code=201)
async def create_mineru_review_session(payload: MineruReviewStartIn):
    """Create local-only review records from a staged answer-book JSON."""
    try:
        return await create_mineru_session(payload.answer_book)
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc


@app.get("/api/mineru/review-sessions/{session_id}")
def get_mineru_review_session(session_id: int):
    """Return session metadata plus a live summary of approved/rejected/pending counts."""
    try:
        return get_mineru_session(session_id)
    except LookupError as exc:
        raise HTTPException(404, str(exc)) from exc


@app.get("/api/mineru/review-sessions/{session_id}/items")
def get_mineru_review_items(session_id: int):
    return list_mineru_items(session_id)


@app.post("/api/mineru/review-items/{item_id}")
async def decide_mineru_review_item(item_id: int, payload: MineruReviewDecisionIn):
    """Approve writes the verified answer back to 8014; reject records the decision locally.

    Approval follows an all-or-nothing contract: 8014 must accept the PUT before the
    local review item is marked approved, so a failed upstream write never leaves the
    local state inconsistent.
    """
    try:
        if payload.action == "reject":
            return reject_mineru_item(item_id, payload.note)
        return await approve_mineru_item(
            item_id,
            payload.std_answer,
            payload.full_solution,
            payload.note,
            payload.overwrite_verified,
        )
    except LookupError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc
    except RuntimeError as exc:
        # 8014 timeout / write failure: local state was not changed, safe to retry.
        raise HTTPException(502, str(exc)) from exc
    except httpx.HTTPStatusError as exc:
        raise HTTPException(exc.response.status_code, "8014 写回失败：" + exc.response.text[:180]) from exc


@app.get("/api/mineru/review-sessions/{session_id}/audit")
def get_mineru_review_audit(session_id: int):
    """Audit trail for a review session (approve/reject actions and before/after values)."""
    return {"session_id": session_id, "entries": list_mineru_audit_log(session_id=session_id)}


@app.post("/api/mineru/review-sessions/{session_id}/sync")
async def sync_mineru_review_session(session_id: int):
    """Sync every section that has at least one approved item back into the local cache.

    This closes the upstream loop: answers verified through MinerU are now visible to
    the assignment/grading pipeline without a manual re-sync.
    """
    with connection() as conn:
        rows = conn.execute(
            """SELECT DISTINCT section_no FROM mineru_review_items
               WHERE session_id=? AND status='approved' AND section_no IS NOT NULL""",
            (session_id,),
        ).fetchall()
    sections = [row["section_no"] for row in rows]
    if not sections:
        raise HTTPException(422, "当前批次没有已确认项，无需同步")

    results = []
    errors = []
    async def _sync_one(section: str):
        import time as _t
        _t0 = _t.time()
        print(f"[publish_sync] START section={section}", flush=True)
        try:
            result = await _sync_section_into_local_cache(section, limit=80)
            mark_section_synced(section)
            print(f"[publish_sync] DONE section={section} in {_t.time()-_t0:.1f}s", flush=True)
            return (section, result, None)
        except HTTPException as exc:
            print(f"[publish_sync] FAIL section={section} in {_t.time()-_t0:.1f}s", flush=True)
            return (section, None, exc.detail)
        except Exception as exc:  # noqa: BLE001
            print(f"[publish_sync] FAIL section={section} in {_t.time()-_t0:.1f}s: {str(exc)[:120]}", flush=True)
            return (section, None, str(exc)[:200])
    pairs = await asyncio.gather(*[_sync_one(s) for s in sections])
    for section, result, err in pairs:
        if err is not None:
            errors.append({"section_no": section, "error": err})
        else:
            results.append({"section_no": section, "sync": result})

    return {
        "session_id": session_id,
        "sections_synced": len(results),
        "sections": [r["section_no"] for r in results],
        "results": results,
        "errors": errors,
    }


@app.get("/api/mineru/pending-sync")
def get_mineru_pending_sync():
    """Sections that have been written back to 8014 and should be re-synced soon."""
    return {"items": list_mineru_pending_sync()}


# ---------------------------------------------------------------------------
# VLM-recognised stems awaiting teacher review (Route 1: OCR-completion loop)
# ---------------------------------------------------------------------------
class AiStemDecisionIn(BaseModel):
    content_text: str = ""
    std_answer: str = ""
    full_solution: str = ""
    note: str = ""


@app.get("/api/agent/ai-stem-candidates")
def list_ai_stem_candidates_endpoint(status: str | None = "pending", section_no: str | None = None):
    """Teacher queue of VLM-recognised stems that need confirmation before publication."""
    return {"items": list_ai_stem_candidates(status=status, section_no=section_no)}


@app.post("/api/agent/ai-stem-candidates/{candidate_id}/approve")
async def approve_ai_stem_endpoint(candidate_id: int, payload: AiStemDecisionIn):
    """Confirm a recognised stem: write content_text back to 8014 + local cache."""
    try:
        return await approve_ai_stem_candidate(
            candidate_id, payload.content_text, payload.std_answer, payload.full_solution, payload.note
        )
    except LookupError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(502, str(exc)) from exc


@app.post("/api/agent/ai-stem-candidates/{candidate_id}/reject")
def reject_ai_stem_endpoint(candidate_id: int, payload: AiStemDecisionIn):
    """Reject a candidate; it is recorded locally but never published."""
    try:
        return reject_ai_stem_candidate(candidate_id, payload.note)
    except LookupError as exc:
        raise HTTPException(404, str(exc)) from exc


@app.post("/api/agent/ai-stem-candidates/approve-section/{section_no}")
async def approve_section_ai_stems(section_no: str):
    """Batch-approve every pending stem candidate for a section (uses VLM text as-is)."""
    candidates = list_ai_stem_candidates(status="pending", section_no=section_no)
    results: list[dict] = []
    errors: list[dict] = []
    for candidate in candidates:
        try:
            results.append(await approve_ai_stem_candidate(candidate["id"]))
        except Exception as exc:  # noqa: BLE001
            errors.append({"candidate_id": candidate["id"], "error": str(exc)[:200]})
    return {"approved": len(results), "errors": errors, "results": results}


@app.post("/api/mineru/pending-sync/{section_no}")
def mark_mineru_section_synced(section_no: str):
    """Mark a section as synced (used after a manual /api/agent/sync-section)."""
    return mark_section_synced(section_no)


@app.get("/answer-import-review", response_class=HTMLResponse)
def mineru_review_page():
    return FileResponse(Path(__file__).with_name("answer_import_review.html"))


@app.post("/api/questions", status_code=201)
def create_question(question: QuestionIn, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    report = validate_question(
        question.content,
        source_type="manual",
        source_confidence=None,
        crop_image_path=None,
    )
    if not report["publish_allowed"]:
        raise HTTPException(
            422,
            "题目未通过数学校验，不能发布：" + (first_issue_message(report) or "存在结构问题"),
        )
    with connection() as conn:
        cursor = conn.execute("""INSERT INTO questions(content,chapter,difficulty,question_type,answer,rubric,source_page,review_status,owner_teacher_id)
          VALUES(?,?,?,?,?,?,?,?,?)""", (
            question.content, question.chapter, question.difficulty,
            question.question_type, question.answer, question.rubric,
            question.source_page, "published", actor["id"] if actor["role"] == "teacher" else None,
        ))
        audit(actor, "question.create", "question", cursor.lastrowid, actor["id"] if actor["role"] == "teacher" else None)
        return {"id": cursor.lastrowid, "message": "题目已入库，待教师审核后可参与组卷。"}


@app.get("/api/questions/{question_id}")
def get_question(question_id: int, request: Request):
    """Return a single question by id (any review_status, not just published)."""
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        row = conn.execute("SELECT * FROM questions WHERE id=?", (question_id,)).fetchone()
    if not row:
        raise HTTPException(404, f"题目 {question_id} 不存在")
    scope = teacher_id_for_scope(actor)
    if scope is not None and row["owner_teacher_id"] not in (None, scope):
        raise HTTPException(404, f"题目 {question_id} 不存在")
    return dict(row)


@app.put("/api/questions/{question_id}")
async def update_question(question_id: int, payload: QuestionUpdateIn, request: Request):
    """Manual edit of an existing question.  Provided fields are written; missing
    fields are left untouched.  Saving is NOT blocked by the math validator — we
    only surface a non-fatal warning when the new content looks suspicious."""
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        row = conn.execute("SELECT * FROM questions WHERE id=?", (question_id,)).fetchone()
        if not row:
            raise HTTPException(404, f"题目 {question_id} 不存在")
        scope = teacher_id_for_scope(actor)
        if scope is not None and row["owner_teacher_id"] not in (None, scope):
            raise HTTPException(404, f"题目 {question_id} 不存在")
        if scope is not None and row["owner_teacher_id"] is None:
            raise HTTPException(403, "官方题库为只读；请复制为个人题目后再修改，避免影响其他教师")
        fields = {}
        if payload.content is not None:
            fields["content"] = payload.content
        if payload.chapter is not None:
            fields["chapter"] = payload.chapter
        if payload.difficulty is not None:
            fields["difficulty"] = payload.difficulty
        if payload.question_type is not None:
            fields["question_type"] = payload.question_type
        if payload.answer is not None:
            fields["answer"] = payload.answer
        if payload.rubric is not None:
            fields["rubric"] = payload.rubric
        if payload.knowledge_points is not None:
            fields["knowledge_points"] = payload.knowledge_points
        if payload.review_status is not None:
            fields["review_status"] = payload.review_status
        if not fields:
            return {"id": question_id, "message": "无字段变更", "question": dict(row)}
        set_sql = ", ".join(f"{k}=?" for k in fields)
        conn.execute(
            f"UPDATE questions SET {set_sql} WHERE id=?",
            list(fields.values()) + [question_id],
        )
        new_row = conn.execute("SELECT * FROM questions WHERE id=?", (question_id,)).fetchone()
    audit(actor, "question.update", "question", question_id, row["owner_teacher_id"])
    # Best-effort push the edited fields back to the 8014 evidence DB so the
    # source of truth stays consistent with the manual correction.
    sync_status = {"synced": False, "skipped": "未请求同步"}
    if payload.sync_8014:
        sync_status = await _sync_to_8014(row["source_problem_id"], fields)
    # Non-fatal validation hint so the teacher knows if the text looks off.
    validation_warning = None
    try:
        rep = validate_question(
            new_row["content"], source_type="manual",
            source_confidence=None, crop_image_path=None,
        )
        if not rep["publish_allowed"]:
            validation_warning = first_issue_message(rep)
    except Exception:  # noqa: BLE001 - validator must never block the save
        validation_warning = None
    return {
        "id": question_id,
        "message": "题目已更新",
        "question": dict(new_row),
        "validation_warning": validation_warning,
        "sync_8014_status": sync_status,
    }


# Teacher-facing OCR-garble backlog.  The CSV is produced by the offline audit
# script (scan_garble_precise / fix_fullwidth_and_audit); this endpoint simply
# surfaces it so the edit panel can walk the backlog in priority order.
# Deployments can override the repository default with GARBLE_AUDIT_CSV.
# Degrades gracefully if absent.
_GARBLE_AUDIT_CSV = Path(settings.garble_audit_csv)
_LEGACY_OCR_MARKERS = re.compile(r"[锟�叫咱呗]")

# 8014 is an internal evidence service.  Never open its SQLite database from
# 8001: all cross-service writes use the authenticated client below.
async def _sync_to_8014(source_problem_id: object, fields: dict) -> dict:
    """Best-effort evidence-service sync for teacher edits; never blocks local save."""
    if not source_problem_id:
        return {"synced": False, "skipped": "本题无 source_problem_id，8014 中无对应记录"}
    try:
        async with evidence_client(timeout=25) as client:
            if "content" in fields:
                response = await client.put(evidence_url(f"/problems/{source_problem_id}/content"), json={"content_text": fields["content"]})
                response.raise_for_status()
            answer_payload = {target: fields[source] for source, target in {"answer": "std_answer", "rubric": "full_solution"}.items() if source in fields}
            if answer_payload:
                response = await client.put(evidence_url(f"/problems/{source_problem_id}/answer"), json=answer_payload)
                response.raise_for_status()
        return {"synced": True, "problem_id": str(source_problem_id)}
    except httpx.TimeoutException:
        return {"synced": False, "reason": "8014 内部服务响应超时"}
    except httpx.HTTPStatusError as exc:
        return {"synced": False, "reason": f"8014 写回失败：HTTP {exc.response.status_code}"}
    except Exception as exc:  # noqa: BLE001 - local edit is already persisted
        return {"synced": False, "reason": f"8014 写回异常：{str(exc)[:160]}"}


@app.get("/api/evidence/images/{image_path:path}")
async def get_evidence_image(image_path: str, request: Request):
    """Authenticated 8001 proxy for private 8014 source images."""
    require_roles(request, {"admin", "teacher"})
    if not image_path or image_path.startswith(("/", "\\")) or ".." in image_path.split("/"):
        raise HTTPException(400, "无效的图片路径")
    try:
        async with evidence_client(timeout=30) as client:
            response = await client.get(evidence_url("/images/" + quote(image_path, safe="/")))
            response.raise_for_status()
        return Response(content=response.content, media_type=response.headers.get("content-type", "image/jpeg"))
    except httpx.HTTPStatusError as exc:
        raise HTTPException(exc.response.status_code, "证据图片不存在或不可访问") from exc
    except httpx.HTTPError as exc:
        raise HTTPException(502, "证据图片服务暂不可用") from exc


def _live_garble_reasons(question: dict) -> tuple[str, list[str]]:
    """Classify current local text without writing or changing publication.

    The historic CSV is only a snapshot. This deliberately conservative scan
    combines the existing full-width/ASCII-salad rules with recurring OCR glyph
    substitutions, so corrected questions drop out while suspicious text stays
    visible for a teacher to check against the source image.
    """
    reasons: list[str] = []
    high_risk = False
    for label, key in (("题干", "content"), ("答案", "answer"), ("评分参考", "rubric")):
        text = str(question.get(key) or "").strip()
        if not text:
            # A rubric is optional.  Its absence must not make an otherwise
            # usable question appear to be corrupted.
            if key != "rubric":
                reasons.append(f"{label}为空")
                high_risk = True
            continue
        if _looks_corrupt(text):
            reasons.append(f"{label}含全角/替换字符")
            high_risk = True
        if len(_LEGACY_OCR_MARKERS.findall(text)) >= 3:
            reasons.append(f"{label}含重复 OCR 代字")
            high_risk = True
        elif _scan_looks_corrupt(text):
            reasons.append(f"{label}含可疑公式或字母序列")
    return ("high" if high_risk else "medium"), reasons


def _live_garble_queue(review_status: str | None = None) -> list[dict]:
    with connection() as conn:
        sql = "SELECT id, content, answer, rubric, review_status FROM questions"
        args: list[str] = []
        if review_status:
            sql += " WHERE review_status=?"
            args.append(review_status)
        rows = [dict(row) for row in conn.execute(sql, args).fetchall()]
    items = []
    for row in rows:
        risk, reasons = _live_garble_reasons(row)
        if not reasons:
            continue
        items.append({
            "question_id": row["id"],
            "review_status": row["review_status"],
            "risk": risk,
            "garble_hint": "；".join(reasons),
            "preview": (row["content"] or "")[:80],
        })
    return sorted(items, key=lambda item: (item["risk"] != "high", item["question_id"]))


@app.get("/api/garble-queue")
def garble_queue(review_status: str | None = None, source: Literal["audit", "live"] = "audit"):
    """Return the OCR-audit queue, optionally using live local review status.

    The CSV is an immutable audit trail, so its status column can be stale after
    a teacher fixes an item. Filtering therefore consults ``questions`` rather
    than trusting the CSV snapshot.
    """
    if source == "live":
        items = _live_garble_queue(review_status)
        high = sum(item["risk"] == "high" for item in items)
        return {
            "items": items,
            "note": f"实时扫描：{len(items)} 道疑似问题（高风险 {high} 道）；仅供复核，不会自动修改状态",
        }
    if not _GARBLE_AUDIT_CSV.exists():
        return {"items": [], "note": "未找到 garble_audit.csv（离线审计脚本未运行）"}
    items = []
    with _GARBLE_AUDIT_CSV.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            try:
                qid = int(row.get("question_id") or row.get("id") or "")
            except ValueError:
                continue
            items.append({
                "question_id": qid,
                "assignments_using": int(row.get("assignments_using") or 0),
                "review_status": row.get("review_status") or "",
                "garble_hint": (row.get("garble_hint") or "")[:60],
                "preview": (row.get("content_preview") or row.get("preview") or "")[:80],
            })
    if review_status:
        ids = [item["question_id"] for item in items]
        if ids:
            placeholders = ",".join("?" for _ in ids)
            with connection() as conn:
                rows = conn.execute(
                    f"SELECT id FROM questions WHERE review_status=? AND id IN ({placeholders})",
                    [review_status, *ids],
                ).fetchall()
            allowed_ids = {row["id"] for row in rows}
            items = [item for item in items if item["question_id"] in allowed_ids]
        else:
            items = []
    # priority: most-used assignments first
    items.sort(key=lambda x: -x["assignments_using"])
    label = f"状态为 {review_status} 的" if review_status else ""
    return {"items": items, "note": f"共 {len(items)} 道{label}待修订题目"}


def _require_class(conn: sqlite3.Connection, class_id: int, actor: dict | None = None) -> sqlite3.Row:
    row = conn.execute("SELECT id,name,semester,teacher_user_id FROM classes WHERE id=?", (class_id,)).fetchone()
    if not row:
        raise HTTPException(404, "班级不存在；请先在“班级与名单”中创建班级")
    if actor:
        scope = teacher_id_for_scope(actor)
        if scope is not None and row["teacher_user_id"] != scope:
            # Do not reveal whether another teacher's class exists.
            raise HTTPException(404, "班级不存在")
    return row


def _require_assignment(conn: sqlite3.Connection, assignment_id: int, actor: dict | None = None) -> sqlite3.Row:
    row = conn.execute("""SELECT a.*,c.teacher_user_id FROM assignments a
                        LEFT JOIN classes c ON c.id=a.class_id WHERE a.id=?""", (assignment_id,)).fetchone()
    if not row:
        raise HTTPException(404, "作业不存在")
    if actor:
        if actor["role"] == "student":
            enrolled = conn.execute("SELECT 1 FROM students WHERE class_id=? AND user_id=?", (row["class_id"], actor["id"])).fetchone()
            if not enrolled:
                raise HTTPException(404, "作业不存在")
            return row
        scope = teacher_id_for_scope(actor)
        if scope is not None and row["teacher_user_id"] != scope:
            raise HTTPException(404, "作业不存在")
    return row


def _require_roster(conn: sqlite3.Connection, class_id: int) -> None:
    if not conn.execute("SELECT 1 FROM students WHERE class_id=? LIMIT 1", (class_id,)).fetchone():
        raise HTTPException(422, "该班尚未导入学生名单，不能发布作业")


def _normalise_students(students: list[dict]) -> tuple[list[tuple[str, str]], int]:
    """Validate a roster without silently guessing identity columns."""
    clean: list[tuple[str, str]] = []
    invalid = 0
    seen: set[str] = set()
    for row in students:
        no = str(row.get("student_no", "")).strip()
        name = str(row.get("name", "")).strip()
        if not no or not name or not _STUDENT_NO_RE.fullmatch(no) or no in seen:
            invalid += 1
            continue
        seen.add(no)
        clean.append((no, name))
    return clean, invalid


def _parse_roster_file(filename: str, raw: bytes) -> list[dict]:
    """Read an explicit 学号/姓名 roster from CSV or XLSX; never infer names."""
    suffix = Path(filename or "").suffix.lower()
    rows: list[list[object]] = []
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - deployment dependency
            raise HTTPException(503, "服务器未安装 Excel 导入依赖 openpyxl，请改用 UTF-8 CSV") from exc
        try:
            book = load_workbook(BytesIO(raw), read_only=True, data_only=True)
            sheet = book.active
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        except Exception as exc:
            raise HTTPException(422, "Excel 文件无法读取；请确认是 .xlsx 格式") from exc
    elif suffix == ".csv":
        text = None
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise HTTPException(422, "CSV 编码无法读取，请另存为 UTF-8 CSV")
        rows = list(csv.reader(text.splitlines()))
    else:
        raise HTTPException(422, "仅支持 .xlsx 或 .csv 名单文件")
    if not rows:
        raise HTTPException(422, "名单文件为空")
    headers = [str(v or "").strip().lower().replace(" ", "") for v in rows[0]]
    no_names = {"学号", "student_no", "studentno", "studentnumber"}
    name_names = {"姓名", "name", "student_name", "studentname"}
    try:
        no_index = next(i for i, h in enumerate(headers) if h in no_names)
        name_index = next(i for i, h in enumerate(headers) if h in name_names)
    except StopIteration as exc:
        raise HTTPException(422, "首行必须包含“学号”和“姓名”两列；其他列可保留") from exc
    return [
        {"student_no": str(row[no_index] or "").strip(), "name": str(row[name_index] or "").strip()}
        for row in rows[1:]
        if len(row) > max(no_index, name_index)
    ]


def _save_roster(class_id: int, students: list[dict]) -> dict:
    clean, invalid = _normalise_students(students)
    with connection() as conn:
        _require_class(conn, class_id)
        inserted = updated = 0
        for student_no, name in clean:
            existing = conn.execute(
                "SELECT id,name FROM students WHERE class_id=? AND student_no=?", (class_id, student_no)
            ).fetchone()
            if existing:
                if existing["name"] != name:
                    conn.execute("UPDATE students SET name=? WHERE id=?", (name, existing["id"]))
                    updated += 1
                continue
            conn.execute(
                "INSERT INTO students(class_id,student_no,name) VALUES(?,?,?)", (class_id, student_no, name)
            )
            inserted += 1
    return {"inserted": inserted, "updated": updated, "invalid_or_duplicate": invalid,
            "accepted": len(clean)}


@app.get("/api/classes")
def list_classes(request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    scope = teacher_id_for_scope(actor)
    with connection() as conn:
        sql = (
            """SELECT c.id,c.name,c.semester,c.created_at,COUNT(DISTINCT st.id) AS student_count,
                      COUNT(DISTINCT a.id) AS assignment_count
               FROM classes c
               LEFT JOIN students st ON st.class_id=c.id
               LEFT JOIN assignments a ON a.class_id=c.id
            """
        )
        args: list[object] = []
        if scope is not None:
            sql += " WHERE c.teacher_user_id=?"; args.append(scope)
        rows = conn.execute(sql + " GROUP BY c.id ORDER BY c.semester DESC,c.name", args).fetchall()
    return [dict(row) for row in rows]


@app.post("/api/classes", status_code=201)
def create_class(payload: ClassIn, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    name, semester = payload.name.strip(), payload.semester.strip()
    with connection() as conn:
        try:
            cur = conn.execute("INSERT INTO classes(name,semester,teacher_user_id) VALUES(?,?,?)", (name, semester, actor["id"]))
        except sqlite3.IntegrityError as exc:
            raise HTTPException(409, "该学期已存在同名班级") from exc
    audit(actor, "class.create", "class", cur.lastrowid, actor["id"] if actor["role"] == "teacher" else None,
          {"name": name, "semester": semester})
    return {"id": cur.lastrowid, "name": name, "semester": semester}


@app.get("/api/classes/{class_id}/students")
def list_class_students(class_id: int, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        _require_class(conn, class_id, actor)
        rows = conn.execute(
            "SELECT id,student_no,name,created_at FROM students WHERE class_id=? ORDER BY student_no", (class_id,)
        ).fetchall()
    return [dict(row) for row in rows]


@app.post("/api/classes/{class_id}/students/import")
def import_roster_json(class_id: int, payload: StudentListIn, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn: _require_class(conn, class_id, actor)
    result = _save_roster(class_id, payload.students)
    audit(actor, "roster.import", "class", class_id, actor["id"] if actor["role"] == "teacher" else None, result)
    return result


@app.post("/api/classes/{class_id}/students/import-file")
async def import_roster_file(class_id: int, request: Request, file: UploadFile = File(...)):
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn: _require_class(conn, class_id, actor)
    filename = file.filename or "名单"
    raw = await file.read()
    if len(raw) > 10 * 1024 * 1024:
        raise HTTPException(413, "名单文件不能超过 10 MB")
    result = _save_roster(class_id, _parse_roster_file(filename, raw))
    audit(actor, "roster.import", "class", class_id, actor["id"] if actor["role"] == "teacher" else None, result)
    return {"filename": filename, **result}


@app.post("/api/classes/{class_id}/invites", status_code=201)
def create_class_invite(class_id: int, payload: InviteCreateIn, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    raw_code = "MATH-" + __import__("secrets").token_urlsafe(18)
    expires_at = datetime.now(timezone.utc) + timedelta(days=payload.expires_days)
    with connection() as conn:
        row = _require_class(conn, class_id, actor)
        cur = conn.execute("""INSERT INTO class_invites(class_id,code_hash,expires_at,max_uses,created_by)
                            VALUES(?,?,?,?,?)""", (class_id, token_hash(raw_code), expires_at.isoformat(), payload.max_uses, actor["id"] or 0))
    audit(actor, "invite.create", "class_invite", cur.lastrowid, row["teacher_user_id"],
          {"class_id": class_id, "expires_at": expires_at.isoformat(), "max_uses": payload.max_uses})
    return {"id": cur.lastrowid, "invite_code": raw_code, "expires_at": expires_at.isoformat(), "max_uses": payload.max_uses,
            "student_activation_url": f"/student-activate?code={raw_code}"}


@app.post("/api/auth/student-activate", status_code=201)
def activate_student(payload: StudentActivateIn, request: Request, response: Response):
    """One-time class invitation + existing roster identity binds a student account."""
    with connection() as conn:
        invite = conn.execute("""SELECT i.*,c.teacher_user_id FROM class_invites i JOIN classes c ON c.id=i.class_id
                                WHERE i.code_hash=? AND i.revoked_at IS NULL""", (token_hash(payload.invite_code),)).fetchone()
        if not invite or datetime.fromisoformat(invite["expires_at"]).astimezone(timezone.utc) <= datetime.now(timezone.utc) or invite["used_count"] >= invite["max_uses"]:
            raise HTTPException(403, "邀请码无效、已过期或使用次数已满")
        roster = conn.execute("SELECT id,name,user_id FROM students WHERE class_id=? AND student_no=?", (invite["class_id"], payload.student_no.strip())).fetchone()
        if not roster or roster["name"].strip() != payload.name.strip():
            raise HTTPException(403, "学号或姓名与该班导入名单不一致")
        if roster["user_id"] is not None:
            raise HTTPException(409, "该名单学生已激活账号，请直接登录")
        from .auth import hash_password
        try:
            user_id = conn.execute("INSERT INTO users(username,display_name,password_hash,role) VALUES(?,?,?,'student')",
                                   (payload.username.strip(), payload.name.strip(), hash_password(payload.password))).lastrowid
        except sqlite3.IntegrityError as exc:
            raise HTTPException(409, "该登录名已存在") from exc
        conn.execute("UPDATE students SET user_id=? WHERE id=?", (user_id, roster["id"]))
        conn.execute("UPDATE class_invites SET used_count=used_count+1 WHERE id=?", (invite["id"],))
    user, token, expires_at = login(payload.username, payload.password)
    response.set_cookie(settings.session_cookie_name, token, httponly=True, secure=settings.cookie_secure, samesite="lax", expires=expires_at)
    audit(user, "student.activate", "class", invite["class_id"], invite["teacher_user_id"], ip=request.client.host if request.client else None)
    return {"user": user, "class_id": invite["class_id"]}


@app.post("/api/assignments", status_code=201)
async def create_assignment(payload: AssignmentIn, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    if payload.basic_ratio + payload.advanced_ratio > 1:
        raise HTTPException(422, "基础和提高比例之和不能超过 1")
    levels = (["基础"] * round(payload.question_count * payload.basic_ratio) +
              ["提高"] * round(payload.question_count * payload.advanced_ratio))
    levels += ["综合"] * (payload.question_count - len(levels))
    with connection() as conn:
        class_row = _require_class(conn, payload.class_id, actor)
        _require_roster(conn, payload.class_id)
        picked, used = [], set()
        for level in levels:
            row = conn.execute("SELECT * FROM questions WHERE chapter=? AND difficulty=? AND review_status='published' AND id NOT IN ({}) ORDER BY RANDOM() LIMIT 1".format(",".join("?" * len(used)) if used else "0"), [payload.chapter, level, *used]).fetchone()
            if row is None:
                row = conn.execute("SELECT * FROM questions WHERE chapter=? AND review_status='published' AND id NOT IN ({}) ORDER BY RANDOM() LIMIT 1".format(",".join("?" * len(used)) if used else "0"), [payload.chapter, *used]).fetchone()
            if row: picked.append(dict(row)); used.add(row["id"])
        if not picked: raise HTTPException(404, "该章节暂无已发布题目")
        # New teacher-created assignments use 20 points for a valid submission
        # and 80 points for answer quality.  Keep the legacy endpoint aligned
        # with the Agent builder so reports always use a 100-point scale.
        quality_total, completion_points = 80.0, 20.0
        parent_scores = [round(quality_total / len(picked), 2) for _ in picked]
        parent_scores[-1] = round(quality_total - sum(parent_scores[:-1]), 2)
        semester = payload.semester.strip() or class_row["semester"]
        cur = conn.execute("INSERT INTO assignments(title,chapter,class_name,class_id,due_at,total_score,semester,status,score_policy,completion_points) VALUES(?,?,?,?,?,?,?,?,?,?)", (payload.title, payload.chapter, class_row["name"], payload.class_id, payload.due_at.isoformat(), 100, semester, "draft", "completion20_quality80_v1", completion_points))
        assignment_id = cur.lastrowid
        conn.executemany(
            "INSERT INTO assignment_questions(assignment_id,question_id,sort_order,score,original_no) VALUES(?,?,?,?,?)",
            [(assignment_id, q["id"], i + 1, parent_scores[i], str(q.get("source_problem_no") or i + 1)) for i, q in enumerate(picked)],
        )
    ai_note = await run_workflow({"task": "assignment_review", "chapter": payload.chapter, "question_ids": list(used)})
    audit(actor, "assignment.create", "assignment", assignment_id, class_row["teacher_user_id"],
          {"class_id": payload.class_id, "question_count": len(picked)})
    return {"id": assignment_id, "questions": picked, "ai_review": ai_note}


@app.post("/api/agent/sync-section")
async def sync_section(section_no: str, limit: int = 80):
    """Teacher-visible operation: refresh a chapter's working cache from 8014."""
    sections = _parse_sections(section_no)
    safe_limit = max(1, min(limit, 80))
    results = await asyncio.gather(
        *[_sync_section_into_local_cache(s, safe_limit) for s in sections]
    )
    if len(results) == 1:
        return results[0]
    return {
        "section_no": "、".join(sections), "sections": results,
        "source_total": sum(item["source_total"] for item in results),
        "synced_count": sum(item["synced_count"] for item in results),
        "skipped_unreadable": sum(item["skipped_unreadable"] for item in results),
        "ai_candidate_count": sum(item["ai_candidate_count"] for item in results),
        "pending_review_count": sum(item["pending_review_count"] for item in results),
        "pending_review_items": [dict(item, section_no=result["section_no"]) for result in results for item in result["pending_review_items"]],
        "unresolved_items": [dict(item, section_no=result["section_no"]) for result in results for item in result["unresolved_items"]],
        "available_by_level": {level: sum(result["available_by_level"][level] for result in results) for level in ("基础", "提高", "综合")},
        "question_ids": [question_id for result in results for question_id in result["question_ids"]],
    }


@app.get("/api/agent/tools")
def agent_tool_manifest(request: Request):
    """Teacher/admin-visible metadata only; no tool can be invoked through this endpoint."""
    require_roles(request, {"admin", "teacher"})
    from .agent_tools import TOOL_ROUTER_VERSION, registry
    return {
        "router_version": TOOL_ROUTER_VERSION,
        "tools": registry.manifest(),
        "policy": "工具仅由 LangGraph 确定性路由调用；学生、OCR 与 PDF 文本不能指定工具或参数。",
    }


@app.get("/api/agent/sections")
async def evidence_sections():
    return await list_evidence_sections()


@app.post("/api/agent/assignments", status_code=201)
async def create_evidence_backed_assignment(payload: AssignmentIn):
    """One-call publish: extract → stratify → assemble for the chosen section(s).

    Delegates the ingestion pipeline to the Agent Orchestrator and then runs the
    existing Dify AI-review workflow over the selected source problems.
    """
    # Top-level hard cap so a stuck external dependency (VLM / Pix2Text / Dify)
    # never freezes the teacher's browser.  When we trip the cap we tell the
    # teacher the truth: the cache could not refresh in time, but they can
    # either retry or open the draft directly from the local cache without
    # waiting on the slow path.
    ASSIGNMENT_DRAFT_TIMEOUT = float(os.environ.get("ASSIGNMENT_DRAFT_TIMEOUT", "90"))
    with connection() as conn:
        class_row = _require_class(conn, payload.class_id)
        _require_roster(conn, payload.class_id)
    try:
        pipeline = await asyncio.wait_for(
            publish_homework(
                sections=_parse_sections(payload.chapter),
                title=payload.title,
                class_id=payload.class_id,
                class_name=class_row["name"],
                due_at=payload.due_at,
                question_count=payload.question_count,
                basic_ratio=payload.basic_ratio,
                advanced_ratio=payload.advanced_ratio,
                subpart_limit=payload.subpart_limit,
            ),
            timeout=ASSIGNMENT_DRAFT_TIMEOUT,
        )
    except asyncio.TimeoutError:
        raise HTTPException(
            504,
            "题库同步暂未响应（外部服务较慢），请稍候再试；如反复失败可先在『同步本章节』单独同步后再生成作业。",
        ) from None
    ai_note = await run_workflow({
        "task": "assignment_review",
        "chapter": payload.chapter,
        "source_problem_ids": pipeline["source_problem_ids"],
    })
    return {
        "id": pipeline["assignment_id"],
        "actual_composition": pipeline["composition"],
        "knowledge_sync": {
            "section_no": "、".join(pipeline["sections"]),
            "synced_count": sum(s["synced_count"] for s in pipeline["sync"]),
            "available_by_level": {
                level: sum(s["available_by_level"][level] for s in pipeline["sync"])
                for level in ("基础", "提高", "综合")
            },
            "unresolved_items": [u for s in pipeline["sync"] for u in s["unresolved_items"]],
        },
        "pdf_path": pipeline["pdf_path"],
        "page_count": pipeline["page_count"],
        "problems": pipeline["problems"],
        "ai_review": ai_note,
    }


@app.post("/api/agent/pipeline/publish", status_code=201)
async def pipeline_publish(payload: AssignmentIn, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    """Full pipeline surface: returns sync + stratify + assemble diagnostics."""
    with connection() as conn:
        class_row = _require_class(conn, payload.class_id, actor)
        _require_roster(conn, payload.class_id)
    return await publish_homework(
        sections=_parse_sections(payload.chapter),
        title=payload.title,
        class_id=payload.class_id,
        class_name=class_row["name"],
        due_at=payload.due_at,
        question_count=payload.question_count,
        basic_ratio=payload.basic_ratio,
        advanced_ratio=payload.advanced_ratio,
        subpart_limit=payload.subpart_limit,
    )


@app.get("/api/assignments")
def list_assignments(request: Request, class_name: str | None = None, include_legacy: bool = False):
    actor = current_user(request)
    sql, args = """SELECT a.*,
        (SELECT COUNT(*) FROM submissions s WHERE s.assignment_id=a.id) AS submission_count,
        (SELECT COUNT(*) FROM submissions s WHERE s.assignment_id=a.id AND s.released_at IS NOT NULL) AS released_submission_count
        FROM assignments a""", []
    clauses = [] if include_legacy else ["a.class_id IS NOT NULL"]
    if class_name:
        clauses.append("a.class_name=?")
        args.append(class_name)
    scope = teacher_id_for_scope(actor)
    if actor["role"] == "student":
        clauses.append("a.status='published'")
        clauses.append("a.class_id IN (SELECT class_id FROM students WHERE user_id=?)")
        args.append(actor["id"])
    elif scope is not None:
        clauses.append("a.class_id IN (SELECT id FROM classes WHERE teacher_user_id=?)")
        args.append(scope)
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    with connection() as conn: return [dict(r) for r in conn.execute(sql + " ORDER BY a.due_at DESC", args)]


def _activation_path_for_assignment(conn: sqlite3.Connection, assignment: dict) -> str:
    """Create a bounded class activation link that returns to this assignment."""
    pending = conn.execute(
        "SELECT COUNT(*) FROM students WHERE class_id=? AND user_id IS NULL", (assignment["class_id"],)
    ).fetchone()[0]
    if not pending:
        return ""
    raw_code = "MATH-" + __import__("secrets").token_urlsafe(18)
    expires_at = datetime.now(timezone.utc) + timedelta(days=30)
    cur = conn.execute(
        "INSERT INTO class_invites(class_id,code_hash,expires_at,max_uses,created_by) VALUES(?,?,?,?,?)",
        (assignment["class_id"], token_hash(raw_code), expires_at.isoformat(), pending, 0),
    )
    next_path = f"/submit?assignment_id={assignment['id']}"
    return "/student-activate?" + urlencode({"code": raw_code, "next": next_path})


def _ensure_assignment_notification(conn: sqlite3.Connection, assignment: sqlite3.Row | dict,
                                    actor: dict) -> sqlite3.Row:
    """Get the publish notification, backfilling safely for pre-feature work."""
    item = dict(assignment)
    row = conn.execute("SELECT * FROM assignment_notifications WHERE assignment_id=?", (item["id"],)).fetchone()
    if row and row["activation_path"] and row["submit_path"].startswith("/student-entry"):
        return row
    activation_path = row["activation_path"] if row and row["activation_path"] else _activation_path_for_assignment(conn, item)
    submit_path = f"/student-entry?assignment_id={item['id']}"
    if row:
        conn.execute("UPDATE assignment_notifications SET submit_path=?,activation_path=? WHERE id=?",
                     (submit_path, activation_path, row["id"]))
        row = conn.execute("SELECT * FROM assignment_notifications WHERE id=?", (row["id"],)).fetchone()
        _record_notification_event(conn, row["id"], actor, "backfilled_student_entry")
        return row
    conn.execute(
        "INSERT INTO assignment_notifications(assignment_id,class_id,title_snapshot,due_at_snapshot,submit_path,activation_path,created_by) VALUES(?,?,?,?,?,?,?)",
        (item["id"], item["class_id"], item["title"], item["due_at"], submit_path, activation_path, actor["id"]),
    )
    row = conn.execute("SELECT * FROM assignment_notifications WHERE assignment_id=?", (item["id"],)).fetchone()
    _record_notification_event(conn, row["id"], actor, "created_on_publish")
    return row

def _notification_payload(notification: sqlite3.Row | dict, request: Request) -> dict:
    item = dict(notification)
    base_url = str(request.base_url).rstrip("/")
    submit_url = base_url + item["submit_path"]
    activation_url = base_url + item["activation_path"] if item.get("activation_path") else ""
    due_text = str(item["due_at_snapshot"]).replace("T", " ")[:16]
    message = (
        f"【作业通知】{item['title_snapshot']}\n"
        f"截止时间：{due_text}\n"
        f"请打开以下链接：已有账号可登录提交；首次使用可激活账号后提交。\n{submit_url}"
    )
    return {**item, "submit_url": submit_url, "activation_url": activation_url, "message": message}


def _record_notification_event(conn: sqlite3.Connection, notification_id: int, actor: dict,
                               action: str, payload: dict | None = None) -> None:
    conn.execute(
        "INSERT INTO assignment_notification_events(notification_id,actor_user_id,action,payload_json) VALUES(?,?,?,?)",
        (notification_id, actor["id"], action, json.dumps(payload or {}, ensure_ascii=False)),
    )


def _assignment_quality_gate(conn: sqlite3.Connection, assignment_id: int) -> dict:
    """Validate every effective question/subpart before students can see a paper."""
    parents = [dict(row) for row in conn.execute(
        """SELECT aq.question_id,aq.original_no,q.content,q.answer,q.rubric,q.review_status
           FROM assignment_questions aq JOIN questions q ON q.id=aq.question_id
           WHERE aq.assignment_id=? ORDER BY aq.sort_order""", (assignment_id,)
    ).fetchall()]
    parts = [dict(row) for row in conn.execute(
        """SELECT question_id,subpart_no,content,answer,rubric FROM assignment_question_parts
           WHERE assignment_id=? ORDER BY question_id,part_order""", (assignment_id,)
    ).fetchall()]
    parts_by_question: dict[int, list[dict]] = {}
    for part in parts:
        parts_by_question.setdefault(part["question_id"], []).append(part)
    items, blocked = [], []
    for parent in parents:
        label = str(parent.get("original_no") or parent["question_id"])
        if parent["review_status"] != "published":
            blocked.append({"label": label, "reason": "题库审核状态不是 published"})
        effective = parts_by_question.get(parent["question_id"]) or [parent]
        for part in effective:
            part_label = label + (f"（{part['subpart_no']}）" if part.get("subpart_no") else "")
            reasons = []
            if len(str(part.get("content") or "").strip()) < 6:
                reasons.append("题干缺失或过短")
            if not str(part.get("answer") or "").strip():
                reasons.append("标准答案缺失")
            if not str(part.get("rubric") or "").strip():
                reasons.append("评分点缺失")
            item = {"label": part_label, "ready": not reasons, "reasons": reasons}
            items.append(item)
            for reason in reasons:
                blocked.append({"label": part_label, "reason": reason})
    return {
        "ready": bool(parents) and not blocked,
        "effective_item_count": len(items), "items": items, "blocked": blocked,
    }


@app.get("/api/assignments/{assignment_id}/quality")
def assignment_quality(assignment_id: int, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        _require_assignment(conn, assignment_id, actor)
        return _assignment_quality_gate(conn, assignment_id)


@app.post("/api/assignments/{assignment_id}/publish")
def publish_assignment(assignment_id: int, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        assignment = _require_assignment(conn, assignment_id, actor)
        if assignment["status"] == "published":
            return {"ok": True, "message": "该作业已发布"}
        quality = _assignment_quality_gate(conn, assignment_id)
        question_count = quality["effective_item_count"]
        if not quality["ready"]:
            preview = "；".join(f"{item['label']}：{item['reason']}" for item in quality["blocked"][:6])
            raise HTTPException(409, "发布前题库质量检查未通过：" + preview)
        conn.execute("UPDATE assignments SET status='published' WHERE id=?", (assignment_id,))
        notification = _ensure_assignment_notification(conn, assignment, actor)
        _record_notification_event(conn, notification["id"], actor, "created_on_publish", {"question_count": question_count})
    audit(actor, "assignment.publish", "assignment", assignment_id, assignment["teacher_user_id"], {"question_count": question_count})
    return {"ok": True, "message": "作业已发布，学生现在可以查看并提交", "notification_id": notification["id"]}


@app.get("/api/assignments/{assignment_id}/notification")
def assignment_notification(assignment_id: int, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        assignment = _require_assignment(conn, assignment_id, actor)
        if assignment["status"] != "published":
            raise HTTPException(409, "请先发布作业，再生成通知")
        row = _ensure_assignment_notification(conn, assignment, actor)
        events = [dict(r) for r in conn.execute(
            "SELECT action,created_at FROM assignment_notification_events WHERE notification_id=? ORDER BY id DESC LIMIT 30",
            (row["id"],),
        ).fetchall()]
    return {**_notification_payload(row, request), "events": events}


@app.post("/api/assignments/{assignment_id}/notification/copied")
def notification_copied(assignment_id: int, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        assignment = _require_assignment(conn, assignment_id, actor)
        row = _ensure_assignment_notification(conn, assignment, actor)
        _record_notification_event(conn, row["id"], actor, "copied")
    return {"ok": True}


@app.get("/api/assignments/{assignment_id}/notification/qr.png")
def notification_qr(assignment_id: int, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        assignment = _require_assignment(conn, assignment_id, actor)
        row = _ensure_assignment_notification(conn, assignment, actor)
        _record_notification_event(conn, row["id"], actor, "qr_downloaded")
    try:
        import qrcode
    except ImportError as exc:
        raise HTTPException(503, "二维码组件未安装") from exc
    image = qrcode.make(_notification_payload(row, request)["submit_url"])
    data = BytesIO()
    image.save(data, format="PNG")
    return Response(content=data.getvalue(), media_type="image/png",
                    headers={"Content-Disposition": f'inline; filename="assignment_{assignment_id}_qr.png"'})


@app.get("/api/assignments/{assignment_id}/notification/roster.csv")
def notification_roster_csv(assignment_id: int, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        assignment = _require_assignment(conn, assignment_id, actor)
        row = _ensure_assignment_notification(conn, assignment, actor)
        students = conn.execute("SELECT student_no,name FROM students WHERE class_id=? ORDER BY student_no", (assignment["class_id"],)).fetchall()
        _record_notification_event(conn, row["id"], actor, "roster_exported", {"student_count": len(students)})
    payload = _notification_payload(row, request)
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["学号", "姓名", "作业名称", "截止时间", "学生提交链接", "通知内容"])
    for student in students:
        writer.writerow([student["student_no"], student["name"], payload["title_snapshot"],
                         payload["due_at_snapshot"], payload["submit_url"], payload["message"]])
    return Response(content="\ufeff" + output.getvalue(), media_type="text/csv; charset=utf-8",
                    headers={"Content-Disposition": f'attachment; filename="assignment_{assignment_id}_notification_roster.csv"'})


@app.get("/api/assignments/{assignment_id}/draft-items")
def assignment_draft_items(assignment_id: int, request: Request):
    """Teacher-only editable copy of a draft's visible questions."""
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        assignment = _require_assignment(conn, assignment_id, actor)
        if assignment["status"] != "draft":
            raise HTTPException(409, "只能编辑尚未发布的作业草稿")
        parents = [dict(r) for r in conn.execute(
            """SELECT aq.question_id,aq.original_no,aq.score,q.content,o.content AS override_content,o.score AS override_score
               FROM assignment_questions aq JOIN questions q ON q.id=aq.question_id
               LEFT JOIN assignment_question_overrides o ON o.assignment_id=aq.assignment_id AND o.question_id=aq.question_id
               WHERE aq.assignment_id=? ORDER BY aq.sort_order""", (assignment_id,)
        ).fetchall()]
        parts = [dict(r) for r in conn.execute(
            """SELECT question_id,subpart_no,part_order,content,score FROM assignment_question_parts
               WHERE assignment_id=? ORDER BY question_id,part_order""", (assignment_id,)
        ).fetchall()]
    by_parent = {}
    for part in parts:
        by_parent.setdefault(part["question_id"], []).append(part)
    items = []
    for parent in parents:
        selected = by_parent.get(parent["question_id"], [])
        if selected:
            for part in selected:
                items.append({"question_id": parent["question_id"], "subpart_no": part["subpart_no"],
                              "label": f"{parent['original_no']}（{part['subpart_no']}）",
                              "content": part["content"], "score": part["score"], "is_subpart": True})
        else:
            items.append({"question_id": parent["question_id"], "subpart_no": "", "label": parent["original_no"],
                          "content": parent["override_content"] or parent["content"],
                          "score": parent["override_score"] if parent["override_score"] is not None else parent["score"],
                          "is_subpart": False})
    return {"assignment_id": assignment_id, "title": assignment["title"], "due_at": assignment["due_at"], "items": items}


def _refresh_assignment_total(conn: sqlite3.Connection, assignment_id: int) -> None:
    total = conn.execute(
        """SELECT COALESCE(SUM(COALESCE(
             (SELECT SUM(p.score) FROM assignment_question_parts p WHERE p.assignment_id=aq.assignment_id AND p.question_id=aq.question_id),
             o.score, aq.score)), 0)
           FROM assignment_questions aq LEFT JOIN assignment_question_overrides o
             ON o.assignment_id=aq.assignment_id AND o.question_id=aq.question_id
           WHERE aq.assignment_id=?""", (assignment_id,)
    ).fetchone()[0]
    policy = conn.execute("SELECT score_policy FROM assignments WHERE id=?", (assignment_id,)).fetchone()
    # Editing the wording/quality points of a modern draft never turns its
    # visible total into 80 by accident; the completion component stays clear.
    visible_total = 100 if policy and policy["score_policy"] == "completion20_quality80_v1" else round(float(total or 0), 2)
    conn.execute("UPDATE assignments SET total_score=? WHERE id=?", (visible_total, assignment_id))


@app.patch("/api/assignments/{assignment_id}/draft-items")
def update_assignment_draft_item(assignment_id: int, payload: AssignmentDraftItemIn, request: Request):
    """Edit only this assignment's copy; shared question bank remains untouched."""
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        assignment = _require_assignment(conn, assignment_id, actor)
        if assignment["status"] != "draft":
            raise HTTPException(409, "作业已发布，不能再改题目；请复制后创建新草稿")
        if conn.execute("SELECT 1 FROM submissions WHERE assignment_id=? LIMIT 1", (assignment_id,)).fetchone():
            raise HTTPException(409, "该草稿已有提交，不能修改题目")
        if not conn.execute("SELECT 1 FROM assignment_questions WHERE assignment_id=? AND question_id=?",
                            (assignment_id, payload.question_id)).fetchone():
            raise HTTPException(404, "该题不属于此作业")
        if payload.subpart_no:
            changed = conn.execute(
                """UPDATE assignment_question_parts SET content=?,score=?
                   WHERE assignment_id=? AND question_id=? AND subpart_no=?""",
                (payload.content.strip(), payload.score, assignment_id, payload.question_id, payload.subpart_no)
            ).rowcount
            if not changed:
                raise HTTPException(404, "未找到该小问")
        else:
            conn.execute(
                """INSERT INTO assignment_question_overrides(assignment_id,question_id,content,score,updated_at)
                   VALUES(?,?,?,?,CURRENT_TIMESTAMP)
                   ON CONFLICT(assignment_id,question_id) DO UPDATE SET
                     content=excluded.content,score=excluded.score,updated_at=CURRENT_TIMESTAMP""",
                (assignment_id, payload.question_id, payload.content.strip(), payload.score)
            )
        _refresh_assignment_total(conn, assignment_id)
    audit(actor, "assignment.draft_item.update", "assignment", assignment_id,
          assignment["teacher_user_id"], {"question_id": payload.question_id, "subpart_no": payload.subpart_no})
    return {"ok": True, "message": "已保存到本次作业草稿；题库原题未修改。"}


@app.patch("/api/assignments/{assignment_id}")
def update_assignment(assignment_id: int, payload: AssignmentUpdateIn, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        assignment = _require_assignment(conn, assignment_id, actor)
        changed = conn.execute(
            "UPDATE assignments SET title=?, due_at=? WHERE id=?",
            (payload.title.strip(), payload.due_at.isoformat(), assignment_id),
        ).rowcount
    if not changed:
        raise HTTPException(404, "作业不存在")
    audit(actor, "assignment.update", "assignment", assignment_id, assignment["teacher_user_id"])
    return {"ok": True, "message": "作业信息已更新"}


@app.delete("/api/assignments/{assignment_id}")
def delete_assignment(assignment_id: int, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        assignment = _require_assignment(conn, assignment_id, actor)
        submission_count = conn.execute("SELECT COUNT(*) FROM submissions WHERE assignment_id=?", (assignment_id,)).fetchone()[0]
        if submission_count:
            raise HTTPException(409, "该作业已有学生提交，不能删除；请保留评分证据。")
        conn.execute("DELETE FROM assignment_questions WHERE assignment_id=?", (assignment_id,))
        conn.execute("DELETE FROM assignments WHERE id=?", (assignment_id,))
    audit(actor, "assignment.delete", "assignment", assignment_id, assignment["teacher_user_id"])
    return {"ok": True, "message": "作业已删除"}


def _demo_submission_rows(conn: sqlite3.Connection, actor: dict) -> list[sqlite3.Row]:
    """Only identify isolated demo accounts/classes; never infer from an ordinary title."""
    rows = conn.execute(
        """SELECT s.id,s.assignment_id,s.student_no,s.student_name,s.file_path,
                  a.title AS assignment_title,c.teacher_user_id,c.name AS class_name,c.semester
           FROM submissions s
           JOIN assignments a ON a.id=s.assignment_id
           JOIN classes c ON c.id=a.class_id
           WHERE UPPER(s.student_no) LIKE 'DEMO%'
              OR UPPER(COALESCE(c.semester,''))='DEMO'
              OR c.name LIKE '%演示%'"""
    ).fetchall()
    scope = teacher_id_for_scope(actor)
    return [row for row in rows if scope is None or row['teacher_user_id'] == scope]


@app.get("/api/demo/submissions/summary")
def demo_submission_summary(request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        rows = _demo_submission_rows(conn, actor)
    return {
        "count": len(rows),
        "items": [{"submission_id": row["id"], "assignment_title": row["assignment_title"],
                   "student_no": row["student_no"], "student_name": row["student_name"],
                   "class_name": row["class_name"]} for row in rows],
        "message": "仅统计演示账号（DEMO*）或演示班级（学期 DEMO / 名称含演示）的提交。",
    }


@app.delete("/api/demo/submissions")
def delete_demo_submissions(payload: DemoCleanupIn, request: Request):
    """Remove disposable demo submissions while preserving real classroom evidence."""
    actor = require_roles(request, {"admin", "teacher"})
    if not payload.confirm:
        raise HTTPException(400, "请先确认仅清理演示提交")
    with connection() as conn:
        rows = _demo_submission_rows(conn, actor)
        ids = [int(row["id"]) for row in rows]
        if not ids:
            return {"ok": True, "deleted_count": 0, "message": "没有可清理的演示提交。"}
        marks = ",".join("?" for _ in ids)
        for table, column in (
            ("agent_learning_attempts", "submission_id"),
            ("agent_learning_traces", "submission_id"),
            ("teacher_evaluation_cases", "source_submission_id"),
            ("submission_question_regions", "submission_id"),
            ("grading_experiences", "submission_id"),
            ("grading_jobs", "submission_id"),
            ("submissions", "id"),
        ):
            conn.execute(f"DELETE FROM {table} WHERE {column} IN ({marks})", ids)
    upload_root = Path(settings.upload_dir).resolve()
    removed_files = 0
    for row in rows:
        path = Path(row["file_path"]).resolve()
        if path.is_file() and upload_root in path.parents:
            try:
                path.unlink()
                removed_files += 1
            except OSError:
                pass
    audit(actor, "demo_submission.cleanup", "submission", ",".join(map(str, ids)),
          None, {"deleted_count": len(ids), "removed_files": removed_files})
    return {"ok": True, "deleted_count": len(ids), "removed_files": removed_files,
            "message": f"已清理 {len(ids)} 份演示提交；真实学生记录未受影响。"}


@app.get("/api/assignments/{assignment_id}/print", response_class=HTMLResponse)
def printable_assignment(assignment_id: int, request: Request):
    actor = current_user(request)
    assignment, rows = _load_assignment_items(assignment_id, actor)
    # Use source textbook numbers (and selected sub-part numbers), then remove
    # the repeated source number from each stem.
    # stem. Wrap raw LaTeX fragments first, then escape HTML so entities inside math
    # (e.g. >) are handled by the browser/MathJax correctly.
    items = "".join(
        f"<section class='subpart' if False else ''><h3>{display_problem_no(dict(r), i)} · {float(r['score']):g}分</h3><div class='problem-body'>{escape(_wrap_latex_for_html(_strip_selected_subpart_prefix(strip_source_problem_prefix(r['content']), str(r.get('original_no') or ''))))}</div><div class='space {'small' if r.get('subpart_no') else ''}'></div></section>"
        for i, r in enumerate(rows)
    )
    return f"""<!doctype html><html lang='zh-CN'><meta charset='utf-8'><title>{escape(assignment['title'])}</title>
    <style>@page{{size:A4;margin:18mm}}body{{font-family:'Microsoft YaHei',sans-serif;color:#111;line-height:1.65}}header{{border-bottom:2px solid #1e3a5f}}h1{{text-align:center}}.meta{{display:flex;justify-content:space-between}}section{{break-inside:avoid;margin-top:18px}}h3{{margin:0 0 8px;font-size:18px}}.problem-body{{font-size:17px;white-space:pre-wrap}}.space{{height:175px;border-bottom:1px dashed #cbd5e1}}.space.small{{height:105px}}</style>
    <script>
    MathJax = {{
      tex: {{ inlineMath: [['$', '$'], ['\\\\(', '\\\\)']], displayMath: [['$$', '$$'], ['\\\\[', '\\\\]']] }},
      svg: {{ fontCache: 'global' }}
    }};
    </script>
    <script id="MathJax-script" async src="https://cdn.jsdelivr.net/npm/mathjax@3/es5/tex-svg.js"></script>
    <header><h1>{escape(assignment['title'])}</h1><div class='meta'><span>班级：{escape(assignment['class_name'])}</span><span>姓名：__________</span><span>学号：__________</span></div><p>章节：{escape(assignment['chapter'])}　截止：{escape(assignment['due_at'])}　总分：{assignment['total_score']}</p></header>{items}</html>"""


@app.get("/api/assignments/{assignment_id}/submissions")
def assignment_submissions(assignment_id: int, request: Request):
    """Teacher's complete result list, including auto-graded submissions.

    The review queue intentionally contains only uncertain work.  This endpoint
    is separate so successful automatic grading does not become invisible.
    """
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        assignment = _require_assignment(conn, assignment_id, actor)
        rows = conn.execute(
            """SELECT s.id,s.student_no,s.student_name,s.status,s.score,s.needs_review,s.submitted_at,
                      j.status AS grading_status,j.result_json
               FROM submissions s LEFT JOIN grading_jobs j ON j.submission_id=s.id
               WHERE s.assignment_id=? ORDER BY s.submitted_at DESC, s.id DESC""",
            (assignment_id,),
        ).fetchall()
    items = []
    for row in rows:
        item = dict(row)
        result = json.loads(item.pop("result_json") or "{}")
        question_results = result.get("results") or []
        item["total_score"] = result.get("total_score", item["score"])
        item["max_score"] = result.get("max_score", assignment["total_score"])
        item["recognized_summary"] = "；".join(
            str(q.get("recognized_work") or "未识别") for q in question_results
        )
        item["qwen_error"] = result.get("qwen_error", "")
        items.append(item)
    return {"assignment_id": assignment_id, "title": assignment["title"], "submissions": items}


def _load_assignment_items(assignment_id: int, actor: dict | None = None) -> tuple[dict, list[dict]]:
    """Return printable rows, expanding safely selected sub-questions."""
    with connection() as conn:
        assignment = _require_assignment(conn, assignment_id, actor)
        parents = [dict(r) for r in conn.execute(
            """SELECT q.content, q.chapter, q.question_type, q.source_problem_no, aq.question_id,
                      aq.sort_order, aq.score, aq.original_no, o.content AS override_content, o.score AS override_score
               FROM assignment_questions aq JOIN questions q ON q.id=aq.question_id
               LEFT JOIN assignment_question_overrides o ON o.assignment_id=aq.assignment_id AND o.question_id=aq.question_id
               WHERE aq.assignment_id=? ORDER BY aq.sort_order""", (assignment_id,)
        ).fetchall()]
        parts = [dict(r) for r in conn.execute(
            """SELECT question_id,subpart_no,part_order,content,answer,rubric,score
               FROM assignment_question_parts WHERE assignment_id=?
               ORDER BY question_id,part_order""", (assignment_id,)
        ).fetchall()]
    by_parent: dict[int, list[dict]] = {}
    for part in parts:
        by_parent.setdefault(part["question_id"], []).append(part)
    rows: list[dict] = []
    for parent in parents:
        selected = by_parent.get(parent["question_id"], [])
        if not selected:
            parent["content"] = parent.get("override_content") or parent["content"]
            parent["score"] = parent.get("override_score") if parent.get("override_score") is not None else parent["score"]
            rows.append(parent)
            continue
        for part in selected:
            rows.append({
                **parent, **part,
                "original_no": f"{parent['original_no']}（{part['subpart_no']}）",
                "content": part["content"],
            })
    return dict(assignment), rows


@app.get("/api/assignments/{assignment_id}/pdf")
def download_assignment_pdf(assignment_id: int, request: Request):
    """Render the real printable A4 homework PDF (original numbers + answer blanks)."""
    assignment, items = _load_assignment_items(assignment_id, current_user(request))
    out_dir = Path(settings.upload_dir) / "assignments"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"assignment_{assignment_id}.pdf"
    try:
        path, _ = build_assignment_pdf(assignment, items, out_path)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"PDF 生成失败：{str(exc)[:220]}")
    return FileResponse(path, filename=f"作业_{assignment_id}.pdf", media_type="application/pdf")


@app.get("/api/assignments/{assignment_id}/latex")
def download_assignment_latex(assignment_id: int, request: Request):
    """Return the LaTeX source of every selected problem as JSON (设计二：含 latex 源码)."""
    assignment, items = _load_assignment_items(assignment_id, current_user(request))
    return export_latex_source(assignment, items)


@app.get("/api/assignments/{assignment_id}/latex.tex")
def download_assignment_latex_tex(assignment_id: int, request: Request):
    """Return a compilable .tex document wrapping the selected problems."""
    assignment, items = _load_assignment_items(assignment_id, current_user(request))
    out_dir = Path(settings.upload_dir) / "assignments"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"assignment_{assignment_id}.tex"
    out_path.write_text(latex_document(assignment, items), encoding="utf-8")
    return FileResponse(out_path, filename=f"作业_{assignment_id}.tex", media_type="application/x-tex")


@app.get("/student-entry", response_class=HTMLResponse)
def student_assignment_entry(assignment_id: int, request: Request):
    """Mobile QR landing page: choose activation or login, then return to work."""
    with connection() as conn:
        assignment = conn.execute("SELECT id,title,status FROM assignments WHERE id=?", (assignment_id,)).fetchone()
        if not assignment or assignment["status"] != "published":
            raise HTTPException(404, "作业不存在或尚未发布")
        notification = conn.execute("SELECT activation_path FROM assignment_notifications WHERE assignment_id=?", (assignment_id,)).fetchone()
    try:
        actor = current_user(request)
    except HTTPException as exc:
        if exc.status_code != 401:
            raise
        actor = None
    submit_path = f"/submit?assignment_id={assignment_id}"
    if actor and actor["role"] == "student":
        return RedirectResponse(submit_path, status_code=303)
    login_path = "/login?" + urlencode({"next": submit_path})
    activation_path = notification["activation_path"] if notification else ""
    activate_link = (f'<a class="btn" href="{escape(activation_path, quote=True)}">首次使用：激活账号并提交</a>'
                     if activation_path else '<p class="hint">本班激活链接尚未生成，请联系教师在作业通知中重新打开“通知”。</p>')
    return HTMLResponse(f"""<!doctype html><html lang="zh-CN"><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>进入作业</title>
<style>:root{{--ink:#102a43;--muted:#60758a;--blue:#2563eb;--line:#dbe5f0}}*{{box-sizing:border-box}}body{{margin:0;background:#f6f8fc;color:var(--ink);font:16px "Microsoft YaHei",sans-serif;display:grid;place-items:center;min-height:100dvh;padding:24px}}.card{{width:min(480px,100%);background:#fff;border:1px solid var(--line);border-radius:20px;padding:30px;box-shadow:0 12px 32px #102a4312}}h1{{margin:0;font-size:27px}}p{{line-height:1.75;color:var(--muted)}}.btn{{display:block;text-align:center;text-decoration:none;margin-top:14px;padding:14px;border-radius:10px;background:var(--blue);color:#fff;font-weight:700}}.btn.secondary{{background:#eff6ff;color:#1d4ed8}}.hint{{font-size:14px}}@media(max-width:640px){{body{{padding:0;background:#fff}}.card{{min-height:100dvh;width:100%;border:0;border-radius:0;box-shadow:none;padding:88px 24px 32px}}h1{{font-size:26px}}.btn{{padding:15px;font-size:17px}}}}</style>
<main class="card"><h1>进入作业</h1><p><b>{escape(assignment["title"])}</b><br>请选择你的情况，完成后会自动进入作业提交页。</p>{activate_link}<a class="btn secondary" href="{escape(login_path, quote=True)}">已有账号：登录并提交</a><p class="hint">激活时会核验教师导入的学号和姓名。</p></main></html>""")


@app.get("/student", response_class=HTMLResponse)
def student_home_page(request: Request):
    try:
        actor = current_user(request)
    except HTTPException as exc:
        if exc.status_code == 401:
            return RedirectResponse("/login?next=/student", status_code=303)
        raise
    if actor["role"] != "student":
        return RedirectResponse("/", status_code=303)
    return FileResponse(Path(__file__).with_name("student_home.html"),
                        headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"})


@app.get("/api/student/assignments")
def student_assignments(request: Request):
    """Student dashboard: one latest submission per assigned homework."""
    actor = require_roles(request, {"student"})
    with connection() as conn:
        rows = conn.execute(
            """SELECT a.id,a.title,a.chapter,a.due_at,a.total_score,a.status AS assignment_status,
                      a.score_policy,a.completion_points,c.name AS class_name,
                      st.student_no,st.name AS roster_name,
                      s.id AS submission_id,s.status AS submission_status,s.score,
                      s.completion_score,s.quality_score,s.quality_max_score,
                      s.submitted_at,s.released_at,s.feedback,j.status AS grading_status,j.result_json
               FROM assignments a
               JOIN students st ON st.class_id=a.class_id AND st.user_id=?
               JOIN classes c ON c.id=a.class_id
               LEFT JOIN submissions s ON s.id=(
                    SELECT s2.id FROM submissions s2
                    WHERE s2.assignment_id=a.id AND s2.student_no=st.student_no
                    ORDER BY s2.id DESC LIMIT 1
               )
               LEFT JOIN grading_jobs j ON j.submission_id=s.id
               WHERE a.status='published'
               ORDER BY datetime(a.due_at) ASC, a.id DESC""",
            (actor["id"],),
        ).fetchall()
    now = datetime.now(timezone.utc)
    items = []
    for row in rows:
        item = dict(row)
        grading = json.loads(item.pop("result_json") or "{}")
        due = item.get("due_at") or ""
        try:
            due_at = datetime.fromisoformat(due.replace("Z", "+00:00"))
            due_at = due_at if due_at.tzinfo else due_at.replace(tzinfo=timezone.utc)
        except ValueError:
            due_at = now
        if item["submission_id"] is None:
            state = "pending" if due_at >= now else "overdue"
        elif item["released_at"]:
            state = "released"
        elif item["grading_status"] in {"queued", "running"}:
            state = "grading"
        elif item["submission_status"] == "review_required":
            state = "reviewing"
        else:
            state = "submitted"
        item["state"] = state
        item["grading"] = grading
        items.append(item)
    return {"assignments": items}


@app.get("/submit", response_class=HTMLResponse)
def student_submit_page(request: Request, assignment_id: int | None = None):
    """Student-only homework page; never leave an unauthenticated visitor on empty /submit."""
    try:
        actor = current_user(request)
    except HTTPException as exc:
        if exc.status_code == 401:
            next_path = "/submit" if assignment_id is None else f"/submit?assignment_id={assignment_id}"
            return RedirectResponse(f"/login?next={next_path}", status_code=303)
        raise
    if actor["role"] != "student":
        return RedirectResponse("/", status_code=303)
    if assignment_id is None:
        return RedirectResponse("/student", status_code=303)
    return FileResponse(Path(__file__).with_name("student_submit.html"), headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"})


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    try:
        actor = current_user(request)
    except HTTPException as exc:
        if exc.status_code != 401:
            raise
    else:
        return RedirectResponse("/student" if actor["role"] == "student" else "/", status_code=303)
    return FileResponse(Path(__file__).with_name("login.html"), headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"})


@app.get("/student-activate", response_class=HTMLResponse)
def student_activate_page():
    return FileResponse(Path(__file__).with_name("student_activate.html"))


# 学生提交加固常量
_ALLOWED_EXT = {".pdf", ".jpg", ".jpeg", ".png", ".doc", ".docx"}
_MAX_UPLOAD_BYTES = 30 * 1024 * 1024  # 30MB 上限
_STUDENT_NO_RE = re.compile(r"[A-Za-z0-9_]{1,32}")


@app.post("/api/assignments/{assignment_id}/submissions", status_code=201)
async def submit_homework(assignment_id: int, background_tasks: BackgroundTasks, request: Request, student_no: str = Form(...), student_name: str = Form(""), file: UploadFile | None = File(None), files: list[UploadFile] = File(default=[])):
    # A shared assignment URL accepts a roster-verified student submission without a login.
    # Signed-in students remain bound to their own roster identity below.
    try:
        actor = current_user(request)
    except HTTPException as exc:
        if exc.status_code != 401:
            raise
        actor = None
    # 1) 学号格式与路径穿越防护：仅允许字母/数字/下划线，避免 folder = upload_dir/student_no 被注入
    student_no = (student_no or "").strip()
    if not _STUDENT_NO_RE.fullmatch(student_no):
        raise HTTPException(422, "学号格式不合法（仅允许字母、数字、下划线，最长 32 位）")
    student_name = (student_name or "").strip()[:40]

    # 2) 文件名与类型校验。兼容旧客户端单文件字段 file，也支持手机端有序 files[]。
    uploads = list(files or [])
    if file is not None:
        uploads.insert(0, file)
    if not uploads:
        raise HTTPException(422, "请至少上传一份作业文件")
    if len(uploads) > 10:
        raise HTTPException(422, "一次最多上传 10 张作业图片")
    upload_meta = []
    for uploaded in uploads:
        filename = (uploaded.filename or "").strip()
        suffix = Path(filename).suffix.lower()
        if not filename:
            raise HTTPException(422, "文件名缺失")
        if suffix not in _ALLOWED_EXT:
            raise HTTPException(415, "仅支持 PDF、图片或 Word 文档")
        upload_meta.append((uploaded, filename, suffix))
    if len(upload_meta) > 1 and any(suffix not in {".jpg", ".jpeg", ".png"} for _, _, suffix in upload_meta):
        raise HTTPException(422, "多页提交请仅选择 JPG 或 PNG 图片；PDF 请作为单个文件上传")

    # 3) 防重复提交：同一作业+学号若已有正在批改（queued/running）的任务则拒绝
    with connection() as conn:
        assignment = _require_assignment(conn, assignment_id, actor)
        if assignment["status"] != "published":
            raise HTTPException(409, "该作业仍是教师草案，发布后学生才能提交")
        if assignment["class_id"] is None:
            raise HTTPException(409, "这是历史演示作业，不能再接收提交；请从已建班级重新发布作业")
        enrolled = conn.execute(
            "SELECT name,user_id FROM students WHERE class_id=? AND student_no=?", (assignment["class_id"], student_no)
        ).fetchone()
        if not enrolled:
            raise HTTPException(403, "该学号不在本班名单中，请联系教师核对班级名单")
        if student_name and student_name != enrolled["name"]:
            raise HTTPException(403, "姓名与该班导入名单不一致，请联系教师核对")
        if actor and actor["role"] == "student" and enrolled["user_id"] != actor["id"]:
            raise HTTPException(403, "当前学生账号与提交学号不一致")
        if not student_name:
            student_name = enrolled["name"]
        if conn.execute(
            """SELECT 1 FROM submissions s JOIN grading_jobs j ON j.submission_id=s.id
               WHERE s.assignment_id=? AND s.student_no=? AND j.status IN ('queued','running') LIMIT 1""",
            (assignment_id, student_no),
        ).fetchone():
            raise HTTPException(409, "该作业已有正在批改的提交，请稍候或联系老师。")

        # 4) 落盘：单文件保持原格式；多张图片按选择顺序合成为多页 PDF。
        # 总大小仍受 30MB 限制，避免以“多页”绕过上传限额。
        folder = Path(settings.upload_dir) / str(assignment_id) / student_no
        folder.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%S')
        written = 0
        page_count = len(upload_meta)
        try:
            raw_pages: list[tuple[bytes, str]] = []
            for uploaded, filename, suffix in upload_meta:
                chunks = []
                while True:
                    chunk = uploaded.file.read(65536)
                    if not chunk:
                        break
                    written += len(chunk)
                    if written > _MAX_UPLOAD_BYTES:
                        raise HTTPException(413, "文件总大小超过 30MB，请压缩后重试")
                    chunks.append(chunk)
                raw = b"".join(chunks)
                if not raw:
                    raise HTTPException(422, "存在空文件，请重新选择")
                raw_pages.append((raw, suffix))
            if len(raw_pages) == 1:
                filename = upload_meta[0][1]
                path = folder / f"{timestamp}_{Path(filename).name}"
                path.write_bytes(raw_pages[0][0])
            else:
                path = folder / f"{timestamp}_mobile_pages.pdf"
                pages = []
                for raw, _suffix in raw_pages:
                    with Image.open(BytesIO(raw)) as image:
                        pages.append(image.convert("RGB"))
                pages[0].save(path, format="PDF", save_all=True, append_images=pages[1:], resolution=150.0)
                if path.stat().st_size > _MAX_UPLOAD_BYTES:
                    raise HTTPException(413, "合成后的多页作业超过 30MB，请减少页数或重新拍摄")
        except HTTPException:
            if 'path' in locals():
                path.unlink(missing_ok=True)
            raise
        except Exception:
            if 'path' in locals():
                path.unlink(missing_ok=True)
            raise HTTPException(422, "图片合并失败，请重新拍摄清晰的 JPG 或 PNG 图片")

        # 始终存绝对路径，避免批改时依赖服务器进程 cwd（曾致相对路径提交无法定位原文件）。
        abs_path = path.resolve()
        completion_points = float(assignment["completion_points"] or 0) if "completion_points" in assignment.keys() else 0.0
        policy = str(assignment["score_policy"] or "legacy") if "score_policy" in assignment.keys() else "legacy"
        cur = conn.execute(
            "INSERT INTO submissions(assignment_id,student_no,student_name,file_path,status,completion_score,score_policy_version) VALUES(?,?,?,?,?,?,?)",
            (assignment_id, student_no, student_name, str(abs_path), "submitted", completion_points, policy),
        )
        submission_id = cur.lastrowid
        job_id = conn.execute("INSERT INTO grading_jobs(submission_id) VALUES(?)", (submission_id,)).lastrowid
    try:
        queue_info = dispatch_grading_job(job_id, background_tasks)
    except QueueUnavailable as exc:
        # Submission is safely archived and remains queued; the scheduler will
        # retry dispatch once Redis recovers.  Tell the caller it was accepted.
        queue_info = {"backend": "rq", "queue_warning": str(exc)}
    audit(actor, "submission.create", "submission", submission_id,
          assignment["teacher_user_id"], {"assignment_id": assignment_id})
    return {"id": submission_id, "grading_job_id": job_id, "page_count": page_count,
            "completion_score": completion_points, "score_policy": policy,
            "queue": queue_info,
            "message": "提交成功，系统已按页归档并进入批改队列。"}


@app.post("/api/grading/run-due")
def queue_due_grading(request: Request):
    require_roles(request, {"admin", "teacher"})
    now = datetime.now(timezone.utc).isoformat()
    queued = enqueue_due_grading(now)
    dispatch = enqueue_pending_grading_jobs()
    return {"queued": queued, "dispatch": dispatch, "message": "已为截止作业创建并投递批改任务；主观题将进入复核队列。"}


def _review_priority(grading_status: str | None, result: dict) -> tuple[int, list[str]]:
    """Deterministic prioritisation: failures and weak evidence appear first."""
    score, reasons = 0, []
    if grading_status == "failed":
        score += 100; reasons.append("初评任务失败")
    if grading_status in {"queued", "running"}:
        score += 70; reasons.append("初评尚未完成")
    if result.get("qwen_error"):
        score += 60; reasons.append("模型调用异常")
    rows = result.get("results") or []
    if not rows and grading_status == "completed":
        score += 55; reasons.append("缺少逐题证据")
    for item in rows:
        if item.get("needs_review"):
            score += 25; reasons.append(f"第 {item.get('sort_order', '?')} 题需复核")
        confidence = item.get("confidence")
        if confidence is not None and float(confidence) < 0.85:
            score += 15; reasons.append(f"第 {item.get('sort_order', '?')} 题置信度偏低")
        if item.get("review_reasons"):
            score += 8
    return score, list(dict.fromkeys(reasons))[:4]


@app.get("/api/reviews")
def list_reviews(
    request: Request,
    class_id: int | None = None,
    assignment_id: int | None = None,
    question_type: str | None = None,
    risk: str | None = None,
):
    """Teacher queue with deterministic priority and transparent filters."""
    actor = require_roles(request, {"admin", "teacher"})
    if question_type and question_type not in {"calc", "proof"}:
        raise HTTPException(422, "题型筛选仅支持 calc 或 proof")
    if risk and risk not in {"qwen_error", "low_confidence", "ocr_failure", "proof", "needs_review"}:
        raise HTTPException(422, "不支持的风险筛选")
    scope = teacher_id_for_scope(actor)
    with connection() as conn:
        sql = (
            """SELECT s.id, s.assignment_id, s.student_no, s.student_name, s.status, s.score,
                      s.feedback, s.submitted_at, j.status AS grading_status, j.result_json, a.title,
                      a.class_id, c.name AS class_name
               FROM submissions s JOIN assignments a ON a.id=s.assignment_id
               JOIN classes c ON c.id=a.class_id
               LEFT JOIN grading_jobs j ON j.submission_id=s.id
               WHERE a.class_id IS NOT NULL AND (s.needs_review=1 OR j.status IN ('queued','running','failed'))
            """
        )
        args: list[object] = []
        if scope is not None:
            sql += " AND c.teacher_user_id=?"; args.append(scope)
        rows = conn.execute(sql + " ORDER BY s.submitted_at DESC", args).fetchall()
    items = []
    for row in rows:
        item = dict(row)
        result = json.loads(item.pop("result_json") or "{}")
        result_rows = result.get("results", []) or []
        item["review_count"] = sum(1 for r in result_rows if r.get("needs_review"))
        item["qwen_error"] = result.get("qwen_error", "")
        item["question_types"] = sorted({
            normalize_question_type(str(
                r.get("question_type")
                or (r.get("evidence") or {}).get("source", {}).get("question_type")
                or (r.get("qwen") or {}).get("question_type")
                or ("proof" if "证明/非计算题" in " ".join(str(x) for x in (r.get("review_reasons") or [])) else "calc")
            ))
            for r in result_rows
        })
        review_text = " ".join(
            str(reason) for r in result_rows
            for reason in [*(r.get("review_reasons") or []), *(r.get("risks") or [])]
        )
        item["risk_flags"] = {
            "qwen_error": bool(item["qwen_error"]),
            "low_confidence": any(float(r.get("confidence") or 0) < 0.85 for r in result_rows),
            "ocr_failure": "识别失败" in review_text or "未识别" in review_text,
            "proof": "proof" in item["question_types"],
            "needs_review": bool(item["review_count"]),
        }
        # Summarise the submission for teachers: stable items can be accepted
        # together; only exception items should demand detailed attention.
        global_blocker = bool(item["qwen_error"]) or item["grading_status"] != "completed"
        stable_items, exception_items = [], []
        for question in result_rows:
            confidence = float(question.get("confidence") or 0)
            risks = list(question.get("review_reasons") or []) + list(question.get("risks") or [])
            stable = (not global_blocker and not question.get("needs_review") and confidence >= 0.85)
            target = stable_items if stable else exception_items
            target.append({
                "sort_order": question.get("sort_order"),
                "confidence": confidence,
                "score": question.get("score"),
                "max_score": question.get("max_score"),
                "reasons": risks or (["模型调用异常"] if item["qwen_error"] else
                                      ["置信度不足"] if confidence < 0.85 else
                                      ["需要教师判断"] if question.get("needs_review") else
                                      ["初评尚未完成"]),
            })
        item["question_count"] = len(result_rows)
        item["stable_question_count"] = len(stable_items)
        item["exception_question_count"] = len(exception_items)
        item["exception_questions"] = exception_items
        item["can_confirm_stable_items"] = bool(stable_items) and not global_blocker
        item["priority"], item["priority_reasons"] = _review_priority(item["grading_status"], result)
        if class_id is not None and int(item["class_id"]) != class_id:
            continue
        if assignment_id is not None and int(item["assignment_id"]) != assignment_id:
            continue
        if question_type and question_type not in item["question_types"]:
            continue
        if risk and not item["risk_flags"].get(risk):
            continue
        items.append(item)
    return sorted(items, key=lambda item: (-item["priority"], item["submitted_at"] or ""))


@app.get("/api/reviews/batch-candidates")
def batch_review_candidates(request: Request):
    """Return only low-risk, not-yet-teacher-confirmed submissions for explicit batch review."""
    actor = require_roles(request, {"admin", "teacher"})
    scope = teacher_id_for_scope(actor)
    with connection() as conn:
        sql = """SELECT s.id,s.student_no,s.student_name,s.score,a.title,j.result_json
                 FROM submissions s JOIN assignments a ON a.id=s.assignment_id
                 JOIN classes c ON c.id=a.class_id JOIN grading_jobs j ON j.submission_id=s.id
                 LEFT JOIN grading_experiences ge ON ge.submission_id=s.id
                 WHERE a.class_id IS NOT NULL AND j.status='completed' AND ge.id IS NULL"""
        params: list[object] = []
        if scope is not None:
            sql += " AND c.teacher_user_id=?"; params.append(scope)
        rows = conn.execute(sql + " ORDER BY s.submitted_at DESC", params).fetchall()
    items = []
    for row in rows:
        result = json.loads(row["result_json"] or "{}")
        question_rows = result.get("results") or []
        reasons = []
        if result.get("qwen_error"):
            reasons.append("模型调用异常")
        if not question_rows:
            reasons.append("缺少逐题证据")
        for question in question_rows:
            if question.get("needs_review"):
                reasons.append(f"第 {question.get('sort_order', '?')} 题要求复核")
            if float(question.get("confidence") or 0) < 0.85:
                reasons.append(f"第 {question.get('sort_order', '?')} 题置信度不足")
        items.append({
            "submission_id": row["id"], "student_no": row["student_no"], "student_name": row["student_name"],
            "title": row["title"], "candidate_score": result.get("total_score", row["score"]),
            "max_score": result.get("max_score"), "eligible": not reasons,
            "blocked_reasons": list(dict.fromkeys(reasons)),
        })
    return {"items": items}


@app.post("/api/reviews/batch-confirm")
def batch_confirm_reviews(payload: BatchReviewIn, request: Request):
    """Teacher explicitly accepts selected low-risk candidates; never publishes them."""
    actor = require_roles(request, {"admin", "teacher"})
    scope = teacher_id_for_scope(actor)
    ids = list(dict.fromkeys(payload.submission_ids))
    confirmed, skipped = [], []
    with connection() as conn:
        for submission_id in ids:
            row = conn.execute(
                """SELECT s.id,s.assignment_id,s.student_name,s.student_no,c.teacher_user_id,j.result_json,j.status AS grading_status,
                          EXISTS(SELECT 1 FROM grading_experiences ge WHERE ge.submission_id=s.id) AS already_confirmed
                   FROM submissions s JOIN assignments a ON a.id=s.assignment_id JOIN classes c ON c.id=a.class_id
                   LEFT JOIN grading_jobs j ON j.submission_id=s.id WHERE s.id=?""",
                (submission_id,),
            ).fetchone()
            if not row or (scope is not None and row["teacher_user_id"] != scope):
                skipped.append({"submission_id": submission_id, "reason": "提交不存在或无权限"}); continue
            if row["already_confirmed"]:
                skipped.append({"submission_id": submission_id, "reason": "该提交已由教师确认"}); continue
            if row["grading_status"] != "completed":
                skipped.append({"submission_id": submission_id, "reason": "初评尚未完成"}); continue
            result = json.loads(row["result_json"] or "{}")
            question_rows = result.get("results") or []
            blocked = bool(result.get("qwen_error")) or not question_rows or any(
                question.get("needs_review") or float(question.get("confidence") or 0) < 0.85
                for question in question_rows
            )
            if blocked:
                skipped.append({"submission_id": submission_id, "reason": "存在低置信度、模型异常或待复核题目，请逐题确认"}); continue
            decisions = []
            for question in question_rows:
                candidate = float(question.get("score") or 0)
                question["teacher_decision"] = {
                    "score": candidate, "feedback": "", "confirmed_by": actor["id"],
                    "confirmed_at": datetime.now(timezone.utc).isoformat(), "batch_low_risk": True,
                }
                decisions.append({
                    "sort_order": question.get("sort_order"), "question_id": question.get("question_id"),
                    "candidate_score": candidate, "max_score": question.get("max_score"),
                    "confirmed_score": candidate, "teacher_feedback": "",
                })
            total = round(sum(float(question.get("score") or 0) for question in question_rows), 2)
            result["teacher_review"] = {
                "version": "batch_low_risk_v1", "confirmed_by": actor["id"],
                "confirmed_at": datetime.now(timezone.utc).isoformat(),
                "overall_feedback": "", "question_decisions": decisions, "computed_total_score": total,
            }
            evidence = json.dumps(result, ensure_ascii=False)
            decision_json = json.dumps(result["teacher_review"], ensure_ascii=False)
            conn.execute("UPDATE grading_jobs SET result_json=? WHERE submission_id=?", (evidence, submission_id))
            conn.execute("UPDATE submissions SET status='graded',score=?,needs_review=0 WHERE id=?", (total, submission_id))
            conn.execute("""INSERT INTO grading_experiences(submission_id,assignment_id,confirmed_score,teacher_feedback,evidence_json,decision_json)
                            VALUES(?,?,?,?,?,?)""", (submission_id, row["assignment_id"], total, "", evidence, decision_json))
            confirmed.append({"submission_id": submission_id, "student_name": row["student_name"] or row["student_no"], "score": total})
    for item in confirmed:
        audit(actor, "submission.batch_review", "submission", item["submission_id"], scope,
              {"score": item["score"], "mode": "batch_low_risk"})
    return {"confirmed": confirmed, "skipped": skipped, "message": f"已确认 {len(confirmed)} 份低风险初评；未自动发布给学生。"}


@app.post("/api/submissions/{submission_id}/confirm-stable-items")
def confirm_stable_items(submission_id: int, request: Request):
    """Accept only deterministic/high-confidence items; leave exceptions for review."""
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        row = conn.execute(
            """SELECT s.id,c.teacher_user_id,j.status AS grading_status,j.result_json
               FROM submissions s JOIN assignments a ON a.id=s.assignment_id
               JOIN classes c ON c.id=a.class_id
               LEFT JOIN grading_jobs j ON j.submission_id=s.id WHERE s.id=?""",
            (submission_id,),
        ).fetchone()
        if not row:
            raise HTTPException(404, "提交不存在")
        scope = teacher_id_for_scope(actor)
        if scope is not None and row["teacher_user_id"] != scope:
            raise HTTPException(404, "提交不存在")
        if row["grading_status"] != "completed":
            raise HTTPException(409, "初评尚未完成，暂不能确认稳定题")
        result = json.loads(row["result_json"] or "{}")
        if result.get("qwen_error"):
            raise HTTPException(409, "模型调用异常，不能自动确认任何题目")
        questions = result.get("results") or []
        confirmed_orders = []
        now = datetime.now(timezone.utc).isoformat()
        for question in questions:
            confidence = float(question.get("confidence") or 0)
            prior = question.get("teacher_decision") or {}
            stable = not question.get("needs_review") and confidence >= 0.85
            if stable and not prior.get("stable_auto_accept"):
                candidate = float(question.get("score") or 0)
                question["teacher_decision"] = {
                    "score": candidate, "feedback": prior.get("feedback", ""),
                    "confirmed_by": actor["id"], "confirmed_at": now,
                    "stable_auto_accept": True,
                }
                confirmed_orders.append(question.get("sort_order"))
        if not confirmed_orders:
            raise HTTPException(409, "没有可一键确认的稳定题；请处理异常小问")
        result["partial_teacher_review"] = {
            "version": "stable_items_v1", "confirmed_by": actor["id"],
            "confirmed_at": now, "confirmed_orders": confirmed_orders,
        }
        conn.execute("UPDATE grading_jobs SET result_json=? WHERE submission_id=?",
                     (json.dumps(result, ensure_ascii=False), submission_id))
    audit(actor, "submission.confirm_stable_items", "submission", submission_id, row["teacher_user_id"],
          {"confirmed_orders": confirmed_orders})
    return {"ok": True, "confirmed_count": len(confirmed_orders),
            "message": f"已确认 {len(confirmed_orders)} 道稳定题；其余异常小问仍待教师裁定，成绩尚未发布。"}
    

@app.get("/api/submissions/{submission_id}/grading")
def grading_evidence(submission_id: int, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        row = conn.execute(
            """SELECT s.*, j.status AS grading_status, j.result_json, a.title,c.teacher_user_id
               FROM submissions s JOIN assignments a ON a.id=s.assignment_id JOIN classes c ON c.id=a.class_id
               LEFT JOIN grading_jobs j ON j.submission_id=s.id WHERE s.id=?""",
            (submission_id,),
        ).fetchone()
    if not row:
        raise HTTPException(404, "提交不存在")
    scope = teacher_id_for_scope(actor)
    if scope is not None and row["teacher_user_id"] != scope:
        raise HTTPException(404, "提交不存在")
    payload = dict(row)
    payload["grading_result"] = json.loads(payload.pop("result_json") or "{}")
    payload["original_file_url"] = f"/api/submissions/{submission_id}/file"
    payload["original_file_name"] = Path(payload["file_path"]).name
    return payload


@app.post("/api/submissions/{submission_id}/regrade")
async def retry_grading(submission_id: int, background_tasks: BackgroundTasks, request: Request):
    """Retry AI grading of an already archived submission without another upload.

    A transient VLM/tunnel outage must never turn into a teacher asking a
    student to upload the same homework again.  The result is still review
    first; this endpoint only replaces the failed *candidate* result.
    """
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        submission = conn.execute(
            """SELECT s.id,s.assignment_id,c.teacher_user_id FROM submissions s
               JOIN assignments a ON a.id=s.assignment_id JOIN classes c ON c.id=a.class_id
               WHERE s.id=?""",
            (submission_id,),
        ).fetchone()
        if not submission:
            raise HTTPException(404, "提交不存在")
        scope = teacher_id_for_scope(actor)
        if scope is not None and submission["teacher_user_id"] != scope:
            raise HTTPException(404, "提交不存在")
        job = conn.execute("SELECT id,status FROM grading_jobs WHERE submission_id=?", (submission_id,)).fetchone()
        if not job:
            raise HTTPException(409, "该提交没有可重试的初评任务")
        if job["status"] in {"queued", "running"}:
            raise HTTPException(409, "该提交正在进行 AI 初评，请稍候")
        conn.execute("UPDATE grading_jobs SET status='queued', result_json=NULL WHERE id=?", (job["id"],))
        conn.execute(
            "UPDATE submissions SET status='submitted', score=NULL, feedback='正在重新调用 AI 初评。', needs_review=1, handwriting_score=NULL WHERE id=?",
            (submission_id,),
        )
    with connection() as conn:
        conn.execute("UPDATE grading_jobs SET rq_job_id=NULL,last_error=NULL WHERE id=?", (job["id"],))
    try:
        queue_info = dispatch_grading_job(job["id"], background_tasks)
    except QueueUnavailable as exc:
        queue_info = {"backend": "rq", "queue_warning": str(exc)}
    audit(actor, "submission.regrade", "submission", submission_id, submission["teacher_user_id"])
    return {"ok": True, "grading_job_id": job["id"], "queue": queue_info,
            "message": "已重新进入 AI 初评队列；原作业已保留，无需重新上传。"}


@app.get("/api/submissions/{submission_id}/file")
def original_submission_file(submission_id: int, request: Request):
    """Serve only the original file belonging to this submission for review."""
    with connection() as conn:
        row = conn.execute("""SELECT s.file_path,c.teacher_user_id FROM submissions s JOIN assignments a ON a.id=s.assignment_id
                            JOIN classes c ON c.id=a.class_id WHERE s.id=?""", (submission_id,)).fetchone()
    if not row:
        raise HTTPException(404, "提交不存在")
    actor = require_roles(request, {"admin", "teacher"})
    scope = teacher_id_for_scope(actor)
    if scope is not None and row["teacher_user_id"] != scope:
        raise HTTPException(404, "提交不存在")
    path = Path(row["file_path"]).resolve()
    upload_root = Path(settings.upload_dir).resolve()
    if not path.is_file() or upload_root not in path.parents:
        # Historical local-machine demo records may have been migrated without
        # their original binary. Return a teacher-friendly page rather than
        # exposing a raw FastAPI JSON error inside the review iframe.
        return HTMLResponse(
            """<!doctype html><meta charset='utf-8'><style>
            body{font-family:'Microsoft YaHei',sans-serif;margin:32px;color:#17324d;line-height:1.7}
            .box{max-width:620px;padding:24px;border:1px solid #fed7aa;background:#fff7ed;border-radius:12px}
            </style><div class='box'><h3>历史作业原件未迁移</h3>
            <p>这是一条旧的本地演示/历史提交记录。当前服务器保留了批改记录，但当时的图片或 PDF 没有同步到上传目录，因此无法预览。</p>
            <p>它不会影响现在的作业、成绩或之后的手写上传。若这份原件仍需要保留，请重新上传；否则可在演示记录清理中移除。</p></div>""",
            status_code=410,
        )
    return FileResponse(path, filename=path.name, content_disposition_type="inline")


def _submission_page_paths(file_path: str) -> list[Path]:
    """Return original image or cached PDF renders, never a user-supplied path."""
    path = Path(file_path).resolve()
    upload_root = Path(settings.upload_dir).resolve()
    if not path.is_file() or upload_root not in path.parents:
        raise HTTPException(404, "原作业文件不存在")
    if path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}:
        return [path]
    if path.suffix.lower() == ".pdf":
        from .grading_pipeline import _render_pdf_pages
        try:
            return _render_pdf_pages(path)
        except ValueError as exc:
            raise HTTPException(409, str(exc)) from exc
    raise HTTPException(409, "该文件格式暂不支持页图确认，请先上传 PDF 或图片")


def _teacher_submission_row(submission_id: int, actor: dict):
    with connection() as conn:
        row = conn.execute(
            """SELECT s.id,s.file_path,s.assignment_id,c.teacher_user_id,j.result_json
               FROM submissions s JOIN assignments a ON a.id=s.assignment_id JOIN classes c ON c.id=a.class_id
               LEFT JOIN grading_jobs j ON j.submission_id=s.id WHERE s.id=?""", (submission_id,)
        ).fetchone()
    if not row:
        raise HTTPException(404, "提交不存在")
    scope = teacher_id_for_scope(actor)
    if scope is not None and row["teacher_user_id"] != scope:
        raise HTTPException(404, "提交不存在")
    return row


@app.get("/api/submissions/{submission_id}/pages")
def submission_pages(submission_id: int, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    row = _teacher_submission_row(submission_id, actor)
    pages = _submission_page_paths(row["file_path"])
    return {"submission_id": submission_id, "page_count": len(pages),
            "pages": [{"page_no": index, "image_url": f"/api/submissions/{submission_id}/pages/{index}/image"} for index in range(1, len(pages) + 1)]}


@app.get("/api/submissions/{submission_id}/pages/{page_no}/image")
def submission_page_image(submission_id: int, page_no: int, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    row = _teacher_submission_row(submission_id, actor)
    pages = _submission_page_paths(row["file_path"])
    if page_no < 1 or page_no > len(pages):
        raise HTTPException(404, "页码不存在")
    return FileResponse(pages[page_no - 1], content_disposition_type="inline")


@app.get("/api/submissions/{submission_id}/question-regions")
def submission_question_regions(submission_id: int, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    _teacher_submission_row(submission_id, actor)
    with connection() as conn:
        rows = conn.execute(
            """SELECT question_id,subpart_no,sort_order,page_no,x,y,width,height,confirmed_at
               FROM submission_question_regions WHERE submission_id=? ORDER BY sort_order,subpart_no""",
            (submission_id,),
        ).fetchall()
    return {"submission_id": submission_id, "mappings": [dict(row) for row in rows]}


@app.post("/api/submissions/{submission_id}/question-regions")
def save_submission_question_regions(submission_id: int, payload: SubmissionRegionBatchIn, request: Request):
    """Persist teacher-confirmed page/region mappings alongside, never inside, the original upload."""
    actor = require_roles(request, {"admin", "teacher"})
    row = _teacher_submission_row(submission_id, actor)
    pages = _submission_page_paths(row["file_path"])
    result = json.loads(row["result_json"] or "{}")
    result_items = result.get("results") or []
    known = {(int(item.get("question_id") or 0), str(item.get("subpart_no") or ""), int(item.get("sort_order") or 0)) for item in result_items}
    seen = set()
    mappings = []
    for item in payload.mappings:
        key = (item.question_id, item.subpart_no or "", item.sort_order)
        if key in seen:
            raise HTTPException(422, "同一题目/小问重复提交映射")
        seen.add(key)
        if key not in known:
            raise HTTPException(422, "映射包含当前作业不存在的题目或小问")
        if item.page_no > len(pages) or item.x + item.width > 1.000001 or item.y + item.height > 1.000001:
            raise HTTPException(422, "页码或裁切区域超出作业页面范围")
        mappings.append(item.model_dump())
    with connection() as conn:
        conn.execute("DELETE FROM submission_question_regions WHERE submission_id=?", (submission_id,))
        now = datetime.now(timezone.utc).isoformat()
        conn.executemany(
            """INSERT INTO submission_question_regions(submission_id,question_id,subpart_no,sort_order,page_no,x,y,width,height,confirmed_by,confirmed_at)
               VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
            [(submission_id, item["question_id"], item["subpart_no"] or "", item["sort_order"], item["page_no"],
              item["x"], item["y"], item["width"], item["height"], actor["id"], now) for item in mappings],
        )
        by_key = {(item["question_id"], item["subpart_no"] or "", item["sort_order"]): item for item in mappings}
        for item in result_items:
            key = (int(item.get("question_id") or 0), str(item.get("subpart_no") or ""), int(item.get("sort_order") or 0))
            if key in by_key:
                item["teacher_page_mapping"] = by_key[key]
        result["teacher_page_mapping_confirmed_at"] = now
        conn.execute("UPDATE grading_jobs SET result_json=? WHERE submission_id=?", (json.dumps(result, ensure_ascii=False), submission_id))
    audit(actor, "submission.question_regions", "submission", submission_id, row["teacher_user_id"],
          {"mapping_count": len(mappings)})
    return {"ok": True, "mapping_count": len(mappings), "message": "题号/小问与页图区域的教师确认已保存。后续重跑初评仍会保留这些映射。"}


@app.post("/api/submissions/{submission_id}/review")
def confirm_review(submission_id: int, decision: ReviewDecisionIn, request: Request):
    """Persist a teacher's per-question decision while preserving AI evidence."""
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        submission = conn.execute("""SELECT s.*,c.teacher_user_id FROM submissions s JOIN assignments a ON a.id=s.assignment_id
                                   JOIN classes c ON c.id=a.class_id WHERE s.id=?""", (submission_id,)).fetchone()
        if not submission:
            raise HTTPException(404, "提交不存在")
        scope = teacher_id_for_scope(actor)
        if scope is not None and submission["teacher_user_id"] != scope:
            raise HTTPException(404, "提交不存在")
        job = conn.execute("SELECT result_json FROM grading_jobs WHERE submission_id=?", (submission_id,)).fetchone()
        result = json.loads((job["result_json"] if job else "") or "{}")
        rows = result.get("results") or []
        if not rows:
            raise HTTPException(409, "尚无逐题初评证据，不能确认成绩")

        provided: dict[int, QuestionReviewDecisionIn] = {}
        for item in decision.question_decisions:
            if item.sort_order in provided:
                raise HTTPException(422, f"第 {item.sort_order} 题重复提交裁定")
            provided[item.sort_order] = item
        known_orders = {int(row.get("sort_order") or 0) for row in rows}
        unknown_orders = set(provided) - known_orders
        if unknown_orders:
            raise HTTPException(422, "裁定中包含不存在的题目序号")

        final_items, decision_items = [], []
        for row in rows:
            item = dict(row)
            order = int(item.get("sort_order") or 0)
            max_score = float(item.get("max_score") or 0)
            candidate_score = float(item.get("score") or 0)
            selected = provided.get(order)
            if selected:
                if selected.score > max_score + 1e-9:
                    raise HTTPException(422, f"第 {order} 题确认分不能超过满分 {max_score:g}")
                final_score = selected.score
                final_feedback = selected.feedback.strip()
                item["teacher_decision"] = {
                    "score": final_score, "feedback": final_feedback,
                    "confirmed_by": actor["id"], "confirmed_at": datetime.now(timezone.utc).isoformat(),
                }
                item["score"] = final_score
                if final_feedback:
                    item["feedback"] = final_feedback
                decision_items.append({
                    "sort_order": order, "question_id": item.get("question_id"),
                    "candidate_score": candidate_score, "max_score": max_score,
                    "confirmed_score": final_score, "teacher_feedback": final_feedback,
                })
            elif (prior_decision := item.get("teacher_decision") or {}).get("stable_auto_accept"):
                final_score = float(prior_decision.get("score", candidate_score))
                item["score"] = final_score
                decision_items.append({
                    "sort_order": order, "question_id": item.get("question_id"),
                    "candidate_score": candidate_score, "max_score": max_score,
                    "confirmed_score": final_score,
                    "teacher_feedback": str(prior_decision.get("feedback") or ""),
                    "stable_auto_accept": True,
                })
            else:
                final_score = candidate_score
            final_items.append(item)

        # A legacy call may still set only a total score. The new path derives
        # it from each controlled item so a client cannot smuggle an inconsistent total.
        computed_total = round(sum(float(item.get("score") or 0) for item in final_items), 2)
        if not provided:
            if decision.score is None:
                raise HTTPException(422, "请填写逐题裁定，或提供总分")
            max_total = sum(float(item.get("max_score") or 0) for item in final_items)
            if decision.score > max_total + 1e-9:
                raise HTTPException(422, f"确认总分不能超过满分 {max_total:g}")
            computed_total = decision.score
        assignment_score = conn.execute(
            "SELECT total_score,score_policy,completion_points FROM assignments WHERE id=?",
            (submission["assignment_id"],),
        ).fetchone()
        completion_score = float(assignment_score["completion_points"] or 0)
        score_policy = str(assignment_score["score_policy"] or "legacy")
        final_total = round(computed_total + completion_score, 2)
        result["results"] = final_items
        result["quality_score"] = computed_total
        result["quality_max_score"] = round(sum(float(item.get("max_score") or 0) for item in final_items), 2)
        result["completion_score"] = completion_score
        result["score_policy"] = score_policy
        result["total_score"] = final_total
        result["max_score"] = float(assignment_score["total_score"] or result["quality_max_score"])
        result["teacher_review"] = {
            "version": "per_question_v1", "confirmed_by": actor["id"],
            "confirmed_at": datetime.now(timezone.utc).isoformat(),
            "overall_feedback": decision.feedback.strip(),
            "question_decisions": decision_items,
            "computed_quality_score": computed_total,
            "completion_score": completion_score,
            "computed_total_score": final_total,
        }
        evidence = json.dumps(result, ensure_ascii=False)
        decision_json = json.dumps(result["teacher_review"], ensure_ascii=False)
        conn.execute("UPDATE grading_jobs SET result_json=? WHERE submission_id=?", (evidence, submission_id))
        conn.execute(
            "UPDATE submissions SET status='graded', score=?, feedback=?, needs_review=0, completion_score=?, quality_score=?, quality_max_score=?, score_policy_version=? WHERE id=?",
            (final_total, decision.feedback.strip(), completion_score, computed_total, result["quality_max_score"], score_policy, submission_id),
        )
        conn.execute(
            """INSERT INTO grading_experiences
               (submission_id,assignment_id,confirmed_score,teacher_feedback,evidence_json,decision_json)
               VALUES(?,?,?,?,?,?)""",
            (submission_id, submission["assignment_id"], final_total, decision.feedback.strip(), evidence, decision_json),
        )
    audit(actor, "submission.review", "submission", submission_id, submission["teacher_user_id"],
          {"score": final_total, "quality_score": computed_total, "question_decision_count": len(decision_items)})
    return {
        "ok": True, "score": final_total, "quality_score": computed_total, "question_decision_count": len(decision_items),
        "message": "教师逐题裁定已确认，原始初评与教师修改均已沉淀为可评测证据。",
    }

REVIEW_QUOTA_PER_CLASS = 2  # 设计文档要求：每班不少于 2 次人工复核


@app.get("/api/reports/grading-deviations")
def grading_deviations(request: Request):
    """Compare AI candidate scores with structured teacher decisions."""
    actor = require_roles(request, {"admin", "teacher"})
    scope = teacher_id_for_scope(actor)
    with connection() as conn:
        sql = """SELECT ge.submission_id,ge.evidence_json,ge.decision_json,a.title
                 FROM grading_experiences ge JOIN submissions s ON s.id=ge.submission_id
                 JOIN assignments a ON a.id=s.assignment_id JOIN classes c ON c.id=a.class_id"""
        params: list[object] = []
        if scope is not None:
            sql += " WHERE c.teacher_user_id=?"; params.append(scope)
        rows = conn.execute(sql, params).fetchall()
        types = {row["id"]: row["question_type"] for row in conn.execute("SELECT id,question_type FROM questions")}
    changes = []
    for row in rows:
        decision = json.loads(row["decision_json"] or "{}")
        for item in decision.get("question_decisions") or []:
            candidate = float(item.get("candidate_score") or 0)
            confirmed = float(item.get("confirmed_score") or 0)
            delta = round(confirmed - candidate, 2)
            changes.append({
                "submission_id": row["submission_id"], "title": row["title"], "sort_order": item.get("sort_order"),
                "question_id": item.get("question_id"), "question_type": types.get(item.get("question_id"), "未分类"),
                "candidate_score": candidate, "confirmed_score": confirmed, "delta": delta,
            })
    changed = [item for item in changes if abs(item["delta"]) > 1e-9]
    by_type = {}
    for kind in sorted({item["question_type"] for item in changes}):
        rows_for_type = [item for item in changes if item["question_type"] == kind]
        by_type[kind] = {
            "item_count": len(rows_for_type),
            "changed_count": sum(1 for item in rows_for_type if abs(item["delta"]) > 1e-9),
            "change_rate": round(sum(1 for item in rows_for_type if abs(item["delta"]) > 1e-9) / len(rows_for_type) * 100, 1) if rows_for_type else 0.0,
            "average_delta": round(sum(item["delta"] for item in rows_for_type) / len(rows_for_type), 2) if rows_for_type else 0.0,
        }
    return {
        "reviewed_submission_count": len(rows), "reviewed_item_count": len(changes),
        "changed_item_count": len(changed),
        "change_rate": round(len(changed) / len(changes) * 100, 1) if changes else 0.0,
        "average_absolute_delta": round(sum(abs(item["delta"]) for item in changes) / len(changes), 2) if changes else 0.0,
        "by_question_type": by_type,
        "largest_changes": sorted(changed, key=lambda item: abs(item["delta"]), reverse=True)[:10],
    }


@app.get("/api/reports/review-quota")
def review_quota(request: Request):
    """人工复核配额：每个班级（按学期）已完成的教师复核次数；少于配额则提示。"""
    actor = require_roles(request, {"admin", "teacher"})
    scope = teacher_id_for_scope(actor)
    with connection() as conn:
        sql = (
            """SELECT a.class_name, a.semester, COUNT(DISTINCT ge.id) AS reviewed
               FROM grading_experiences ge
               JOIN submissions s ON s.id=ge.submission_id
               JOIN assignments a ON a.id=ge.assignment_id
               JOIN classes c ON c.id=a.class_id
               WHERE COALESCE(a.is_demo,0)=0
            """
        )
        args: list[object] = []
        if scope is not None: sql += " AND c.teacher_user_id=?"; args.append(scope)
        rows = conn.execute(sql + " GROUP BY a.class_name, a.semester", args).fetchall()
    items = []
    for r in rows:
        reviewed = r["reviewed"]
        items.append({
            "class_name": r["class_name"], "semester": r["semester"],
            "teacher_reviews": reviewed, "quota": REVIEW_QUOTA_PER_CLASS,
            "meets_quota": reviewed >= REVIEW_QUOTA_PER_CLASS,
        })
    return {"quota_per_class": REVIEW_QUOTA_PER_CLASS, "classes": items,
            "classes_below_quota": [i for i in items if not i["meets_quota"]]}


@app.get("/api/reviews/batch-release-candidates")
def batch_release_candidates(request: Request):
    """Only teacher-confirmed, graded, not-yet-released work may be batch released."""
    actor=require_roles(request,{"admin","teacher"}); scope=teacher_id_for_scope(actor)
    with connection() as conn:
        sql="""SELECT s.id AS submission_id,s.student_no,s.student_name,s.score,a.title,
                     a.class_name,COUNT(*) OVER (PARTITION BY a.id) AS assignment_submission_count
              FROM submissions s JOIN assignments a ON a.id=s.assignment_id
              JOIN classes c ON c.id=a.class_id JOIN grading_experiences ge ON ge.submission_id=s.id
              WHERE s.status='graded' AND s.released_at IS NULL"""
        args=[]
        if scope is not None: sql+=" AND c.teacher_user_id=?";args.append(scope)
        rows=conn.execute(sql+" ORDER BY a.id DESC,s.id DESC",args).fetchall()
    return {"items":[dict(x) for x in rows]}


@app.post("/api/reviews/batch-release")
def batch_release(payload: BatchReviewIn, request: Request):
    actor=require_roles(request,{"admin","teacher"});scope=teacher_id_for_scope(actor)
    released=[];skipped=[];now=datetime.now(timezone.utc).isoformat()
    with connection() as conn:
        for submission_id in list(dict.fromkeys(payload.submission_ids)):
            row=conn.execute("""SELECT s.id,s.status,s.released_at,c.teacher_user_id,
                       EXISTS(SELECT 1 FROM grading_experiences ge WHERE ge.submission_id=s.id) AS confirmed
                       FROM submissions s JOIN assignments a ON a.id=s.assignment_id JOIN classes c ON c.id=a.class_id
                       WHERE s.id=?""",(submission_id,)).fetchone()
            if not row or (scope is not None and row["teacher_user_id"]!=scope):
                skipped.append({"submission_id":submission_id,"reason":"提交不存在或无权限"});continue
            if row["released_at"]: skipped.append({"submission_id":submission_id,"reason":"成绩已发布"});continue
            if row["status"]!="graded" or not row["confirmed"]:
                skipped.append({"submission_id":submission_id,"reason":"尚未完成教师确认，不能发布"});continue
            conn.execute("UPDATE submissions SET released_at=? WHERE id=?",(now,submission_id))
            released.append(submission_id)
    for submission_id in released:
        audit(actor,"submission.batch_release","submission",submission_id,None,{})
    return {"released":released,"skipped":skipped,"message":f"已向学生发布 {len(released)} 份确认后的成绩与错题复盘。"}

@app.post("/api/submissions/{submission_id}/release")
def release_submission(submission_id: int, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    with connection() as conn:
        row = conn.execute("SELECT s.status,c.teacher_user_id FROM submissions s JOIN assignments a ON a.id=s.assignment_id JOIN classes c ON c.id=a.class_id WHERE s.id=?", (submission_id,)).fetchone()
        if not row:
            raise HTTPException(404, "提交不存在")
        scope = teacher_id_for_scope(actor)
        if scope is not None and row["teacher_user_id"] != scope:
            raise HTTPException(404, "提交不存在")
        if row["status"] != "graded":
            raise HTTPException(409, "请先完成教师复核，再发布给学生")
        confirmed = conn.execute("SELECT 1 FROM grading_experiences WHERE submission_id=? LIMIT 1", (submission_id,)).fetchone()
        if not confirmed:
            raise HTTPException(409, "请先逐题确认，或在“批量确认低风险初评”中留下教师裁定后再发布")
        conn.execute("UPDATE submissions SET released_at=? WHERE id=?", (datetime.now(timezone.utc).isoformat(), submission_id))
    return {"ok": True, "message": "成绩与错题复盘已发布给学生"}


@app.get("/api/student/submissions/{submission_id}/file")
def student_submission_file(submission_id: int, request: Request):
    """Let a student reopen only their own archived upload."""
    actor = require_roles(request, {"student"})
    with connection() as conn:
        row = conn.execute(
            """SELECT s.file_path FROM submissions s
               JOIN assignments a ON a.id=s.assignment_id
               JOIN students st ON st.class_id=a.class_id AND st.student_no=s.student_no
               WHERE s.id=? AND st.user_id=?""",
            (submission_id, actor["id"]),
        ).fetchone()
    if not row:
        raise HTTPException(404, "未找到你的提交记录")
    path = Path(row["file_path"])
    if not path.is_file():
        raise HTTPException(404, "原作业文件暂不可用，请联系教师")
    return FileResponse(path, filename=path.name)


@app.get("/api/student/released-submissions")
def student_released_submissions(request: Request):
    actor = require_roles(request, {"student"})
    with connection() as conn:
        rows = conn.execute("SELECT s.id,s.assignment_id,s.score,s.completion_score,s.quality_score,s.quality_max_score,s.feedback,s.released_at,a.title,a.due_at,a.total_score,j.result_json FROM submissions s JOIN assignments a ON a.id=s.assignment_id JOIN students st ON st.class_id=a.class_id AND st.user_id=? LEFT JOIN grading_jobs j ON j.submission_id=s.id WHERE s.status=? AND s.released_at IS NOT NULL ORDER BY s.released_at DESC", (actor["id"], "graded")).fetchall()
    return [{"id": r["id"], "assignment_id": r["assignment_id"], "title": r["title"], "score": r["score"], "completion_score": r["completion_score"], "quality_score": r["quality_score"], "quality_max_score": r["quality_max_score"], "max_score": r["total_score"], "feedback": r["feedback"], "released_at": r["released_at"], "grading": json.loads(r["result_json"] or "{}")} for r in rows]



@app.get("/api/student/released-submissions/{submission_id}/mistakes")
def student_released_mistakes(submission_id: int, request: Request):
    actor = require_roles(request, {"student"})
    with connection() as conn:
        row = conn.execute("SELECT j.result_json FROM submissions s JOIN assignments a ON a.id=s.assignment_id JOIN students st ON st.class_id=a.class_id AND st.user_id=? LEFT JOIN grading_jobs j ON j.submission_id=s.id WHERE s.id=? AND s.status=? AND s.released_at IS NOT NULL", (actor["id"], submission_id, "graded")).fetchone()
        if not row:
            raise HTTPException(404, "未找到已发布的作业结果")
        results = json.loads(row["result_json"] or "{}").get("results") or []
        ids = [int(item["question_id"]) for item in results if item.get("question_id") is not None]
        qrows = conn.execute("SELECT id,source_problem_id,content FROM questions WHERE id IN (" + ",".join("?" for _ in ids) + ")", ids).fetchall() if ids else []
    questions = {r["id"]: dict(r) for r in qrows}
    mistakes = []
    for item in results:
        if item.get("correct") is True and float(item.get("score") or 0) >= float(item.get("max_score") or 0):
            continue
        q = questions.get(item.get("question_id"), {})
        mistakes.append({"question_id": item.get("question_id"), "source_problem_id": q.get("source_problem_id"), "content": q.get("content"), "recognized_work": item.get("recognized_work"), "feedback": item.get("feedback"), "score": item.get("score"), "max_score": item.get("max_score"), "needs_review": item.get("needs_review")})
    return {"submission_id": submission_id, "mistakes": mistakes}



@app.get("/api/reports/weak-points")
def weak_points(request: Request, class_name: str | None = None, semester: str | None = None, top: int = 10):
    """薄弱知识点建议：按知识点（无标签时回退章节）聚合得分率，低于阈值给出补习建议。"""
    actor = require_roles(request, {"admin", "teacher"})
    scope = teacher_id_for_scope(actor)
    with connection() as conn:
        qmap = {row["id"]: (row["knowledge_points"] or row["chapter"])
                for row in conn.execute("SELECT id, knowledge_points, chapter FROM questions")}
        clauses, params = ["a.class_id IS NOT NULL", "COALESCE(a.is_demo,0)=0"], []
        if class_name:
            clauses.append("a.class_name=?"); params.append(class_name)
        if semester:
            clauses.append("a.semester=?"); params.append(semester)
        if scope is not None:
            clauses.append("c.teacher_user_id=?"); params.append(scope)
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        rows = conn.execute(
            f"""SELECT ge.evidence_json FROM grading_experiences ge
                JOIN submissions s ON s.id=ge.submission_id
                JOIN assignments a ON a.id=ge.assignment_id
                JOIN classes c ON c.id=a.class_id {where}""",
            params,
        ).fetchall()
    agg: dict[str, dict] = {}
    for r in rows:
        ev = json.loads(r["evidence_json"] or "{}")
        for item in ev.get("results", []):
            qid = int(item.get("question_id") or 0)
            kp = qmap.get(qid)
            if not kp:
                continue
            max_s = float(item.get("max_score") or 0) or 1.0
            sc = float(item.get("score") or 0)
            bucket = agg.setdefault(kp, {"kp": kp, "score_sum": 0.0, "max_sum": 0.0, "count": 0})
            bucket["score_sum"] += sc
            bucket["max_sum"] += max_s
            bucket["count"] += 1
    points = []
    for b in agg.values():
        rate = round(b["score_sum"] / b["max_sum"], 3) if b["max_sum"] else 0
        points.append({"knowledge_point": b["kp"], "avg_rate": rate,
                       "sample_count": b["count"],
                       "suggestion": ("建议加强练习与错题重做" if rate < 0.7 else "掌握良好")})
    points.sort(key=lambda x: x["avg_rate"])
    return {"weak_points": points[:top], "threshold": 0.7}


@app.get("/api/reports/semester-summary")
def semester_summary(request: Request, class_name: str | None = None, semester: str | None = None):
    """学期末分数汇总：按学生聚合总分/平均分/排名，含班级均分、分布与书写整洁度。"""
    actor = require_roles(request, {"admin", "teacher"})
    scope = teacher_id_for_scope(actor)
    with connection() as conn:
        clauses = ["s.status='graded'", "a.class_id IS NOT NULL", "COALESCE(a.is_demo,0)=0"]
        params = []
        if class_name:
            clauses.append("a.class_name=?"); params.append(class_name)
        if semester:
            clauses.append("a.semester=?"); params.append(semester)
        if scope is not None:
            clauses.append("c.teacher_user_id=?"); params.append(scope)
        where = " AND ".join(clauses)
        rows = conn.execute(
            f"""SELECT s.student_no, s.student_name, s.score, s.handwriting_score,
                       s.completion_score, s.quality_score, s.quality_max_score,
                       a.id AS assignment_id, a.title AS assignment_title,
                       a.class_name, a.semester, a.total_score
                FROM submissions s JOIN assignments a ON a.id=s.assignment_id
                JOIN classes c ON c.id=a.class_id
                WHERE {where}""",
            params,
        ).fetchall()
    students: dict[str, dict] = {}
    for r in rows:
        no = r["student_no"]
        st = students.setdefault(no, {"student_no": no, "student_name": r["student_name"],
                                       "score_sum": 0.0, "max_sum": 0.0, "count": 0, "hw_scores": [], "records": []})
        # New assignments already use 100.  Normalising legacy work here makes
        # a semester average comparable even while teachers keep old records.
        source_max = float(r["total_score"] or 0)
        normalized = round((float(r["score"] or 0) / source_max * 100), 2) if source_max else 0.0
        st["score_sum"] += normalized
        st["max_sum"] += 100.0
        st["count"] += 1
        st["records"].append({"assignment_id": r["assignment_id"], "title": r["assignment_title"],
                              "score_100": normalized, "raw_score": r["score"], "raw_max": source_max,
                              "completion_score": r["completion_score"], "quality_score": r["quality_score"],
                              "quality_max_score": r["quality_max_score"]})
        if r["handwriting_score"] is not None:
            st["hw_scores"].append(float(r["handwriting_score"]))
    for st in students.values():
        st["average"] = round(st["score_sum"] / st["count"], 1) if st["count"] else 0
        st["avg_rate"] = round(st["score_sum"] / st["max_sum"], 3) if st["max_sum"] else 0
        st["handwriting_avg"] = round(sum(st["hw_scores"]) / len(st["hw_scores"]), 1) if st["hw_scores"] else None
        st.pop("hw_scores", None)
        st["records"].sort(key=lambda x: x["assignment_id"], reverse=True)
    ranked = sorted(students.values(), key=lambda x: x["average"], reverse=True)
    for i, st in enumerate(ranked, 1):
        st["rank"] = i
    avg_class = round(sum(s["average"] for s in ranked) / len(ranked), 1) if ranked else 0
    dist = {"优秀": 0, "良好": 0, "及格": 0, "不及格": 0}
    for s in ranked:
        a = s["average"]
        dist["优秀" if a >= 90 else "良好" if a >= 75 else "及格" if a >= 60 else "不及格"] += 1
    return {"student_count": len(ranked), "class_average": avg_class,
            "distribution": dist, "students": ranked}


@app.get("/api/reports/summary")
async def summary(request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    scope = teacher_id_for_scope(actor)
    with connection() as conn:
        suffix, args = ("", []) if scope is None else (" AND class_id IN (SELECT id FROM classes WHERE teacher_user_id=?)", [scope])
        demo_suffix = " AND COALESCE(is_demo,0)=0"
        local = dict(conn.execute(f"""SELECT (SELECT COUNT(*) FROM questions) question_count,
          (SELECT COUNT(*) FROM assignments WHERE status='published' AND class_id IS NOT NULL{demo_suffix}{suffix}) assignment_count,
          (SELECT COUNT(*) FROM submissions s JOIN assignments a ON a.id=s.assignment_id WHERE a.class_id IS NOT NULL AND COALESCE(a.is_demo,0)=0{suffix}) submission_count,
          (SELECT COUNT(*) FROM grading_jobs WHERE status='queued') review_queue""", args * 2).fetchone())
        hw_where = "a.class_id IS NOT NULL AND COALESCE(a.is_demo,0)=0 AND s.handwriting_score IS NOT NULL" + (" AND c.teacher_user_id=?" if scope is not None else "")
        hw = conn.execute(
            """SELECT AVG(s.handwriting_score) FROM submissions s
               JOIN assignments a ON a.id=s.assignment_id
               JOIN classes c ON c.id=a.class_id WHERE """ + hw_where, args).fetchone()[0]
        below_where = "" if scope is None else " WHERE c.teacher_user_id=?"
        below = conn.execute(
            """SELECT COUNT(*) FROM (SELECT a.class_name, a.semester
                FROM grading_experiences ge JOIN submissions s ON s.id=ge.submission_id
                JOIN assignments a ON a.id=ge.assignment_id JOIN classes c ON c.id=a.class_id WHERE COALESCE(a.is_demo,0)=0""" + (" AND c.teacher_user_id=?" if scope is not None else "") +
                " GROUP BY a.class_name, a.semester HAVING COUNT(DISTINCT ge.id) < ?)",
            args + [REVIEW_QUOTA_PER_CLASS]).fetchone()[0]
    evidence = await evidence_status()
    return {
        **local,
        "knowledge_total": evidence["problem_count"] if evidence["connected"] else None,
        "synced_cache_count": local["question_count"],
        "handwriting_avg": round(hw, 1) if hw is not None else None,
        "classes_below_review_quota": below,
        "review_quota_per_class": REVIEW_QUOTA_PER_CLASS,
    }


@app.get("/api/agent/capabilities")
def agent_capabilities():
    """Capabilities exposed to the teacher-facing agent shell."""
    return {
        "agent": "高数教材智能体 v1",
        "tasks": [
            {"id": "assignment", "name": "智能生成作业", "status": "available",
             "description": "按章节、难度和题量从题库组卷，生成留白作业单。"},
            {"id": "publish", "name": "发布与收作业", "status": "available",
             "description": "发布给班级，保留学生照片、PDF 和手写文件提交入口。"},
            {"id": "grading", "name": "智能批改", "status": "review-first",
             "description": "按评分点初评；低置信度或主观题必须转教师复核。"},
            {"id": "material", "name": "教材资料库", "status": "backend",
             "description": "检索教材、答案 PDF 与教师确认经验，不要求先建完整答案库。"},
            {"id": "answer_import", "name": "MinerU 答案导入审核", "status": "available",
             "description": "审核 MinerU 解析的答案书，确认后写回 8014 并同步到本地缓存。"},
            {"id": "reports", "name": "教学报表", "status": "available",
             "description": "复核配额、薄弱知识点建议、学期末分数汇总与书写整洁度维度，辅助教学决策。"},
            {"id": "ai_stem", "name": "AI 题干候选复核", "status": "review-first",
             "description": "VLM 已识别但双模型校验未过的题干，教师勾选确认后写回 8014 并进入本地缓存。"},
        ],
        "evidence_policy": "模型结果仅作候选；涉及成绩发布时保留评分证据与教师复核入口。",
    }


@app.get("/api/agent/knowledge-status")
async def knowledge_status():
    """Expose the 8014 evidence library health to the teacher portal."""
    return await evidence_status()


@app.get("/api/agent/retrieve-section")
async def retrieve_section(section_no: str, limit: int = 30):
    """Evidence packet used by the assignment and grading agents."""
    if not section_no.strip():
        raise HTTPException(422, "请输入章节编号，例如 1.1")
    return await retrieve_section_problems(section_no.strip(), max(1, min(limit, 80)))


class LearningAnswerIn(BaseModel):
    student_answer: str = Field(min_length=1, max_length=5000)
    mode: str = "diagnose"


@app.get("/api/learning/problems")
async def learning_problems():
    """Return only student-safe problem fields from the authoritative 8014 library."""
    try:
        async with evidence_client(timeout=20) as client:
            response = await client.get(evidence_url("").rstrip("/") + "/problems", params={"size": 500})
            response.raise_for_status()
        items = response.json().get("items", [])
    except httpx.HTTPError as exc:
        raise HTTPException(502, "学习题库暂不可用") from exc
    safe_items = [item for item in items if item.get("answer_status") == "verified" and not str(item.get("answer_invalid_reason") or "").strip() and len(str(item.get("content_text") or "").strip()) >= 12]
    return [{"id": item.get("id"), "section_no": item.get("section_no"), "problem_no": item.get("problem_no"), "sub_no": item.get("sub_no"), "ptype": item.get("ptype"), "difficulty": item.get("difficulty"), "content_text": item.get("content_text")} for item in safe_items]


def _candidate_quality_gate(item: dict) -> dict:
    """Read-only quality gate. Passing means eligible for teacher confirmation, never auto-publish."""
    reasons = []
    if item.get("match_status") != "pending":
        reasons.append("候选已处理")
    if len(str(item.get("content_text") or "").strip()) < 12:
        reasons.append("题干过短或缺失")
    if int(item.get("source_image_count") or 0) < 1:
        reasons.append("缺少答案来源图片")
    try:
        ai = json.loads(item.get("ai_review_json") or "{}")
    except Exception:
        ai = {}
    if item.get("ai_review_status") != "completed":
        reasons.append("尚未完成 AI 核验")
    confidence = float(ai.get("confidence") or 0)
    if item.get("ai_review_status") == "completed" and confidence < 0.92:
        reasons.append(f"AI 置信度不足（{confidence:.2f} < 0.92）")
    if ai.get("risks"):
        reasons.append("AI 标记风险：" + "；".join(str(x) for x in ai["risks"][:3]))
    answer = str(ai.get("std_answer") or "").strip()
    if item.get("ai_review_status") == "completed" and not answer:
        reasons.append("AI 未给出标准答案")
    if len(answer) > 3000:
        reasons.append("标准答案过长，疑似 OCR 串题")
    return {
        "can_publish_after_teacher_confirmation": not reasons,
        "recommended_action": "teacher_approve" if not reasons else (
            "run_ai_review" if item.get("ai_review_status") != "completed" else "fix_or_reject"
        ),
        "reasons": reasons,
        "ai_confidence": confidence,
    }


class CandidatePublishIn(BaseModel):
    confirm: bool = False
    note: str = ""
    content_text: str | None = None
    std_answer: str | None = None
    full_solution: str | None = None
    ptype: str | None = None
    manual_correction: bool = False


async def _pending_candidate(candidate_id: int) -> dict:
    # A repair-batch item can already have been published once (or hidden by
    # candidate de-duplication).  Always resolve its frozen ID directly rather
    # than searching only the mutable pending queue.
    async with evidence_client(timeout=30) as client:
        response = await client.get(
            evidence_url("").rstrip("/") + f"/answer-import-candidates/{candidate_id}"
        )
    if response.status_code == 404:
        raise HTTPException(404, "未找到该题库候选")
    if response.status_code >= 400:
        raise HTTPException(502, "题库候选服务暂不可用")
    return response.json()


@app.get("/api/teacher/question-bank/review-queue")
async def teacher_question_bank_review_queue(request: Request):
    """Authenticated teacher queue; candidate OCR/AI data never reaches students."""
    require_roles(request, {"admin", "teacher"})
    async with evidence_client(timeout=30) as client:
        response = await client.get(evidence_url("").rstrip("/") + "/answer-import-candidates", params={"status": "pending", "limit": 300})
        coverage = await client.get(evidence_url("").rstrip("/") + "/answer-library/coverage")
    if response.status_code >= 400:
        raise HTTPException(502, "题库候选服务暂不可用")
    items = response.json().get("items", [])
    safe_items = []
    for item in items:
        gate = _candidate_quality_gate(item)
        try:
            ai = json.loads(item.get("ai_review_json") or "{}")
        except Exception:
            ai = {}
        safe_items.append({
            "id": item.get("id"), "section_no": item.get("section_no"), "problem_no": item.get("problem_no"),
            "content_text": item.get("content_text"), "ptype": item.get("ptype"),
            "source_image_count": item.get("source_image_count"), "ai_review_status": item.get("ai_review_status"),
            "ai_suggestion": {"std_answer": ai.get("std_answer"), "full_solution": ai.get("full_solution"),
                              "ptype": ai.get("ptype"), "confidence": ai.get("confidence"), "risks": ai.get("risks") or []},
            "quality_gate": gate,
        })
    return {
        "summary": {
            "pending_candidates": len(safe_items),
            "ready_for_teacher_approval": sum(x["quality_gate"]["can_publish_after_teacher_confirmation"] for x in safe_items),
            "needs_ai_review": sum(x["quality_gate"]["recommended_action"] == "run_ai_review" for x in safe_items),
            "needs_fix_or_reject": sum(x["quality_gate"]["recommended_action"] == "fix_or_reject" for x in safe_items),
            "coverage": coverage.json() if coverage.status_code < 400 else {},
        },
        "items": safe_items,
    }


class CandidateBatchStartIn(BaseModel):
    section_no: str
    images_base64: list[str] = []
    filename: str = "teacher-answer-upload"


@app.post("/api/teacher/question-bank/create-candidates")
async def teacher_create_candidates(req: CandidateBatchStartIn, request: Request):
    require_roles(request, {"admin", "teacher"})
    if not req.section_no.strip() or not req.images_base64:
        raise HTTPException(422, "请选择章节并上传至少一张答案页图片")
    if len(req.images_base64) > 30:
        raise HTTPException(413, "一次最多上传 30 张图片")
    payload = {"section_no": req.section_no.strip(), "images_base64": req.images_base64,
               "filename": req.filename or "teacher-answer-upload", "only_unverified": True}
    async with evidence_client(timeout=120) as client:
        response = await client.post(evidence_url("").rstrip("/") + "/answer-library/section-batch", json=payload)
    if response.status_code >= 400:
        raise HTTPException(response.status_code, response.json().get("detail", "创建候选失败"))
    return response.json()


@app.post("/api/teacher/question-bank/candidates/{candidate_id}/run-ai-review")
async def teacher_run_candidate_ai_review(candidate_id: int, request: Request):
    require_roles(request, {"admin", "teacher"})
    async with evidence_client(timeout=620) as client:
        response = await client.post(evidence_url("").rstrip("/") + f"/answer-import-candidates/{candidate_id}/ai-review")
    if response.status_code >= 400:
        raise HTTPException(response.status_code, response.json().get("detail", "AI 核验失败"))
    return response.json()


@app.post("/api/teacher/question-bank/candidates/{candidate_id}/publish")
async def teacher_publish_candidate(candidate_id: int, req: CandidatePublishIn, request: Request):
    """The only UI path that publishes a candidate: explicit teacher confirmation + two gates."""
    actor = require_roles(request, {"admin", "teacher"})
    if req.confirm is not True:
        raise HTTPException(400, "发布需要 confirm=true")
    item = await _pending_candidate(candidate_id)
    gate = _candidate_quality_gate(item)
    ai = json.loads(item.get("ai_review_json") or "{}")
    content = (req.content_text if req.content_text is not None else item.get("content_text") or "").strip()
    answer = (req.std_answer if req.std_answer is not None else ai.get("std_answer") or "").strip()
    if req.manual_correction:
        if int(item.get("source_image_count") or 0) < 1 or len(content) < 3 or not answer:
            raise HTTPException(409, "人工修订发布需要来源图、非空题干和标准答案")
    elif not gate["can_publish_after_teacher_confirmation"]:
        raise HTTPException(409, "质量门禁未通过：" + "；".join(gate["reasons"]))
    payload = {
        "action": "approved", "content_text": content, "std_answer": answer,
        "full_solution": str(req.full_solution if req.full_solution is not None else ai.get("full_solution") or ""),
        "ptype": req.ptype if req.ptype in {"calc", "proof"} else (ai.get("ptype") if ai.get("ptype") in {"calc", "proof"} else item.get("ptype")),
        "note": (("教师人工 LaTeX 修订确认；" if req.manual_correction else "教师确认发布；") + req.note.strip())[:2000],
    }
    async with evidence_client(timeout=30) as client:
        response = await client.post(evidence_url("").rstrip("/") + f"/answer-import-candidates/{candidate_id}/review", json=payload)
    if response.status_code >= 400:
        raise HTTPException(response.status_code, response.json().get("detail", "发布失败"))

    # A batch is a frozen review snapshot.  Once a teacher manually corrects
    # and publishes an exception, reflect that canonical revision in the
    # snapshot too; otherwise the exception queue misleadingly keeps showing
    # the pre-edit OCR text after a successful verified writeback.
    synced_batches = []
    if req.manual_correction:
        render_ok, _ = _latex_render_gate(content)
        with connection() as conn:
            batch_rows = conn.execute(
                "SELECT DISTINCT batch_id FROM question_repair_batch_items WHERE candidate_id=?",
                (candidate_id,),
            ).fetchall()
            conn.execute(
                """UPDATE question_repair_batch_items
                   SET normalized_content=?, normalized_answer=?, normalized_solution=?,
                       render_ok=?, answer_consistent=1, bucket='manual_published',
                       reasons_json='[]', publish_status='published'
                   WHERE candidate_id=?""",
                (content, answer, payload["full_solution"], int(render_ok), candidate_id),
            )
            for batch_row in batch_rows:
                batch_id = batch_row["batch_id"]
                counts = {
                    row["bucket"]: row["n"]
                    for row in conn.execute(
                        "SELECT bucket, COUNT(*) AS n FROM question_repair_batch_items WHERE batch_id=? GROUP BY bucket",
                        (batch_id,),
                    ).fetchall()
                }
                conn.execute(
                    """UPDATE question_repair_batches
                       SET pass_count=?, sample_count=?, exception_count=?, qwen_count=(
                         SELECT COUNT(*) FROM question_repair_batch_items WHERE batch_id=? AND qwen_used=1
                       )
                       WHERE id=?""",
                    (counts.get("pass", 0), counts.get("sample", 0), counts.get("exception", 0), batch_id, batch_id),
                )
                conn.execute(
                    "INSERT INTO question_repair_batch_audit(batch_id,event_type,actor_user_id,payload_json) VALUES(?,?,?,?)",
                    (batch_id, "manual_exception_published", actor["id"], json.dumps({"candidate_id": candidate_id}, ensure_ascii=False)),
                )
                synced_batches.append(batch_id)
    return {**response.json(), "quality_gate": gate, "published_by_teacher": True, "synced_repair_batches": synced_batches}


import random

REPAIR_RULE_VERSION = "latex-repair-v1.1"

def _subquestion_coverage_gate(stem: str, answer: str, solution: str) -> tuple[bool, str | None]:
    """Do not publish a multi-part problem when its answer only covers a subset."""
    expected = {int(x) for x in re.findall(r"[（(]\s*(\d{1,2})\s*[）)]", str(stem or ""))}
    if len(expected) < 2:
        return True, None
    provided_text = str(answer or "") + "\n" + str(solution or "")
    covered = {int(x) for x in re.findall(r"[（(]\s*(\d{1,2})\s*[）)]", provided_text)}
    missing = sorted(expected - covered)
    if missing:
        return False, f"子问覆盖不足：题干 {len(expected)} 问，答案/推导仅标注 {len(covered)} 问，缺少 {len(missing)} 问"
    return True, None


def _ensure_repair_batch_schema() -> None:
    with connection() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS question_repair_batches (
          id INTEGER PRIMARY KEY AUTOINCREMENT, status TEXT NOT NULL DEFAULT 'processing',
          rule_version TEXT NOT NULL, qwen_model TEXT NOT NULL DEFAULT 'local-qwen', created_by INTEGER,
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP, processed_at TEXT, teacher_confirmed_at TEXT,
          teacher_confirmed_by INTEGER, total_count INTEGER NOT NULL DEFAULT 0, pass_count INTEGER NOT NULL DEFAULT 0,
          sample_count INTEGER NOT NULL DEFAULT 0, exception_count INTEGER NOT NULL DEFAULT 0,
          qwen_count INTEGER NOT NULL DEFAULT 0, notes TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS question_repair_batch_items (
          id INTEGER PRIMARY KEY AUTOINCREMENT, batch_id INTEGER NOT NULL, candidate_id INTEGER NOT NULL,
          snapshot_json TEXT NOT NULL, normalized_content TEXT, normalized_answer TEXT, normalized_solution TEXT,
          qwen_used INTEGER NOT NULL DEFAULT 0, render_ok INTEGER NOT NULL DEFAULT 0,
          answer_consistent INTEGER NOT NULL DEFAULT 0, confidence REAL NOT NULL DEFAULT 0,
          bucket TEXT NOT NULL, reasons_json TEXT NOT NULL DEFAULT '[]', publish_status TEXT NOT NULL DEFAULT 'pending',
          UNIQUE(batch_id,candidate_id)
        );
        CREATE TABLE IF NOT EXISTS question_repair_batch_audit (
          id INTEGER PRIMARY KEY AUTOINCREMENT, batch_id INTEGER NOT NULL, event_type TEXT NOT NULL,
          actor_user_id INTEGER, payload_json TEXT NOT NULL DEFAULT '{}', created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_repair_batch_items_bucket ON question_repair_batch_items(batch_id,bucket);
        """)

def _repair_latex(text: str) -> str:
    return str(text or "").replace("\\frac12", "\\frac{1}{2}").replace("\\frac 12", "\\frac{1}{2}").replace("\\(", "$").replace("\\)", "$").replace("\\[", "$$").replace("\\]", "$$").strip()

def _latex_render_gate(text: str) -> tuple[bool, list[str]]:
    t=str(text or "").strip(); reasons=[]
    if len(t)<12: reasons.append("题干过短或缺失")
    if t.count("$")%2: reasons.append("LaTeX 分隔符未闭合")
    if t.count("\\begin{") != t.count("\\end{"): reasons.append("LaTeX 环境未闭合")
    if "\\begin{aligned}" in t and "\\end{aligned}" not in t: reasons.append("aligned 环境不完整")
    if "\\frac{" in t and t.count("{") != t.count("}"): reasons.append("LaTeX 花括号不平衡")
    lines=[line.strip() for line in t.splitlines() if line.strip()]
    short_lines=sum(1 for line in lines if len(line)<=2)
    if len(lines)>=8 and short_lines/len(lines)>=0.25: reasons.append("OCR 文本结构异常（大量断裂行）")
    if re.search(r"(?m)^\s*解(?:设|：|\s)", t): reasons.append("题干混入解答过程")
    return not reasons,reasons

def _repair_bucket(item: dict) -> dict:
    try: ai=json.loads(item.get("ai_review_json") or "{}")
    except Exception: ai={}
    stem,answer,solution=(_repair_latex(item.get("content_text")),_repair_latex(ai.get("std_answer")),_repair_latex(ai.get("full_solution")))
    confidence=float(ai.get("confidence") or 0); render_ok,reasons=_latex_render_gate(stem)
    source_ok=int(item.get("source_image_count") or 0)>=1; completed=item.get("ai_review_status")=="completed"
    coverage_ok, coverage_reason = _subquestion_coverage_gate(stem, answer, solution)
    answer_consistent=bool(answer) and len(answer)<=3000 and not ai.get("risks") and coverage_ok
    if not source_ok: reasons.append("无来源图")
    if not completed: reasons.append("Qwen 尚未重建")
    if completed and confidence<.92: reasons.append(f"低置信度（{confidence:.2f}）")
    if coverage_reason: reasons.append(coverage_reason)
    if completed and not answer_consistent and not coverage_reason: reasons.append("答案缺失、冲突或 AI 标记风险")
    return {"stem":stem,"answer":answer,"solution":solution,"qwen_used":completed,"render_ok":render_ok,"answer_consistent":answer_consistent,"confidence":confidence,"bucket":"pass" if not reasons and confidence>=.98 else ("sample" if not reasons else "exception"),"reasons":reasons}

@app.get("/api/teacher/question-bank/coverage")
async def teacher_question_bank_coverage(request: Request):
    """Return source-level coverage without exposing raw import internals to the daily UI."""
    require_roles(request, {"admin", "teacher"})
    rows = []
    async with evidence_client(timeout=45) as client:
        for status in ("approved", "pending", "rejected"):
            response = await client.get(
                evidence_url("").rstrip("/") + "/answer-import-candidates",
                params={"status": status, "limit": 300},
            )
            if response.status_code >= 400:
                raise HTTPException(502, "题库来源统计暂不可用")
            rows.extend((status, item) for item in response.json().get("items", []))
    sources: dict[str, dict] = {}
    for status, item in rows:
        source = str(item.get("source_document") or item.get("source_pdf") or "未标记来源")
        record = sources.setdefault(source, {"source": source, "approved": 0, "pending": 0, "rejected": 0})
        record[status] += 1
    return {
        "sources": sorted(sources.values(), key=lambda x: x["source"]),
        "totals": {
            "approved": sum(x["approved"] for x in sources.values()),
            "pending": sum(x["pending"] for x in sources.values()),
            "rejected": sum(x["rejected"] for x in sources.values()),
        },
    }


class RepairBatchCreateIn(BaseModel):
    candidate_limit:int=Field(default=300,ge=1,le=300)

async def _process_repair_batch(batch_id: int) -> None:
    """Rebuild frozen candidates by ID, not through the mutable de-duplicated queue."""
    with connection() as conn:
        rows = conn.execute(
            "SELECT candidate_id FROM question_repair_batch_items WHERE batch_id=?",
            (batch_id,),
        ).fetchall()
    current: dict[int, dict] = {}
    total = len(rows)
    async with evidence_client(timeout=620) as client:
        for index, row in enumerate(rows, start=1):
            candidate_id = int(row["candidate_id"])
            try:
                # A repair batch is an immutable audit snapshot.  The public
                # pending list de-duplicates candidate rows, so it must never
                # be used to resolve these historical IDs.
                response = await client.get(
                    evidence_url("").rstrip("/") + f"/answer-import-candidates/{candidate_id}"
                )
                if response.status_code >= 400:
                    continue
                item = response.json()
                if item.get("ai_review_status") != "completed":
                    await client.post(
                        evidence_url("").rstrip("/") + f"/answer-import-candidates/{candidate_id}/ai-review"
                    )
                    response = await client.get(
                        evidence_url("").rstrip("/") + f"/answer-import-candidates/{candidate_id}"
                    )
                    if response.status_code < 400:
                        item = response.json()
                current[candidate_id] = item
            except Exception:
                # A failed source image/model invocation remains an exception;
                # later aggregation records its reason instead of publishing it.
                item = None
            # Persist progress per item: a long Qwen batch should never look stalled.
            if item:
                result = _repair_bucket(item)
                with connection() as progress_conn:
                    progress_conn.execute(
                        "UPDATE question_repair_batch_items SET normalized_content=?,normalized_answer=?,normalized_solution=?,qwen_used=?,render_ok=?,answer_consistent=?,confidence=?,bucket=?,reasons_json=? WHERE batch_id=? AND candidate_id=?",
                        (result["stem"], result["answer"], result["solution"], int(result["qwen_used"]), int(result["render_ok"]), int(result["answer_consistent"]), result["confidence"], result["bucket"], json.dumps(result["reasons"], ensure_ascii=False), batch_id, candidate_id),
                    )
                    counts = {r["bucket"]: r["n"] for r in progress_conn.execute("SELECT bucket,COUNT(*) n FROM question_repair_batch_items WHERE batch_id=? GROUP BY bucket", (batch_id,))}
                    qwen = progress_conn.execute("SELECT COUNT(*) n FROM question_repair_batch_items WHERE batch_id=? AND qwen_used=1", (batch_id,)).fetchone()["n"]
                    progress_conn.execute(
                        "UPDATE question_repair_batches SET pass_count=?,sample_count=?,exception_count=?,qwen_count=?,notes=? WHERE id=?",
                        (counts.get("pass", 0), counts.get("sample", 0), counts.get("exception", 0), qwen, f"正在从来源图重建：已处理 {index}/{total} 题；Qwen 已完成 {qwen} 题。", batch_id),
                    )
                    progress_conn.commit()
    with connection() as conn:
        for row in rows:
            item = current.get(int(row["candidate_id"]))
            if not item:
                continue
            result = _repair_bucket(item)
            conn.execute(
                "UPDATE question_repair_batch_items SET normalized_content=?,normalized_answer=?,normalized_solution=?,qwen_used=?,render_ok=?,answer_consistent=?,confidence=?,bucket=?,reasons_json=? WHERE batch_id=? AND candidate_id=?",
                (
                    result["stem"], result["answer"], result["solution"], int(result["qwen_used"]),
                    int(result["render_ok"]), int(result["answer_consistent"]), result["confidence"],
                    result["bucket"], json.dumps(result["reasons"], ensure_ascii=False), batch_id, row["candidate_id"],
                ),
            )
        counts = {r["bucket"]: r["n"] for r in conn.execute(
            "SELECT bucket,COUNT(*) n FROM question_repair_batch_items WHERE batch_id=? GROUP BY bucket",
            (batch_id,),
        )}
        qwen = conn.execute(
            "SELECT COUNT(*) n FROM question_repair_batch_items WHERE batch_id=? AND qwen_used=1",
            (batch_id,),
        ).fetchone()["n"]
        conn.execute(
            "UPDATE question_repair_batches SET status='ready',processed_at=CURRENT_TIMESTAMP,pass_count=?,sample_count=?,exception_count=?,qwen_count=?,notes=? WHERE id=?",
            (
                counts.get("pass", 0), counts.get("sample", 0), counts.get("exception", 0), qwen,
                f"批量重建完成：Qwen 已完成 {qwen} 题；其余题按异常原因保留。",
                batch_id,
            ),
        )
        conn.execute(
            "INSERT INTO question_repair_batch_audit(batch_id,event_type,payload_json) VALUES(?,?,?)",
            (
                batch_id, "processed",
                json.dumps(
                    {
                        "rule_version": REPAIR_RULE_VERSION, "qwen_count": qwen,
                        "pass": counts.get("pass", 0), "sample": counts.get("sample", 0),
                        "exception": counts.get("exception", 0), "lookup": "frozen_candidate_id",
                    },
                    ensure_ascii=False,
                ),
            ),
        )


@app.post("/api/teacher/question-bank/repair-batches/{batch_id}/rebuild-qwen")
def rebuild_repair_batch_qwen(batch_id: int, request: Request):
    """Queue source-image Qwen rebuilds for every still-unrebuilt item in a frozen batch."""
    actor = require_roles(request, {"admin", "teacher"})
    view = _repair_batch_view(batch_id)
    batch = view["batch"]
    if batch["status"] == "rebuilding":
        raise HTTPException(409, "该批次已经在重建队列中，请稍后刷新查看进度")
    # A batch can be published after its safe items were sampled and adopted,
    # while exception items remain intentionally pending.  Those leftovers are
    # valid targets for a later source-image rebuild.
    if batch["status"] not in {"ready", "partial_failure", "published"}:
        raise HTTPException(409, "请等待当前批次准备完成后再启动批量重建")
    eligible = [
        item for item in view["exceptions"]
        if "Qwen 尚未重建" in (item.get("reasons") or [])
        and int(item.get("source_image_count") or 0) > 0
    ]
    if not eligible:
        raise HTTPException(409, "当前批次没有可从来源图批量重建的题目")
    if not rq_enabled():
        raise HTTPException(503, "批量重建需要生产队列；当前未启用 RQ")
    try:
        redis, Queue, Retry = _rq_objects()
        redis.ping()
        queue = Queue(settings.rq_queue_name, connection=redis)
        job = queue.enqueue(
            "app.rq_tasks.rebuild_question_candidates_task",
            batch_id,
            job_timeout=max(settings.rq_job_timeout_seconds, 3600),
            result_ttl=7 * 24 * 3600,
            failure_ttl=30 * 24 * 3600,
            retry=Retry(max=settings.rq_retry_max, interval=[60, 300, 900]),
        )
    except Exception as exc:
        raise HTTPException(503, f"无法投递 Qwen 重建队列：{str(exc)[:160]}") from exc
    with connection() as conn:
        conn.execute("UPDATE question_repair_batches SET status='rebuilding',notes=? WHERE id=?",
                     (f"已投递 Qwen 来源图重建：{len(eligible)} 题", batch_id))
        conn.execute(
            "INSERT INTO question_repair_batch_audit(batch_id,event_type,actor_user_id,payload_json) VALUES(?,?,?,?)",
            (batch_id, "qwen_rebuild_queued", actor["id"],
             json.dumps({"eligible_count": len(eligible), "rq_job_id": job.id}, ensure_ascii=False)),
        )
    return {
        "batch_id": batch_id, "queued_count": len(eligible), "rq_job_id": job.id,
        "message": f"已将 {len(eligible)} 道题加入 Qwen 来源图重建队列；页面无需保持打开，完成后会自动重新分组。",
    }


@app.post("/api/teacher/question-bank/repair-batches")
async def create_repair_batch(req:RepairBatchCreateIn,request:Request,background_tasks:BackgroundTasks):
    actor=require_roles(request,{"admin","teacher"}); _ensure_repair_batch_schema()
    async with evidence_client(timeout=30) as client:
        response=await client.get(evidence_url("").rstrip("/")+"/answer-import-candidates",params={"status":"pending","limit":req.candidate_limit})
    if response.status_code>=400: raise HTTPException(502,"题库候选服务暂不可用")
    items=response.json().get("items",[])
    if not items: raise HTTPException(409,"没有待处理候选，无法创建修复批次")
    with connection() as conn:
        cur=conn.execute("INSERT INTO question_repair_batches(status,rule_version,created_by,total_count,notes) VALUES('processing',?,?,?,'冻结候选、来源图与规则版本')",(REPAIR_RULE_VERSION,actor["id"],len(items))); batch_id=cur.lastrowid
        for item in items:
            result=_repair_bucket(item)
            conn.execute("INSERT INTO question_repair_batch_items(batch_id,candidate_id,snapshot_json,normalized_content,normalized_answer,normalized_solution,qwen_used,render_ok,answer_consistent,confidence,bucket,reasons_json) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",(batch_id,item["id"],json.dumps(item,ensure_ascii=False),result["stem"],result["answer"],result["solution"],int(result["qwen_used"]),int(result["render_ok"]),int(result["answer_consistent"]),result["confidence"],result["bucket"],json.dumps(result["reasons"],ensure_ascii=False)))
        conn.execute("INSERT INTO question_repair_batch_audit(batch_id,event_type,actor_user_id,payload_json) VALUES(?,?,?,?)",(batch_id,"created",actor["id"],json.dumps({"frozen_candidates":len(items),"rule_version":REPAIR_RULE_VERSION},ensure_ascii=False)))
    background_tasks.add_task(_process_repair_batch, batch_id)
    return {"batch_id":batch_id,"status":"processing","message":"修复批次已冻结，正在执行规则清洗、缺失项 Qwen 重建与自动分桶；页面会显示处理状态。"}

def _repair_batch_view(batch_id:int)->dict:
    _ensure_repair_batch_schema()
    with connection() as conn:
        b=conn.execute("SELECT * FROM question_repair_batches WHERE id=?",(batch_id,)).fetchone()
        if not b: raise HTTPException(404,"未找到修复批次")
        rows=conn.execute("SELECT * FROM question_repair_batch_items WHERE batch_id=? ORDER BY id",(batch_id,)).fetchall()
        audits=conn.execute("SELECT event_type,created_at,payload_json FROM question_repair_batch_audit WHERE batch_id=? ORDER BY id DESC",(batch_id,)).fetchall()
    items=[]
    for r in rows:
        snap=json.loads(r["snapshot_json"])
        items.append({"id":r["id"],"candidate_id":r["candidate_id"],"section_no":snap.get("section_no"),"problem_no":snap.get("problem_no"),"content_text":r["normalized_content"],"std_answer":r["normalized_answer"],"full_solution":r["normalized_solution"],"bucket":r["bucket"],"render_ok":bool(r["render_ok"]),"answer_consistent":bool(r["answer_consistent"]),"confidence":r["confidence"],"qwen_used":bool(r["qwen_used"]),"reasons":json.loads(r["reasons_json"] or "[]"),"source_image_count":snap.get("source_image_count",0),"publish_status":r["publish_status"]})
    samples=[x for x in items if x["bucket"]=="sample"]; rng=random.Random(f"repair-sample-{batch_id}")
    updated = [x for x in items if x.get("publish_status") in {"published", "verified_auto"}]
    provisional = [x for x in items if x.get("publish_status") == "provisional"]
    exceptions = [x for x in items if x["bucket"] == "exception" and x.get("publish_status") not in {"provisional", "published", "verified_auto"}]
    return {"batch":dict(b),"samples":rng.sample(samples,min(5,len(samples))) if samples else [],
            "updated": updated,
            "updated_counts": {"teacher_published": sum(x.get("publish_status") == "published" for x in updated),
                               "auto_verified": sum(x.get("publish_status") == "verified_auto" for x in updated)},
            "exceptions": exceptions,
            "provisional": provisional,
            "passed_count":sum(x["bucket"]=="pass" for x in items),
            "audit":[{"event_type":a["event_type"],"created_at":a["created_at"],"payload":json.loads(a["payload_json"] or "{}")} for a in audits]}

@app.get("/api/teacher/question-bank/repair-batches")
def get_repair_batches(request:Request):
    require_roles(request,{"admin","teacher"}); _ensure_repair_batch_schema()
    with connection() as conn: rows=conn.execute("SELECT * FROM question_repair_batches ORDER BY id DESC LIMIT 20").fetchall()
    return {"items":[dict(x) for x in rows]}

@app.get("/api/teacher/question-bank/repair-batches/{batch_id}")
def get_repair_batch(batch_id:int,request:Request):
    require_roles(request,{"admin","teacher"}); return _repair_batch_view(batch_id)

class RepairBatchConfirmIn(BaseModel):
    samples_checked:bool=False
    note:str=""

@app.post("/api/teacher/question-bank/repair-batches/{batch_id}/confirm")
async def confirm_repair_batch(batch_id:int,req:RepairBatchConfirmIn,request:Request):
    actor=require_roles(request,{"admin","teacher"})
    if not req.samples_checked: raise HTTPException(400,"请先勾选“我已核对随机抽样题”")
    view=_repair_batch_view(batch_id)
    if view["batch"]["status"]!="ready": raise HTTPException(409,"该批次不是可确认状态")
    publish_items=list(view["samples"])
    with connection() as conn: pass_rows=conn.execute("SELECT * FROM question_repair_batch_items WHERE batch_id=? AND bucket='pass'",(batch_id,)).fetchall()
    for r in pass_rows:
        snap=json.loads(r["snapshot_json"]); publish_items.append({"candidate_id":r["candidate_id"],"content_text":r["normalized_content"],"std_answer":r["normalized_answer"],"full_solution":r["normalized_solution"],"ptype":snap.get("ptype")})
    published=0; failures=[]
    async with evidence_client(timeout=35) as client:
        for x in publish_items:
            payload={"action":"approved","content_text":x["content_text"],"std_answer":x["std_answer"],"full_solution":x["full_solution"] or "","ptype":x.get("ptype") if x.get("ptype") in {"calc","proof"} else "calc","note":"修复批次教师抽样确认发布"}
            response=await client.post(evidence_url("").rstrip("/")+f"/answer-import-candidates/{x['candidate_id']}/review",json=payload)
            if response.status_code<400:
                published+=1
                with connection() as conn: conn.execute("UPDATE question_repair_batch_items SET publish_status='published' WHERE batch_id=? AND candidate_id=?",(batch_id,x["candidate_id"]))
            else: failures.append({"candidate_id":x["candidate_id"],"detail":response.text[:200]})
    with connection() as conn:
        conn.execute("UPDATE question_repair_batches SET status=?,teacher_confirmed_at=CURRENT_TIMESTAMP,teacher_confirmed_by=?,notes=? WHERE id=?",("published" if not failures else "partial_failure",actor["id"],req.note[:1000],batch_id))
        conn.execute("INSERT INTO question_repair_batch_audit(batch_id,event_type,actor_user_id,payload_json) VALUES(?,?,?,?)",(batch_id,"teacher_confirmed",actor["id"],json.dumps({"published":published,"failures":failures,"sample_count":len(view["samples"])},ensure_ascii=False)))
    return {"batch_id":batch_id,"published":published,"failures":failures,"exceptions_kept":len(view["exceptions"])}


class RepairBatchProvisionalIn(BaseModel):
    confirm: bool = False

@app.post("/api/teacher/question-bank/repair-batches/{batch_id}/provisional-import")
async def provisional_import_repair_batch(batch_id: int, req: RepairBatchProvisionalIn, request: Request):
    actor = require_roles(request, {"admin", "teacher"})
    if not req.confirm:
        raise HTTPException(400, "请确认将本批恢复结果纳入待核验题库")
    view = _repair_batch_view(batch_id)
    if view["batch"]["status"] == "rebuilding":
        raise HTTPException(409, "请等待当前重建任务完成")
    eligible = [x for x in view["exceptions"]
                if int(x.get("source_image_count") or 0) > 0
                and len(str(x.get("content_text") or "").strip()) >= 12
                and bool(str(x.get("std_answer") or "").strip())]
    imported, failures = 0, []
    async with evidence_client(timeout=35) as client:
        for item in eligible:
            payload = {"action": "provisional", "content_text": item["content_text"],
                       "std_answer": item["std_answer"], "full_solution": item.get("full_solution") or "",
                       "ptype": "calc", "note": "批量来源图重建：待核验题库，不启用自动评分"}
            response = await client.post(evidence_url("").rstrip("/") + f"/answer-import-candidates/{item['candidate_id']}/review", json=payload)
            if response.status_code < 400:
                imported += 1
                with connection() as conn:
                    conn.execute("UPDATE question_repair_batch_items SET publish_status='provisional' WHERE batch_id=? AND candidate_id=?", (batch_id, item["candidate_id"]))
            else:
                failures.append(item["candidate_id"])
    with connection() as conn:
        conn.execute("INSERT INTO question_repair_batch_audit(batch_id,event_type,actor_user_id,payload_json) VALUES(?,?,?,?)",
                     (batch_id, "provisional_imported", actor["id"], json.dumps({"imported": imported, "failures": failures}, ensure_ascii=False)))
    return {"imported": imported, "failures": failures, "manual_remaining": len(view["exceptions"]) - imported}

@app.get("/question-bank-review", response_class=HTMLResponse)
def teacher_question_bank_review_page(request: Request):
    try:
        require_roles(request, {"admin", "teacher"})
    except HTTPException as exc:
        if exc.status_code == 401:
            return RedirectResponse("/login", status_code=303)
        raise
    return FileResponse(Path(__file__).with_name("question_bank_review.html"), headers={"Cache-Control": "no-store, max-age=0"})


class TeacherEvaluationCaseIn(BaseModel):
    source_submission_id: int
    question_id: int
    case_type: str
    teacher_correct: bool
    expected_route: str
    expected_diagnosis: str = ""
    requires_teacher_review: bool = False
    disagreement_reason: str = ""
    teacher_note: str = ""


def _teacher_eval_candidate_rows(actor: dict) -> list[dict]:
    scope = teacher_id_for_scope(actor)
    with connection() as conn:
        sql = """SELECT s.id AS submission_id,a.title,c.name AS class_name,j.result_json,
                        q.id AS question_id,q.content,q.question_type,q.answer,ge.confirmed_at
                 FROM submissions s JOIN assignments a ON a.id=s.assignment_id
                 JOIN classes c ON c.id=a.class_id JOIN grading_jobs j ON j.submission_id=s.id
                 JOIN grading_experiences ge ON ge.submission_id=s.id
                 JOIN assignment_questions aq ON aq.assignment_id=a.id
                 JOIN questions q ON q.id=aq.question_id WHERE j.status='completed'"""
        args=[]
        if scope is not None:
            sql += " AND c.teacher_user_id=?"; args.append(scope)
        rows=conn.execute(sql+" ORDER BY ge.confirmed_at DESC, s.id DESC",args).fetchall()
    result=[]
    for row in rows:
        payload=json.loads(row["result_json"] or "{}")
        item=next((x for x in (payload.get("results") or []) if int(x.get("question_id") or -1)==row["question_id"]),None)
        if item:
            result.append({"submission_id":row["submission_id"],"question_id":row["question_id"],
              "assignment_title":row["title"],"class_name":row["class_name"],
              "question_type":normalize_question_type(row["question_type"]),"problem_text":str(row["content"] or "")[:4000],
              "student_answer":str(item.get("recognized_work") or "")[:4000],
              "standard_answer":str((item.get("evidence") or {}).get("standard_answer") or row["answer"] or "")[:4000],
              "ai_initial_correct":item.get("correct"),"confidence":item.get("confidence"),
              "needs_review":bool(item.get("needs_review")),"risks":item.get("risks") or [],
              "teacher_confirmed_at":row["confirmed_at"]})
    return result


@app.get("/api/teacher/evaluation-cases/candidates")
def teacher_evaluation_case_candidates(request: Request):
    return {"items":_teacher_eval_candidate_rows(require_roles(request,{"admin","teacher"}))}


@app.post("/api/teacher/evaluation-cases")
def create_teacher_evaluation_case(req: TeacherEvaluationCaseIn, request: Request):
    actor=require_roles(request,{"admin","teacher"})
    if req.case_type not in {"proof","handwriting_steps","multi_part_page","ai_teacher_disagreement"}:
        raise HTTPException(422,"请选择有效案例类型")
    if req.expected_route not in {"diagnose_misconception","independent_solve","teacher_review"}:
        raise HTTPException(422,"请选择有效的安全路由")
    candidate=next((x for x in _teacher_eval_candidate_rows(actor) if x["submission_id"]==req.source_submission_id and x["question_id"]==req.question_id),None)
    if not candidate: raise HTTPException(404,"未找到可标注的、已教师复核作业条目")
    if not candidate["student_answer"] or not candidate["standard_answer"]:
        raise HTTPException(422,"该条目缺少必要数学文本，不能作为文本回归样本。")
    with connection() as conn:
        conn.execute("""INSERT INTO teacher_evaluation_cases(source_submission_id,question_id,case_type,question_type,problem_text,student_answer,standard_answer,ai_initial_correct,teacher_correct,expected_route,expected_diagnosis,requires_teacher_review,disagreement_reason,teacher_note,created_by)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
          ON CONFLICT(source_submission_id,question_id) DO UPDATE SET case_type=excluded.case_type,teacher_correct=excluded.teacher_correct,expected_route=excluded.expected_route,expected_diagnosis=excluded.expected_diagnosis,requires_teacher_review=excluded.requires_teacher_review,disagreement_reason=excluded.disagreement_reason,teacher_note=excluded.teacher_note,created_by=excluded.created_by,created_at=CURRENT_TIMESTAMP""",
          (req.source_submission_id,req.question_id,req.case_type,candidate["question_type"],candidate["problem_text"],candidate["student_answer"],candidate["standard_answer"],None if candidate["ai_initial_correct"] is None else int(bool(candidate["ai_initial_correct"])),int(req.teacher_correct),req.expected_route,req.expected_diagnosis.strip()[:120],int(req.requires_teacher_review),req.disagreement_reason.strip()[:500],req.teacher_note.strip()[:1000],actor["id"]))
    audit(actor,"teacher_evaluation_case.save","submission",req.source_submission_id,None,{"question_id":req.question_id,"case_type":req.case_type})
    return {"ok":True,"message":"已保存为脱敏教师标注案例；不会保存学生姓名、学号或原始图片。"}


@app.get("/api/teacher/evaluation-cases")
def list_teacher_evaluation_cases(request: Request):
    actor=require_roles(request,{"admin","teacher"}); scope=teacher_id_for_scope(actor)
    with connection() as conn:
        sql="SELECT id,case_type,question_type,ai_initial_correct,teacher_correct,expected_route,expected_diagnosis,requires_teacher_review,disagreement_reason,created_at FROM teacher_evaluation_cases"
        rows=conn.execute(sql+(" WHERE created_by=?" if scope is not None else "")+" ORDER BY id DESC LIMIT 100",([scope] if scope is not None else [])).fetchall()
    return {"items":[dict(x) for x in rows]}


@app.get("/teacher-evaluation-cases", response_class=HTMLResponse)
def teacher_evaluation_cases_page(request: Request):
    try: require_roles(request,{"admin","teacher"})
    except HTTPException as exc:
        if exc.status_code==401: return RedirectResponse("/login?next=/teacher-evaluation-cases",status_code=303)
        raise
    return FileResponse(Path(__file__).with_name("teacher_evaluation_cases.html"),headers={"Cache-Control":"no-store, max-age=0"})


@app.get("/api/reports/teaching-actions")
def teaching_actions(request: Request):
    """Teacher-home action counts.  Excludes demo/legacy assignments from live teaching reminders."""
    actor=require_roles(request,{"admin","teacher"});scope=teacher_id_for_scope(actor)
    now=datetime.now(timezone.utc)
    with connection() as conn:
        sql="""SELECT a.id,a.due_at,a.status,a.class_id,
                (SELECT COUNT(*) FROM students st WHERE st.class_id=a.class_id) AS roster_count,
                (SELECT COUNT(*) FROM submissions s WHERE s.assignment_id=a.id) AS submitted_count
                FROM assignments a JOIN classes c ON c.id=a.class_id
                WHERE a.class_id IS NOT NULL AND COALESCE(a.is_demo,0)=0"""
        args=[]
        if scope is not None: sql+=" AND c.teacher_user_id=?";args.append(scope)
        assignments=conn.execute(sql,args).fetchall()
        pending_sql="""SELECT COUNT(*) FROM submissions s JOIN assignments a ON a.id=s.assignment_id
          JOIN classes c ON c.id=a.class_id LEFT JOIN grading_jobs j ON j.submission_id=s.id
          WHERE a.class_id IS NOT NULL AND COALESCE(a.is_demo,0)=0
          AND (s.needs_review=1 OR j.status IN ('queued','running','failed'))"""
        release_sql="""SELECT COUNT(*) FROM submissions s JOIN assignments a ON a.id=s.assignment_id
          JOIN classes c ON c.id=a.class_id JOIN grading_experiences ge ON ge.submission_id=s.id
          WHERE a.class_id IS NOT NULL AND COALESCE(a.is_demo,0)=0 AND s.status='graded' AND s.released_at IS NULL"""
        if scope is not None:
            pending_sql+=" AND c.teacher_user_id=?";release_sql+=" AND c.teacher_user_id=?"
        pending=conn.execute(pending_sql,([scope] if scope is not None else [])).fetchone()[0]
        ready_release=conn.execute(release_sql,([scope] if scope is not None else [])).fetchone()[0]
    active=[]
    for item in assignments:
        try: due=datetime.fromisoformat(str(item["due_at"]).replace("Z","+00:00"))
        except ValueError: continue
        if item["status"]=="published" and due>=now: active.append(item)
    expected=sum(int(x["roster_count"] or 0) for x in active)
    submitted=sum(min(int(x["submitted_count"] or 0),int(x["roster_count"] or 0)) for x in active)
    return {"active_assignment_count":len(active),"submitted_count":submitted,"expected_submission_count":expected,
        "submission_rate":round(submitted/expected*100,1) if expected else None,
        "pending_review_count":pending,"ready_to_release_count":ready_release}

@app.get("/api/reports/agent-traces")
def agent_traces(request: Request, limit: int = 30):
    """Teacher-only observability feed; never returns student answer or standard answer."""
    actor = require_roles(request, {"admin", "teacher"})
    scope = teacher_id_for_scope(actor)
    limit = max(1, min(limit, 100))
    with connection() as conn:
        where, params = "", []
        if scope is not None:
            where = " JOIN submissions s ON s.id=t.submission_id JOIN assignments a ON a.id=s.assignment_id JOIN classes c ON c.id=a.class_id WHERE c.teacher_user_id=?"
            params.append(scope)
        rows = conn.execute(
            "SELECT t.id,t.created_at,t.agent_trace_id,t.submission_id,t.question_id,t.mode,t.qwen_used,t.teacher_review_needed,t.latency_ms,t.execution_trace_json FROM agent_learning_traces t" + where + " ORDER BY t.id DESC LIMIT ?",
            [*params, limit],
        ).fetchall()
    return [{"id": r["id"], "created_at": r["created_at"], "trace_id": r["agent_trace_id"],
             "submission_id": r["submission_id"], "question_id": r["question_id"], "mode": r["mode"],
             "qwen_used": bool(r["qwen_used"]), "teacher_review_needed": bool(r["teacher_review_needed"]),
             "latency_ms": r["latency_ms"], "events": json.loads(r["execution_trace_json"] or "[]")} for r in rows]


@app.get("/api/reports/agent-evaluation")
def agent_evaluation(request: Request):
    """Teacher-facing reliability and cost/latency summary for the released-mistake Agent."""
    actor = require_roles(request, {"admin", "teacher"})
    scope = teacher_id_for_scope(actor)
    with connection() as conn:
        where = ""
        params: list[object] = []
        if scope is not None:
            where = " JOIN submissions s ON s.id=t.submission_id JOIN assignments a ON a.id=s.assignment_id JOIN classes c ON c.id=a.class_id WHERE c.teacher_user_id=?"
            params.append(scope)
        rows = conn.execute("SELECT t.qwen_called,t.qwen_success,t.qwen_adopted,t.teacher_review_needed,t.handwriting_ocr_failed,t.latency_ms,t.question_type,t.learning_outcome,t.execution_trace_json FROM agent_learning_traces t" + where, params).fetchall()
    total = len(rows)
    def rate(key: str) -> float:
        return round((sum(int(r[key] or 0) for r in rows) / total * 100), 1) if total else 0.0
    with connection() as conn:
        attempt_where = ""
        attempt_params: list[object] = []
        if scope is not None:
            attempt_where = " JOIN submissions s ON s.id=t.submission_id JOIN assignments a ON a.id=s.assignment_id JOIN classes c ON c.id=a.class_id WHERE c.teacher_user_id=?"
            attempt_params.append(scope)
        attempts = conn.execute(
            "SELECT a.prior_trace_id,a.correct,a.input_kind,a.created_at AS attempt_created_at,t.created_at AS trace_created_at,t.mode FROM agent_learning_attempts a JOIN agent_learning_traces t ON t.agent_trace_id=a.prior_trace_id" + attempt_where,
            attempt_params,
        ).fetchall()
    attempted_traces = {r["prior_trace_id"] for r in attempts}
    corrected_traces = {r["prior_trace_id"] for r in attempts if int(r["correct"] or 0)}
    reattempt_minutes = []
    earliest_attempt = {}
    for attempt in attempts:
        trace_id = attempt["prior_trace_id"]
        current = earliest_attempt.get(trace_id)
        if current is None or str(attempt["attempt_created_at"] or "") < str(current["attempt_created_at"] or ""):
            earliest_attempt[trace_id] = attempt
    for attempt in earliest_attempt.values():
        try:
            started = datetime.fromisoformat(str(attempt["trace_created_at"]).replace("Z", "+00:00"))
            retried = datetime.fromisoformat(str(attempt["attempt_created_at"]).replace("Z", "+00:00"))
            reattempt_minutes.append(max(0.0, (retried - started).total_seconds() / 60))
        except (TypeError, ValueError):
            pass
    diagnosis_counter = {}
    for row in rows:
        for event in json.loads(row["execution_trace_json"] or "[]"):
            for code in event.get("diagnosis_codes") or []:
                diagnosis_counter[code] = diagnosis_counter.get(code, 0) + 1
    effectiveness = {}
    for mode in ("diagnose", "hint", "solution"):
        mode_attempts = [r for r in attempts if r["mode"] == mode]
        effectiveness[mode] = {
            "attempt_count": len(mode_attempts),
            "verified_correct_rate": round(sum(int(r["correct"] or 0) for r in mode_attempts) / len(mode_attempts) * 100, 1) if mode_attempts else 0.0,
        }
    reattempt_by_input = {}
    for input_kind in ("text", "image"):
        input_attempts = [r for r in attempts if (r["input_kind"] or "text") == input_kind]
        reattempt_by_input[input_kind] = {
            "attempt_count": len(input_attempts),
            "verified_correct_rate": round(sum(int(r["correct"] or 0) for r in input_attempts) / len(input_attempts) * 100, 1) if input_attempts else 0.0,
        }
    return {
        "sample_size": total,
        "sympy_direct_rate": round(100 - rate("qwen_called"), 1) if total else 0.0,
        "qwen_fallback_rate": rate("qwen_called"),
        "qwen_success_rate": rate("qwen_success"),
        "qwen_adopted_rate": rate("qwen_adopted"),
        "teacher_review_rate": rate("teacher_review_needed"),
        "handwriting_ocr_failure_rate": rate("handwriting_ocr_failed"),
        "average_response_ms": round(sum(float(r["latency_ms"] or 0) for r in rows) / total, 1) if total else 0.0,
        "review_rate_by_type": {kind: round(sum(int(r["teacher_review_needed"] or 0) for r in rows if r["question_type"] == kind) / sum(1 for r in rows if r["question_type"] == kind) * 100, 1) for kind in sorted({r["question_type"] for r in rows})},
        "learning_outcomes": {kind: sum(1 for r in rows if r["learning_outcome"] == kind) for kind in ("reworked", "corrected", "requested_solution", "not_resolved")},
        "verified_reattempt_count": len(attempted_traces),
        "verified_corrected_count": len(corrected_traces),
        "image_reattempt_count": sum(1 for r in attempts if r["input_kind"] == "image"),
        "verified_reattempt_rate": round(len(attempted_traces) / total * 100, 1) if total else 0.0,
        "verified_corrected_rate": round(len(corrected_traces) / len(attempted_traces) * 100, 1) if attempted_traces else 0.0,
        "hint_effectiveness": effectiveness,
        "reattempt_by_input": reattempt_by_input,
        "average_minutes_to_first_reattempt": round(sum(reattempt_minutes) / len(reattempt_minutes), 1) if reattempt_minutes else None,
        "common_diagnoses": [{"code": code, "count": count} for code, count in sorted(diagnosis_counter.items(), key=lambda item: (-item[1], item[0]))[:10]],
    }


@app.get("/agent-observability", response_class=HTMLResponse)
def agent_observability_page(request: Request):
    try:
        require_roles(request, {"admin", "teacher"})
    except HTTPException as exc:
        if exc.status_code == 401:
            return RedirectResponse("/login", status_code=303)
        raise
    return FileResponse(Path(__file__).with_name("agent_observability.html"))


@app.get("/agent-evaluation", response_class=HTMLResponse)
def agent_evaluation_page(request: Request):
    try:
        require_roles(request, {"admin", "teacher"})
    except HTTPException as exc:
        if exc.status_code == 401:
            return RedirectResponse("/login", status_code=303)
        raise
    return FileResponse(Path(__file__).with_name("agent_evaluation.html"))


class MistakeReviewIn(BaseModel):
    mode: str = "diagnose"
    student_steps: str = Field(default="", max_length=12000)


@app.post("/api/student/released-submissions/{submission_id}/mistakes/{question_id}/review")
async def review_released_mistake(submission_id: int, question_id: int, req: MistakeReviewIn, request: Request):
    """Only a released, owned wrong answer may enter the learning Agent."""
    actor = require_roles(request, {"student"})
    if req.mode not in {"diagnose", "hint", "solution"}:
        raise HTTPException(422, "不支持的复盘模式")
    with connection() as conn:
        row = conn.execute("SELECT j.result_json FROM submissions s JOIN assignments a ON a.id=s.assignment_id JOIN students st ON st.class_id=a.class_id AND st.user_id=? LEFT JOIN grading_jobs j ON j.submission_id=s.id WHERE s.id=? AND s.status=? AND s.released_at IS NOT NULL", (actor["id"], submission_id, "graded")).fetchone()
        if not row:
            raise HTTPException(404, "未找到已发布的作业结果")
        results = json.loads(row["result_json"] or "{}").get("results") or []
        item = next((x for x in results if int(x.get("question_id") or -1) == question_id), None)
        if not item or (item.get("correct") is True and float(item.get("score") or 0) >= float(item.get("max_score") or 0)):
            raise HTTPException(404, "该题不是可复盘的已发布错题")
        question = conn.execute("SELECT source_problem_id,question_type FROM questions WHERE id=?", (question_id,)).fetchone()
        if not question or not question["source_problem_id"]:
            raise HTTPException(409, "该错题尚未关联到可靠题库原题")
        source_problem_id = question["source_problem_id"]
        student_answer = str(item.get("recognized_work") or "").strip()
        teacher_feedback = str(item.get("feedback") or "")
    started = time.perf_counter()
    try:
        async with evidence_client(timeout=620) as client:
            response = await client.post(evidence_url("").rstrip("/") + f"/agent/problems/{source_problem_id}/learn", json={"student_answer": student_answer or "未识别到作答", "student_steps": req.student_steps, "mode": req.mode, "teacher_feedback": teacher_feedback})
    except httpx.HTTPError as exc:
        raise HTTPException(502, "错题复盘 Agent 暂不可用") from exc
    latency_ms = round((time.perf_counter() - started) * 1000, 1)
    if response.status_code >= 400:
        try:
            detail = response.json().get("detail", "错题复盘请求失败")
        except ValueError:
            detail = "错题复盘请求失败"
        raise HTTPException(response.status_code, detail)
    result = response.json()
    agent_trace_id = result.pop("trace_id", None)
    execution_trace = result.pop("execution_trace", []) or []
    diagnosis_codes = [str(item.get("code")) for item in (result.get("diagnosis") or {}).get("diagnoses", []) if item.get("code")]
    execution_trace.append({"node": "diagnosis_summary", "skills": ["misconception_diagnosis"], "success": True, "diagnosis_codes": diagnosis_codes})
    qwen_called = any("independent_solving" in event.get("skills", []) for event in execution_trace)
    qwen_success = any("independent_solving" in event.get("skills", []) and event.get("success") for event in execution_trace)
    qwen_adopted = bool((result.get("solution_comparison") or {}).get("consistent"))
    qwen_used = qwen_called
    needs_teacher_review = result.get("action") == "teacher_review"
    question_type = "proof" if "证明" in str(question["question_type"] or "") else "calc"
    failure_codes = [event["error_code"] for event in execution_trace if event.get("error_code")]
    with connection() as conn:
        conn.execute(
            "INSERT INTO agent_learning_traces(student_user_id,submission_id,question_id,source_problem_id,mode,qwen_used,teacher_review_needed,handwriting_ocr_failed,latency_ms,agent_trace_id,execution_trace_json,qwen_called,qwen_success,qwen_adopted,question_type,learning_outcome) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (actor["id"], submission_id, question_id, source_problem_id, req.mode, int(qwen_used), int(needs_teacher_review), int(not student_answer), latency_ms, agent_trace_id, json.dumps(execution_trace, ensure_ascii=False), int(qwen_called), int(qwen_success), int(qwen_adopted), question_type, "requested_solution" if req.mode == "solution" else None),
        )
    return {**result, "trace": {
        "trace_id": agent_trace_id, "mode": req.mode, "qwen_used": qwen_used,
        "needs_teacher_review": needs_teacher_review, "latency_ms": latency_ms,
        "failure_codes": failure_codes,
    }}


def _record_handwriting_failure(actor_id: int, submission_id: int, question_id: int, source_problem_id: str, question_type: str, mode: str, latency_ms: float, error_code: str) -> None:
    event = {"node": "answer_perception", "skills": ["answer_perception"], "latency_ms": latency_ms, "success": False, "error_code": error_code, "action": "stop"}
    with connection() as conn:
        conn.execute(
            "INSERT INTO agent_learning_traces(student_user_id,submission_id,question_id,source_problem_id,mode,qwen_used,teacher_review_needed,handwriting_ocr_failed,latency_ms,agent_trace_id,execution_trace_json,qwen_called,qwen_success,qwen_adopted,question_type,learning_outcome) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (actor_id, submission_id, question_id, source_problem_id, mode, 0, 1, 1, latency_ms, str(uuid.uuid4()), json.dumps([event], ensure_ascii=False), 0, 0, 0, question_type, "not_resolved"),
        )


class MistakeReviewImageIn(MistakeReviewIn):
    image_base64: str = Field(min_length=100, max_length=12_000_000)


class MistakeReviewStepImagesIn(MistakeReviewIn):
    image_base64_list: list[str] = Field(min_length=1, max_length=6)




@app.post("/api/student/released-submissions/{submission_id}/mistakes/{question_id}/review-image")
async def review_released_mistake_image(submission_id: int, question_id: int, req: MistakeReviewImageIn, request: Request):
    """Safe handwritten re-perception for an already released wrong answer only."""
    actor = require_roles(request, {"student"})
    if req.mode not in {"diagnose", "hint", "solution"}:
        raise HTTPException(422, "不支持的复盘模式")
    with connection() as conn:
        row = conn.execute("SELECT j.result_json FROM submissions s JOIN assignments a ON a.id=s.assignment_id JOIN students st ON st.class_id=a.class_id AND st.user_id=? LEFT JOIN grading_jobs j ON j.submission_id=s.id WHERE s.id=? AND s.status=? AND s.released_at IS NOT NULL", (actor["id"], submission_id, "graded")).fetchone()
        if not row: raise HTTPException(404, "未找到已发布的作业结果")
        results = json.loads(row["result_json"] or "{}").get("results") or []
        item = next((x for x in results if int(x.get("question_id") or -1) == question_id), None)
        if not item or (item.get("correct") is True and float(item.get("score") or 0) >= float(item.get("max_score") or 0)):
            raise HTTPException(404, "该题不是可复盘的已发布错题")
        question = conn.execute("SELECT source_problem_id,question_type FROM questions WHERE id=?", (question_id,)).fetchone()
        if not question or not question["source_problem_id"]:
            raise HTTPException(409, "该错题尚未关联到可靠题库原题")
        source_problem_id = question["source_problem_id"]
        teacher_feedback = str(item.get("feedback") or "")
    started = time.perf_counter()
    try:
        async with evidence_client(timeout=620) as client:
            response = await client.post(evidence_url("").rstrip("/") + f"/agent/problems/{source_problem_id}/learn-image", json={"image_base64": req.image_base64, "mode": req.mode, "teacher_feedback": teacher_feedback, "student_steps": req.student_steps})
    except httpx.HTTPError as exc:
        _record_handwriting_failure(actor["id"], submission_id, question_id, source_problem_id, "proof" if "证明" in str(question["question_type"] or "") else "calc", req.mode, round((time.perf_counter() - started) * 1000, 1), "PERCEPTION_UNAVAILABLE")
        raise HTTPException(502, "手写错题复盘 Agent 暂不可用") from exc
    latency_ms = round((time.perf_counter() - started) * 1000, 1)
    if response.status_code >= 400:
        detail = response.json().get("detail", "手写识别失败")
        error_code = "EMPTY_OCR_RESULT" if response.status_code == 422 else "PERCEPTION_UNAVAILABLE"
        _record_handwriting_failure(actor["id"], submission_id, question_id, source_problem_id, "proof" if "证明" in str(question["question_type"] or "") else "calc", req.mode, latency_ms, error_code)
        raise HTTPException(response.status_code, detail)
    result = response.json()
    agent_trace_id = result.pop("trace_id", None)
    execution_trace = result.pop("execution_trace", []) or []
    diagnosis_codes = [str(item.get("code")) for item in (result.get("diagnosis") or {}).get("diagnoses", []) if item.get("code")]
    execution_trace.append({"node": "diagnosis_summary", "skills": ["misconception_diagnosis"], "success": True, "diagnosis_codes": diagnosis_codes})
    qwen_called = any("independent_solving" in event.get("skills", []) for event in execution_trace)
    qwen_success = any("independent_solving" in event.get("skills", []) and event.get("success") for event in execution_trace)
    qwen_adopted = bool((result.get("solution_comparison") or {}).get("consistent"))
    qwen_used = qwen_called
    needs_teacher_review = result.get("action") == "teacher_review"
    question_type = "proof" if "证明" in str(question["question_type"] or "") else "calc"
    failure_codes = [event["error_code"] for event in execution_trace if event.get("error_code")]
    with connection() as conn:
        conn.execute("INSERT INTO agent_learning_traces(student_user_id,submission_id,question_id,source_problem_id,mode,qwen_used,teacher_review_needed,handwriting_ocr_failed,latency_ms,agent_trace_id,execution_trace_json,qwen_called,qwen_success,qwen_adopted,question_type,learning_outcome) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (actor["id"], submission_id, question_id, source_problem_id, req.mode, int(qwen_used), int(needs_teacher_review), 0, latency_ms, agent_trace_id, json.dumps(execution_trace, ensure_ascii=False), int(qwen_called), int(qwen_success), int(qwen_adopted), question_type, "requested_solution" if req.mode == "solution" else None))
    return {**result, "trace": {"trace_id": agent_trace_id, "mode": req.mode, "qwen_used": qwen_used, "needs_teacher_review": needs_teacher_review, "latency_ms": latency_ms, "failure_codes": failure_codes}}


@app.post("/api/student/released-submissions/{submission_id}/mistakes/{question_id}/review-step-images")
async def review_released_mistake_step_images(submission_id: int, question_id: int, req: MistakeReviewStepImagesIn, request: Request):
    """Use ordered handwritten step images only for an owned, released wrong answer."""
    actor = require_roles(request, {"student"})
    if req.mode not in {"diagnose", "hint", "solution"}:
        raise HTTPException(422, "不支持的复盘模式")
    if any(len(image) > 12_000_000 for image in req.image_base64_list) or sum(len(image) for image in req.image_base64_list) > 36_000_000:
        raise HTTPException(413, "步骤图片过大，请压缩后重试")
    with connection() as conn:
        row = conn.execute("SELECT j.result_json FROM submissions s JOIN assignments a ON a.id=s.assignment_id JOIN students st ON st.class_id=a.class_id AND st.user_id=? LEFT JOIN grading_jobs j ON j.submission_id=s.id WHERE s.id=? AND s.status=? AND s.released_at IS NOT NULL", (actor["id"], submission_id, "graded")).fetchone()
        if not row:
            raise HTTPException(404, "未找到已发布的作业结果")
        results = json.loads(row["result_json"] or "{}").get("results") or []
        item = next((x for x in results if int(x.get("question_id") or -1) == question_id), None)
        if not item or (item.get("correct") is True and float(item.get("score") or 0) >= float(item.get("max_score") or 0)):
            raise HTTPException(404, "该题不是可复盘的已发布错题")
        question = conn.execute("SELECT source_problem_id,question_type FROM questions WHERE id=?", (question_id,)).fetchone()
        if not question or not question["source_problem_id"]:
            raise HTTPException(409, "该错题尚未关联到可靠题库原题")
        source_problem_id = question["source_problem_id"]
        teacher_feedback = str(item.get("feedback") or "")
    started = time.perf_counter()
    try:
        async with evidence_client(timeout=620) as client:
            response = await client.post(
                evidence_url("").rstrip("/") + f"/agent/problems/{source_problem_id}/learn-step-images",
                json={"image_base64_list": req.image_base64_list, "mode": req.mode, "teacher_feedback": teacher_feedback, "student_steps": req.student_steps},
            )
    except httpx.HTTPError as exc:
        _record_handwriting_failure(actor["id"], submission_id, question_id, source_problem_id, "proof" if "证明" in str(question["question_type"] or "") else "calc", req.mode, round((time.perf_counter() - started) * 1000, 1), "PERCEPTION_UNAVAILABLE")
        raise HTTPException(502, "分步手写复盘 Agent 暂不可用") from exc
    latency_ms = round((time.perf_counter() - started) * 1000, 1)
    if response.status_code >= 400:
        try:
            detail = response.json().get("detail", "步骤图片识别失败")
        except ValueError:
            detail = "步骤图片识别失败"
        error_code = "EMPTY_OCR_RESULT" if response.status_code == 422 else "PERCEPTION_UNAVAILABLE"
        _record_handwriting_failure(actor["id"], submission_id, question_id, source_problem_id, "proof" if "证明" in str(question["question_type"] or "") else "calc", req.mode, latency_ms, error_code)
        raise HTTPException(response.status_code, detail)
    result = response.json()
    agent_trace_id = result.pop("trace_id", None)
    execution_trace = result.pop("execution_trace", []) or []
    diagnosis_codes = [str(item.get("code")) for item in (result.get("diagnosis") or {}).get("diagnoses", []) if item.get("code")]
    execution_trace.append({"node": "diagnosis_summary", "skills": ["misconception_diagnosis"], "success": True, "diagnosis_codes": diagnosis_codes})
    qwen_called = any("independent_solving" in event.get("skills", []) for event in execution_trace)
    qwen_success = any("independent_solving" in event.get("skills", []) and event.get("success") for event in execution_trace)
    qwen_adopted = bool((result.get("solution_comparison") or {}).get("consistent"))
    needs_teacher_review = result.get("action") == "teacher_review"
    question_type = "proof" if "证明" in str(question["question_type"] or "") else "calc"
    failure_codes = [event["error_code"] for event in execution_trace if event.get("error_code")]
    with connection() as conn:
        conn.execute("INSERT INTO agent_learning_traces(student_user_id,submission_id,question_id,source_problem_id,mode,qwen_used,teacher_review_needed,handwriting_ocr_failed,latency_ms,agent_trace_id,execution_trace_json,qwen_called,qwen_success,qwen_adopted,question_type,learning_outcome) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (actor["id"], submission_id, question_id, source_problem_id, req.mode, int(qwen_called), int(needs_teacher_review), int(not result.get("recognized_step_count")), latency_ms, agent_trace_id, json.dumps(execution_trace, ensure_ascii=False), int(qwen_called), int(qwen_success), int(qwen_adopted), question_type, "requested_solution" if req.mode == "solution" else None))
    return {**result, "trace": {"trace_id": agent_trace_id, "mode": req.mode, "qwen_used": qwen_called, "needs_teacher_review": needs_teacher_review, "latency_ms": latency_ms, "failure_codes": failure_codes}}


class MistakeReattemptIn(BaseModel):
    prior_trace_id: str = Field(min_length=8, max_length=120)
    answer: str = Field(min_length=1, max_length=12000)


@app.post("/api/student/released-submissions/{submission_id}/mistakes/{question_id}/reattempt")
async def verify_mistake_reattempt(submission_id: int, question_id: int, req: MistakeReattemptIn, request: Request):
    """Verify a post-hint reattempt against the protected source answer."""
    actor = require_roles(request, {"student"})
    with connection() as conn:
        row = conn.execute(
            """SELECT j.result_json FROM submissions s JOIN assignments a ON a.id=s.assignment_id
               JOIN students st ON st.class_id=a.class_id AND st.user_id=?
               LEFT JOIN grading_jobs j ON j.submission_id=s.id
               WHERE s.id=? AND s.status=? AND s.released_at IS NOT NULL""",
            (actor["id"], submission_id, "graded"),
        ).fetchone()
        if not row:
            raise HTTPException(404, "未找到已发布的作业结果")
        results = json.loads(row["result_json"] or "{}").get("results") or []
        item = next((x for x in results if int(x.get("question_id") or -1) == question_id), None)
        if not item:
            raise HTTPException(404, "未找到该错题")
        question = conn.execute("SELECT source_problem_id FROM questions WHERE id=?", (question_id,)).fetchone()
        trace = conn.execute(
            "SELECT id FROM agent_learning_traces WHERE agent_trace_id=? AND student_user_id=? AND submission_id=? AND question_id=?",
            (req.prior_trace_id, actor["id"], submission_id, question_id),
        ).fetchone()
        if not question or not question["source_problem_id"] or not trace:
            raise HTTPException(409, "请先从本题的 Agent 反馈进入重做验证")
        source_problem_id = question["source_problem_id"]
    try:
        async with evidence_client(timeout=620) as client:
            response = await client.post(
                evidence_url("").rstrip("/") + f"/agent/problems/{source_problem_id}/learn",
                json={"student_answer": req.answer, "mode": "diagnose", "student_steps": ""},
            )
    except httpx.HTTPError as exc:
        raise HTTPException(502, "重做验证暂不可用") from exc
    if response.status_code >= 400:
        raise HTTPException(response.status_code, "重做验证请求失败")
    result = response.json()
    verification = result.get("verification") or {}
    correct = bool(verification.get("correct")) and float(verification.get("confidence") or 0) >= 0.85
    with connection() as conn:
        conn.execute(
            "INSERT INTO agent_learning_attempts(prior_trace_id,student_user_id,submission_id,question_id,answer_text,verification_json,correct,input_kind) VALUES(?,?,?,?,?,?,?,?)",
            (req.prior_trace_id, actor["id"], submission_id, question_id, req.answer,
             json.dumps(verification, ensure_ascii=False), int(correct), "text"),
        )
        conn.execute(
            "UPDATE agent_learning_traces SET learning_outcome=? WHERE agent_trace_id=? AND student_user_id=?",
            ("corrected" if correct else "reworked", req.prior_trace_id, actor["id"]),
        )
    return {
        "correct": correct,
        "message": "这次重做已验证正确，做得很好。" if correct else "这次结果还不一致；请对照下一步提示检查关键步骤后再试。",
        "verification": {"method": verification.get("method"), "confidence": verification.get("confidence")},
    }


class MistakeReattemptImageIn(BaseModel):
    prior_trace_id: str = Field(min_length=8, max_length=120)
    image_base64: str = Field(min_length=100, max_length=12_000_000)


@app.post("/api/student/released-submissions/{submission_id}/mistakes/{question_id}/reattempt-image")
async def verify_mistake_reattempt_image(submission_id: int, question_id: int, req: MistakeReattemptImageIn, request: Request):
    """Recognize a handwritten reattempt, then verify it with the protected source answer."""
    actor = require_roles(request, {"student"})
    with connection() as conn:
        row = conn.execute(
            """SELECT j.result_json FROM submissions s JOIN assignments a ON a.id=s.assignment_id
               JOIN students st ON st.class_id=a.class_id AND st.user_id=?
               LEFT JOIN grading_jobs j ON j.submission_id=s.id
               WHERE s.id=? AND s.status=? AND s.released_at IS NOT NULL""",
            (actor["id"], submission_id, "graded"),
        ).fetchone()
        if not row:
            raise HTTPException(404, "未找到已发布的作业结果")
        results = json.loads(row["result_json"] or "{}").get("results") or []
        item = next((x for x in results if int(x.get("question_id") or -1) == question_id), None)
        if not item:
            raise HTTPException(404, "未找到该错题")
        question = conn.execute("SELECT source_problem_id FROM questions WHERE id=?", (question_id,)).fetchone()
        trace = conn.execute(
            "SELECT id FROM agent_learning_traces WHERE agent_trace_id=? AND student_user_id=? AND submission_id=? AND question_id=?",
            (req.prior_trace_id, actor["id"], submission_id, question_id),
        ).fetchone()
        if not question or not question["source_problem_id"] or not trace:
            raise HTTPException(409, "请先从本题的 Agent 反馈进入重做验证")
        source_problem_id = question["source_problem_id"]
    try:
        async with evidence_client(timeout=620) as client:
            response = await client.post(
                evidence_url("").rstrip("/") + f"/agent/problems/{source_problem_id}/learn-image",
                json={"image_base64": req.image_base64, "mode": "diagnose", "student_steps": ""},
            )
    except httpx.HTTPError as exc:
        raise HTTPException(502, "手写重做验证暂不可用") from exc
    if response.status_code >= 400:
        detail = response.json().get("detail", "手写图片识别失败")
        raise HTTPException(response.status_code, detail)
    result = response.json()
    verification = result.get("verification") or {}
    perception = result.get("perception") or {}
    recognized_work = str(result.get("recognized_work") or "").strip()
    recognition_confidence = perception.get("confidence")
    correct = bool(verification.get("correct")) and float(verification.get("confidence") or 0) >= 0.85
    with connection() as conn:
        conn.execute(
            "INSERT INTO agent_learning_attempts(prior_trace_id,student_user_id,submission_id,question_id,answer_text,verification_json,correct,input_kind) VALUES(?,?,?,?,?,?,?,?)",
            (req.prior_trace_id, actor["id"], submission_id, question_id, recognized_work,
             json.dumps({"verification": verification, "perception": {"confidence": recognition_confidence}}, ensure_ascii=False),
             int(correct), "image"),
        )
        conn.execute(
            "UPDATE agent_learning_traces SET learning_outcome=? WHERE agent_trace_id=? AND student_user_id=?",
            ("corrected" if correct else "reworked", req.prior_trace_id, actor["id"]),
        )
    if not recognized_work:
        message = "图片中的作答未能可靠识别；请重新拍清楚，或改用文字输入验证。"
    elif correct:
        message = "已识别并验证：这次手写重做正确，做得很好。"
    else:
        message = "已识别，但这次结果还不一致；请对照下一步提示检查关键步骤后再试。"
    return {
        "correct": correct,
        "message": message,
        "recognition_confidence": recognition_confidence,
        "verification": {"method": verification.get("method"), "confidence": verification.get("confidence")},
    }


class LearningOutcomeIn(BaseModel):
    outcome: Literal["reworked", "corrected", "requested_solution", "not_resolved"]


@app.post("/api/student/agent-traces/{trace_id}/outcome")
def record_learning_outcome(trace_id: str, req: LearningOutcomeIn, request: Request):
    actor = require_roles(request, {"student"})
    with connection() as conn:
        updated = conn.execute("UPDATE agent_learning_traces SET learning_outcome=? WHERE agent_trace_id=? AND student_user_id=?", (req.outcome, trace_id, actor["id"])).rowcount
    if not updated:
        raise HTTPException(404, "未找到你的学习记录")
    return {"ok": True, "outcome": req.outcome}


@app.post("/api/learning/problems/{problem_id}/learn")
async def learn_problem(problem_id: str, req: LearningAnswerIn):
    """Proxy the Agent while keeping standard answers server-side."""
    raise HTTPException(410, "预提交通用练习复盘已关闭；请从教师发布的错题详情页进入。")
    # legacy code below is intentionally unreachable
    try:
        async with evidence_client(timeout=620) as client:
            response = await client.post(evidence_url("").rstrip("/") + f"/agent/problems/{problem_id}/learn", json={"student_answer": req.student_answer, "mode": req.mode})
    except httpx.HTTPError as exc:
        raise HTTPException(502, "学习 Agent 暂不可用") from exc
    if response.status_code >= 400:
        try:
            detail = response.json().get("detail", "学习 Agent 请求失败")
        except ValueError:
            detail = "学习 Agent 请求失败"
        raise HTTPException(response.status_code, detail)
    return response.json()


class LearningImageIn(BaseModel):
    image_base64: str = Field(min_length=100, max_length=12_000_000)
    mode: str = "diagnose"


@app.post("/api/learning/problems/{problem_id}/learn-image")
async def learn_problem_image(problem_id: str, req: LearningImageIn):
    raise HTTPException(410, "预提交通用练习复盘已关闭；请从教师发布的错题详情页进入。")
    # legacy code below is intentionally unreachable
    try:
        async with evidence_client(timeout=620) as client:
            response = await client.post(evidence_url("").rstrip("/") + f"/agent/problems/{problem_id}/learn-image", json={"image_base64": req.image_base64, "mode": req.mode})
    except httpx.HTTPError as exc:
        raise HTTPException(502, "手写数学识别暂不可用") from exc
    if response.status_code >= 400:
        try:
            detail = response.json().get("detail", "手写识别失败")
        except ValueError:
            detail = "手写识别失败"
        raise HTTPException(response.status_code, detail)
    return response.json()



@app.get("/learn", response_class=HTMLResponse)
def student_learn_page(request: Request):
    try:
        actor = current_user(request)
    except HTTPException as exc:
        if exc.status_code == 401:
            return RedirectResponse("/login?next=/learn", status_code=303)
        raise
    if actor["role"] != "student":
        return RedirectResponse("/", status_code=303)
    return FileResponse(Path(__file__).with_name("student_learn.html"), headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"})


@app.get("/mistakes", response_class=HTMLResponse)
def student_mistakes_page(request: Request):
    try:
        actor = current_user(request)
    except HTTPException as exc:
        if exc.status_code == 401:
            return RedirectResponse("/login?next=/mistakes", status_code=303)
        raise
    if actor["role"] != "student":
        return RedirectResponse("/", status_code=303)
    return FileResponse(Path(__file__).with_name("student_mistakes.html"), headers={
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache", "Expires": "0",
    })


@app.get("/healthz")
def healthz():
    """Unauthenticated liveness/readiness probe; intentionally exposes no student data."""
    try:
        settings.prepare_dirs()
        with connection() as conn:
            conn.execute("SELECT 1").fetchone()
        queue = queue_health()
        if not queue.get("ok"):
            raise RuntimeError("任务队列不可用")
        return {"status": "ok", "service": "math-agent", "version": app.version, "queue": queue}
    except Exception as exc:
        raise HTTPException(503, "服务暂未就绪") from exc


@app.get("/", response_class=HTMLResponse)
def teacher_agent_portal(request: Request):
    try:
        actor = current_user(request)
    except HTTPException as exc:
        if exc.status_code == 401:
            return RedirectResponse("/login", status_code=303)
        raise
    if actor["role"] == "student":
        return RedirectResponse("/mistakes", status_code=303)
    return FileResponse(Path(__file__).with_name("teacher_portal.html"), headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"})


@app.get("/api-demo", response_class=HTMLResponse)
def dashboard():
    return """<!doctype html><html lang='zh-CN'><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>高数作业助手</title>
    <style>body{margin:0;background:#f5f7fb;font:16px 'Microsoft YaHei',sans-serif;color:#15243a}.hero{padding:56px max(6vw,32px);background:#102a43;color:white}.hero h1{font-size:38px;margin:0 0 12px}.hero p{color:#c7d7e8}.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(210px,1fr));gap:18px;padding:32px max(6vw,32px)}.card{background:#fff;border-radius:14px;padding:22px;box-shadow:0 7px 25px #102a4311}.number{font-size:32px;color:#0b7285;font-weight:700}.hint{color:#62758a}.flow{margin:0 max(6vw,32px) 32px;background:#e6fffb;border-left:4px solid #0b7285;padding:20px;border-radius:8px}a{color:#0b7285}</style>
    <main><section class='hero'><h1>高数作业助手</h1><p>从章节题库到作业提交、批改复核的一体化工作台</p></section><section class='grid' id='metrics'></section><section class='flow'><b>当前可用：</b>示例题库已初始化。教师可通过 <a href='/docs'>接口文档</a> 录题、生成作业、打印作业单，并让学生上传提交件。Dify 与 OCR 通过环境变量接入。</section></main>
    <script>fetch('/api/reports/summary').then(r=>r.json()).then(x=>{let labels={question_count:'已发布题目',assignment_count:'进行中作业',submission_count:'已归档提交',review_queue:'待批改/复核'};metrics.innerHTML=Object.entries(labels).map(([k,v])=>`<article class=card><div class=number>${x[k]}</div><div class=hint>${v}</div></article>`).join('')})</script></html>"""
